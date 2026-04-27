"""Streaming XLSX writer that splits rows across sheets above Excel's
1,048,576-row hard limit (issue #122).

Excel's per-sheet row limit is 2**20 = 1,048,576. Production scans now
routinely exceed that (a 2.5M-row scan hit the limit in #122 and the
``Workbook.save`` call raised ``Row numbers must be between 1 and
1048576``). This helper:

* Uses ``openpyxl`` ``write_only`` mode so memory stays roughly constant
  regardless of row count (peak <500 MB on the 5M-row internal probe).
* Splits the row stream across N data sheets named ``Data_1``,
  ``Data_2``, ... when ``len(rows) > max_rows_per_sheet``. The default
  cap is 1,000,000 (a touch under Excel's hard limit so per-sheet
  totals/formulas still fit a row at the bottom).
* Optionally accepts ``extra_sheets`` so callers (chargeback, forecast)
  can prepend a ``Settings`` sheet whose cells the data sheets reference
  by formula. Per-data-sheet formulas always reference
  ``=Settings!$B$2`` regardless of which ``Data_N`` sheet they land on.
* Optionally writes an ``Index`` sheet listing each data sheet + its row
  count so auditors opening a multi-sheet workbook can navigate.

Returns ``{total_rows, sheet_count, byte_size}`` so callers can surface
"split into N sheets" hints to the frontend (issue #122 frontend toast).
"""

from __future__ import annotations

import io
from typing import Any, BinaryIO, Callable, Dict, Iterable, List, Optional

# Excel's hard cap: 2**20. We default to 1,000,000 (a few thousand under)
# so callers that want to drop a SUM/TOTAL row at the bottom of each
# sheet still have headroom.
EXCEL_MAX_ROWS_PER_SHEET = 1_048_576
DEFAULT_MAX_ROWS_PER_SHEET = 1_000_000


def write_large_workbook(
    rows_iterator: Iterable[Dict[str, Any]],
    columns: List[Dict[str, Any]],
    output: BinaryIO,
    max_rows_per_sheet: int = DEFAULT_MAX_ROWS_PER_SHEET,
    sheet_name_prefix: str = "Data",
    extra_sheets: Optional[Dict[str, List[tuple]]] = None,
    *,
    write_index_sheet: bool = True,
    formula_factory: Optional[Callable[[int, int, str], Optional[str]]] = None,
) -> Dict[str, Any]:
    """Stream ``rows_iterator`` into a multi-sheet XLSX written to ``output``.

    Parameters
    ----------
    rows_iterator
        Any iterable yielding ``dict`` rows. Each dict should carry the
        keys named in ``columns[i]['key']``; missing keys serialize as
        ``""``. Iterators are consumed once — do not pass a generator
        you also intend to iterate elsewhere.
    columns
        Ordered list of column descriptors. Each entry is a dict with:

        * ``key``    — required; row-dict key.
        * ``header`` — required; cell value for the header row.
        * ``width``  — optional; column width hint (ignored in
          write_only mode for body sheets but applied via
          ``column_dimensions`` where supported).
        * ``format`` — optional; reserved for future number-format hints.
    output
        Any binary file-like with ``.write``. Typically ``io.BytesIO``;
        callers wrap that in a ``StreamingResponse``.
    max_rows_per_sheet
        Cap per data sheet. Default 1,000,000 (Excel's hard limit is
        1,048,576; we leave headroom for any per-sheet TOTAL row a
        caller might tack on).
    sheet_name_prefix
        Title prefix for split sheets. With ``"Data"`` and 2.5M rows
        you get ``Data_1``, ``Data_2``, ``Data_3``. Single-sheet
        workbooks (rows fit in one sheet) use ``"<prefix>_1"`` for
        consistency — callers that want plain ``"Data"`` should rename
        post-hoc, but the issue spec says ``Data_1`` so we keep that.
    extra_sheets
        Optional ``{sheet_name: [(row, col, value), ...]}``. ``row`` is
        1-indexed; ``col`` is either a 1-indexed int OR a column letter
        string (``"A"``, ``"B"``, ...). Values may be Python scalars or
        formula strings (``"=B2*C3"``). Used by chargeback to prepend a
        ``Settings`` sheet that data-sheet formulas reference.
    write_index_sheet
        If ``True`` and the row stream produces more than one data
        sheet, an ``Index`` sheet is appended listing each data sheet's
        title + row count. Skipped for single-sheet workbooks (no value).
    formula_factory
        Optional callable ``(row_idx, col_idx, key) -> Optional[str]``
        invoked per cell. If it returns a string starting with ``"="``
        that string is written instead of the row value. ``row_idx`` is
        the 1-indexed row number *within the current data sheet*.
        Lets chargeback land per-row ``=Settings!$B$2 * C{r}`` formulas
        that stay correct after sheet splits.

    Returns
    -------
    dict
        ``{total_rows, sheet_count, byte_size}``.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # pragma: no cover - openpyxl is in requirements.txt
        raise RuntimeError(
            "openpyxl is required for write_large_workbook"
        ) from e

    if max_rows_per_sheet <= 0:
        raise ValueError("max_rows_per_sheet must be positive")
    if max_rows_per_sheet > EXCEL_MAX_ROWS_PER_SHEET:
        # Soft cap to Excel's hard limit; otherwise the openpyxl save would
        # raise the same error issue #122 was filed for.
        max_rows_per_sheet = EXCEL_MAX_ROWS_PER_SHEET

    if not columns:
        raise ValueError("columns must be a non-empty list of {key,header} dicts")

    headers = [c.get("header", c.get("key", "")) for c in columns]
    keys = [c["key"] for c in columns]
    widths = [c.get("width") for c in columns]

    wb = Workbook(write_only=True)

    # ---- extra_sheets first (so data-sheet formulas can reference them) ----
    if extra_sheets:
        # write_only=True forbids random-access cell writes; build these
        # sheets as ordered row streams. We accept (row, col, value) so
        # callers can express sparse layouts; we densely materialise them
        # into rows here.
        for sheet_name, cells in extra_sheets.items():
            ws_extra = wb.create_sheet(title=sheet_name[:31] or "Settings")
            if not cells:
                ws_extra.append([])
                continue
            # Normalise (row, col-letter|int, value) -> (row_int, col_int, value)
            normalised = []
            max_row = 0
            max_col = 0
            for entry in cells:
                if len(entry) != 3:
                    continue
                r, c, v = entry
                r_int = int(r)
                if isinstance(c, str):
                    # Excel column letters -> 1-indexed int.
                    c_int = 0
                    for ch in c.upper():
                        c_int = c_int * 26 + (ord(ch) - ord("A") + 1)
                else:
                    c_int = int(c)
                if r_int < 1 or c_int < 1:
                    continue
                normalised.append((r_int, c_int, v))
                max_row = max(max_row, r_int)
                max_col = max(max_col, c_int)
            # Build a dense 2D buffer then append row-by-row.
            grid: List[List[Any]] = [
                [None] * max_col for _ in range(max_row)
            ]
            for r_int, c_int, v in normalised:
                grid[r_int - 1][c_int - 1] = v
            for row in grid:
                ws_extra.append(row)

    # ---- data sheets --------------------------------------------------------
    total_rows = 0
    data_sheet_titles: List[str] = []
    sheet_row_counts: List[int] = []

    current_ws = None
    current_count = 0
    sheet_idx = 0

    def _open_new_data_sheet():
        nonlocal current_ws, current_count, sheet_idx
        sheet_idx += 1
        title = f"{sheet_name_prefix}_{sheet_idx}"
        # Excel cap on sheet titles is 31 chars.
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        # Column widths via column_dimensions are honoured in write_only.
        for i, w in enumerate(widths, start=1):
            if w:
                try:
                    ws.column_dimensions[get_column_letter(i)].width = float(w)
                except Exception:  # pragma: no cover - defensive
                    pass
        current_ws = ws
        current_count = 0
        data_sheet_titles.append(title)
        sheet_row_counts.append(0)

    # Seed the first sheet eagerly so an empty iterator still produces a
    # valid (header-only) workbook.
    _open_new_data_sheet()

    for row_dict in rows_iterator:
        if current_count >= max_rows_per_sheet:
            _open_new_data_sheet()

        # Header row counts as 1; the data row sits at current_count + 2.
        data_row_idx = current_count + 2

        cells: List[Any] = []
        for col_idx_zero, key in enumerate(keys):
            col_idx = col_idx_zero + 1
            value = row_dict.get(key, "")
            if formula_factory is not None:
                try:
                    formula = formula_factory(data_row_idx, col_idx, key)
                except Exception:
                    formula = None
                if formula is not None:
                    # openpyxl auto-detects formulas via the leading "=".
                    cells.append(formula)
                    continue
            cells.append(value)
        current_ws.append(cells)
        current_count += 1
        total_rows += 1
        sheet_row_counts[-1] = current_count

    # ---- Index sheet (only when split) -------------------------------------
    if write_index_sheet and len(data_sheet_titles) > 1:
        ws_idx = wb.create_sheet(title="Index")
        ws_idx.append(["Sheet", "Row count"])
        for title, count in zip(data_sheet_titles, sheet_row_counts):
            ws_idx.append([title, int(count)])
        ws_idx.append(["TOTAL", int(total_rows)])

    # ---- finalise -----------------------------------------------------------
    # write_only Workbooks must be saved to a path or seekable buffer;
    # ``output`` may be a non-seekable ``StreamingResponse`` body, so we
    # spool to a local BytesIO first then copy.
    spool = io.BytesIO()
    wb.save(spool)
    blob = spool.getvalue()
    output.write(blob)

    return {
        "total_rows": int(total_rows),
        "sheet_count": len(data_sheet_titles),
        "byte_size": len(blob),
        "sheet_titles": list(data_sheet_titles),
    }


def stream_csv(
    rows_iterator: Iterable[Dict[str, Any]],
    columns: List[Dict[str, Any]],
):
    """Generator that yields CSV bytes for ``StreamingResponse``.

    Used by every ``*.xlsx`` endpoint when ``?format=csv`` is requested:
    a single CSV file has no row limit, so a multi-million-row export
    that would split an XLSX into 3 sheets fits in one CSV.
    """
    import csv

    # Use a small in-memory buffer per chunk so we yield bytes back to
    # FastAPI without holding the entire CSV in memory.
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")

    headers = [c.get("header", c.get("key", "")) for c in columns]
    keys = [c["key"] for c in columns]

    writer.writerow(headers)
    yield buf.getvalue().encode("utf-8")
    buf.seek(0)
    buf.truncate(0)

    for row_dict in rows_iterator:
        writer.writerow([row_dict.get(k, "") for k in keys])
        # Flush every row — the row count is what blows up in #122,
        # not the per-row size, so per-row flush keeps memory flat.
        yield buf.getvalue().encode("utf-8")
        buf.seek(0)
        buf.truncate(0)
