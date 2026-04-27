"""Chargeback / cost-center reporting (issue #111, Phase 1).

Maps each ``scanned_files.owner`` value (e.g. ``CONTOSO\\jdoe``) to a
manually-defined cost center and produces:

* per-cost-center totals (gb, file_count, top owners, top dirs)
* a list of unmapped owners so admins can extend the mapping
* an XLSX workbook with FORMULAS so auditors can edit the
  ``cost_per_gb_month`` rate cell and have totals recompute

Phase 1 scope (per issue #111): manual mapping only. AD group
auto-discovery is deferred. Pure Python compute (no pandas);
``fnmatch`` is used so admins can write glob patterns such as
``CONTOSO\\hr_*`` or ``CONTOSO\\*``.

Schema (see ``src/storage/database.py``)::

    cost_centers(id, name, description, cost_per_gb_month, created_at)
    cost_center_owners(cost_center_id, owner_pattern)

Usage::

    from src.reports.chargeback import ChargebackReport
    cb = ChargebackReport(db, config)
    cid = cb.add_center("HR", "Human Resources", 0.05)
    cb.add_owner(cid, "CONTOSO\\hr_*")
    result = cb.compute(scan_id)
    blob = cb.export_xlsx(scan_id)
"""

from __future__ import annotations

import fnmatch
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional

logger = logging.getLogger("file_activity.reports.chargeback")

# Bytes -> GB conversion uses the IEC binary GiB so the sheet's headers
# stay consistent with the rest of the dashboard (which uses 1 GB = 2**30
# bytes via ``size_formatter``).
_BYTES_PER_GB = 1024 ** 3
UNMAPPED_BUCKET = "__unmapped__"


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------


@dataclass
class CostCenterTotals:
    """Per-cost-center aggregation result."""

    id: Optional[int]
    name: str
    description: str = ""
    cost_per_gb_month: float = 0.0
    total_bytes: int = 0
    file_count: int = 0
    top_owners: List[Dict[str, Any]] = field(default_factory=list)
    top_directories: List[Dict[str, Any]] = field(default_factory=list)

    @property
    def total_gb(self) -> float:
        return self.total_bytes / _BYTES_PER_GB

    @property
    def monthly_cost(self) -> float:
        return self.total_gb * float(self.cost_per_gb_month or 0.0)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "name": self.name,
            "description": self.description,
            "cost_per_gb_month": float(self.cost_per_gb_month or 0.0),
            "total_bytes": int(self.total_bytes),
            "total_gb": round(self.total_gb, 4),
            "file_count": int(self.file_count),
            "monthly_cost": round(self.monthly_cost, 4),
            "top_owners": list(self.top_owners),
            "top_directories": list(self.top_directories),
        }


@dataclass
class ChargebackResult:
    """Full chargeback compute result for a single scan."""

    scan_id: int
    centers: List[CostCenterTotals] = field(default_factory=list)
    unmapped_owners: List[Dict[str, Any]] = field(default_factory=list)
    total_bytes: int = 0
    total_file_count: int = 0
    computed_at: str = ""

    @property
    def total_monthly_cost(self) -> float:
        return sum(c.monthly_cost for c in self.centers)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "scan_id": int(self.scan_id),
            "computed_at": self.computed_at or datetime.utcnow().isoformat() + "Z",
            "total_bytes": int(self.total_bytes),
            "total_gb": round(self.total_bytes / _BYTES_PER_GB, 4),
            "total_file_count": int(self.total_file_count),
            "total_monthly_cost": round(self.total_monthly_cost, 4),
            "centers": [c.to_dict() for c in self.centers],
            "unmapped_owners": list(self.unmapped_owners),
        }


# ---------------------------------------------------------------------------
# ChargebackReport
# ---------------------------------------------------------------------------


class ChargebackReport:
    """Cost-center mapping CRUD + per-scan chargeback compute + XLSX export."""

    def __init__(self, db, config: Optional[Dict[str, Any]] = None) -> None:
        self.db = db
        self.config = config or {}

    # -- centers CRUD -----------------------------------------------------

    def list_centers(self) -> List[Dict[str, Any]]:
        with self.db.get_cursor() as cur:
            cur.execute(
                "SELECT id, name, description, cost_per_gb_month, created_at "
                "FROM cost_centers ORDER BY name"
            )
            centers = [dict(r) for r in cur.fetchall()]
            for c in centers:
                cur.execute(
                    "SELECT owner_pattern FROM cost_center_owners "
                    "WHERE cost_center_id = ? ORDER BY owner_pattern",
                    (c["id"],),
                )
                c["owner_patterns"] = [r["owner_pattern"] for r in cur.fetchall()]
        return centers

    def get_center(self, center_id: int) -> Optional[Dict[str, Any]]:
        with self.db.get_cursor() as cur:
            cur.execute(
                "SELECT id, name, description, cost_per_gb_month, created_at "
                "FROM cost_centers WHERE id = ?",
                (center_id,),
            )
            row = cur.fetchone()
            if not row:
                return None
            d = dict(row)
            cur.execute(
                "SELECT owner_pattern FROM cost_center_owners "
                "WHERE cost_center_id = ? ORDER BY owner_pattern",
                (center_id,),
            )
            d["owner_patterns"] = [r["owner_pattern"] for r in cur.fetchall()]
            return d

    def add_center(
        self,
        name: str,
        description: str = "",
        cost_per_gb_month: float = 0.0,
    ) -> int:
        if not name or not str(name).strip():
            raise ValueError("name gerekli")
        try:
            rate = float(cost_per_gb_month or 0.0)
        except (TypeError, ValueError) as e:
            raise ValueError(f"cost_per_gb_month sayisal olmali: {e}")
        if rate < 0:
            raise ValueError("cost_per_gb_month negatif olamaz")
        with self.db.get_cursor() as cur:
            cur.execute(
                "INSERT INTO cost_centers (name, description, cost_per_gb_month) "
                "VALUES (?, ?, ?)",
                (str(name).strip(), description or "", rate),
            )
            return int(cur.lastrowid)

    def update_center(self, center_id: int, **fields: Any) -> bool:
        """Idempotent partial update. Unknown fields are silently ignored."""
        allowed = {"name", "description", "cost_per_gb_month"}
        sets: List[str] = []
        vals: List[Any] = []
        for k, v in fields.items():
            if k not in allowed:
                continue
            if k == "cost_per_gb_month":
                try:
                    v = float(v or 0.0)
                except (TypeError, ValueError) as e:
                    raise ValueError(f"cost_per_gb_month sayisal olmali: {e}")
                if v < 0:
                    raise ValueError("cost_per_gb_month negatif olamaz")
            if k == "name":
                if not v or not str(v).strip():
                    raise ValueError("name bos olamaz")
                v = str(v).strip()
            sets.append(f"{k} = ?")
            vals.append(v)
        if not sets:
            return False
        vals.append(center_id)
        with self.db.get_cursor() as cur:
            cur.execute(
                f"UPDATE cost_centers SET {', '.join(sets)} WHERE id = ?",
                vals,
            )
            return cur.rowcount > 0

    def remove_center(self, center_id: int) -> bool:
        """Idempotent delete. Cascades owner patterns via FK."""
        with self.db.get_cursor() as cur:
            # SQLite FK cascade only fires when PRAGMA foreign_keys=ON; the
            # Database class enables it on connect, but we also explicitly
            # delete the owner rows so the operation is robust either way.
            cur.execute(
                "DELETE FROM cost_center_owners WHERE cost_center_id = ?",
                (center_id,),
            )
            cur.execute("DELETE FROM cost_centers WHERE id = ?", (center_id,))
            return cur.rowcount > 0

    # -- owner patterns ---------------------------------------------------

    def add_owner(self, center_id: int, owner_pattern: str) -> bool:
        """Attach an owner pattern to a center. Idempotent — re-adding the
        same pattern returns ``False`` (no-op) instead of raising."""
        if not owner_pattern or not str(owner_pattern).strip():
            raise ValueError("owner_pattern gerekli")
        pat = str(owner_pattern).strip()
        with self.db.get_cursor() as cur:
            cur.execute(
                "SELECT 1 FROM cost_centers WHERE id = ?",
                (center_id,),
            )
            if not cur.fetchone():
                raise ValueError(f"cost_center bulunamadi: {center_id}")
            try:
                cur.execute(
                    "INSERT INTO cost_center_owners (cost_center_id, owner_pattern) "
                    "VALUES (?, ?)",
                    (center_id, pat),
                )
                return True
            except Exception:
                # Re-add of an existing (cost_center_id, owner_pattern) pair
                # — composite PK already enforces uniqueness. Treat as no-op.
                return False

    def remove_owner(self, center_id: int, owner_pattern: str) -> bool:
        with self.db.get_cursor() as cur:
            cur.execute(
                "DELETE FROM cost_center_owners "
                "WHERE cost_center_id = ? AND owner_pattern = ?",
                (center_id, owner_pattern),
            )
            return cur.rowcount > 0

    # -- compute ----------------------------------------------------------

    def compute(self, scan_id: int) -> ChargebackResult:
        """Aggregate a scan's files by cost center.

        Files whose owner does not match any pattern fall into the
        ``unmapped_owners`` bucket so the admin can extend the mapping.
        """
        centers_meta = self.list_centers()

        # Build (pattern, center_idx) lookup. Order is stable by name so the
        # first matching pattern wins for files whose owner happens to be
        # eligible for two centers (deterministic, easy to explain).
        pattern_table: List[tuple] = []
        for idx, c in enumerate(centers_meta):
            for pat in c.get("owner_patterns") or []:
                pattern_table.append((pat, idx))

        center_totals: List[CostCenterTotals] = [
            CostCenterTotals(
                id=c["id"],
                name=c["name"],
                description=c.get("description") or "",
                cost_per_gb_month=float(c.get("cost_per_gb_month") or 0.0),
            )
            for c in centers_meta
        ]
        # Per-center owner / dir tallies (computed first, then top-N at the end).
        center_owner_bytes: List[Dict[str, int]] = [dict() for _ in centers_meta]
        center_owner_count: List[Dict[str, int]] = [dict() for _ in centers_meta]
        center_dir_bytes: List[Dict[str, int]] = [dict() for _ in centers_meta]
        center_dir_count: List[Dict[str, int]] = [dict() for _ in centers_meta]

        unmapped_owner_bytes: Dict[str, int] = {}
        unmapped_owner_count: Dict[str, int] = {}

        total_bytes = 0
        total_files = 0

        with self.db.get_cursor() as cur:
            cur.execute(
                "SELECT file_path, owner, file_size FROM scanned_files "
                "WHERE scan_id = ?",
                (int(scan_id),),
            )
            for row in cur:
                size = int(row["file_size"] or 0)
                owner = row["owner"] or ""
                path = row["file_path"] or ""
                idx = self._match_owner_idx(owner, pattern_table)

                total_bytes += size
                total_files += 1

                # Directory key: drop the basename, normalise separators so
                # both ``\`` (UNC) and ``/`` (POSIX) collapse to ``/``.
                dir_key = self._directory_of(path)

                if idx is None:
                    bucket = owner or "(no owner)"
                    unmapped_owner_bytes[bucket] = unmapped_owner_bytes.get(bucket, 0) + size
                    unmapped_owner_count[bucket] = unmapped_owner_count.get(bucket, 0) + 1
                    continue

                center_totals[idx].total_bytes += size
                center_totals[idx].file_count += 1
                if owner:
                    center_owner_bytes[idx][owner] = center_owner_bytes[idx].get(owner, 0) + size
                    center_owner_count[idx][owner] = center_owner_count[idx].get(owner, 0) + 1
                if dir_key:
                    center_dir_bytes[idx][dir_key] = center_dir_bytes[idx].get(dir_key, 0) + size
                    center_dir_count[idx][dir_key] = center_dir_count[idx].get(dir_key, 0) + 1

        # Materialise top-10 owners / directories per center.
        for idx, ct in enumerate(center_totals):
            ct.top_owners = self._top_n(
                center_owner_bytes[idx], center_owner_count[idx], n=10, key="owner"
            )
            ct.top_directories = self._top_n(
                center_dir_bytes[idx], center_dir_count[idx], n=10, key="directory"
            )

        unmapped_list = self._top_n(
            unmapped_owner_bytes, unmapped_owner_count, n=None, key="owner"
        )

        result = ChargebackResult(
            scan_id=int(scan_id),
            centers=center_totals,
            unmapped_owners=unmapped_list,
            total_bytes=total_bytes,
            total_file_count=total_files,
            computed_at=datetime.utcnow().isoformat() + "Z",
        )
        return result

    # -- xlsx export ------------------------------------------------------

    def export_xlsx(self, scan_id: int) -> bytes:
        """Build an XLSX workbook with FORMULAS so auditors can edit rates.

        Sheets:
        * ``Summary`` — one row per cost center; ``monthly_cost`` is a
          live formula referencing the corresponding ``cost_per_gb_month``
          cell on the same row, so editing the rate updates the total.
        * ``Detail`` — flat rows of (cost_center, owner, total_gb,
          file_count, monthly_cost-formula). Used for owner-level audits.
        * ``Settings`` — global default rate (cell ``B2``); the
          ``Summary`` sheet's per-row rates use ``=Settings!$B$2`` if the
          per-center rate is missing/zero, so a single cell can drive the
          whole report.

        Returns the workbook bytes; callers wrap it in a
        ``StreamingResponse``.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as e:  # pragma: no cover - openpyxl is in requirements
            raise RuntimeError(
                "openpyxl is required for chargeback xlsx export"
            ) from e

        result = self.compute(scan_id)

        wb = Workbook()

        # ---- Settings sheet (built FIRST so Summary formulas can ref it) ----
        ws_set = wb.active
        ws_set.title = "Settings"
        ws_set["A1"] = "Setting"
        ws_set["B1"] = "Value"
        for c in (ws_set["A1"], ws_set["B1"]):
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")

        # B2 holds the GLOBAL default rate. Auditors may edit this cell.
        # Per-center rates on the Summary sheet override it via their own
        # cell, but Detail-sheet formulas reference Settings!$B$2 so a
        # single-cell change propagates.
        default_rate = 0.0
        if result.centers:
            non_zero = [c.cost_per_gb_month for c in result.centers if c.cost_per_gb_month]
            if non_zero:
                default_rate = float(non_zero[0])
        ws_set["A2"] = "Default cost_per_gb_month"
        ws_set["B2"] = default_rate
        ws_set["A3"] = "Bytes per GB"
        ws_set["B3"] = _BYTES_PER_GB
        ws_set["A4"] = "Computed at (UTC)"
        ws_set["B4"] = result.computed_at
        ws_set["A5"] = "Scan id"
        ws_set["B5"] = result.scan_id
        ws_set["A6"] = "Note"
        ws_set["B6"] = (
            "Edit B2 to change the global rate. Per-center rates on the "
            "Summary sheet override this default for that center only."
        )
        ws_set.column_dimensions["A"].width = 32
        ws_set.column_dimensions["B"].width = 60

        # ---- Summary sheet --------------------------------------------------
        # Columns: A name | B description | C cost_per_gb_month | D total_gb |
        #          E file_count | F monthly_cost (FORMULA = C*D) |
        #          G top_owner
        ws_sum = wb.create_sheet("Summary")
        headers = [
            "cost_center",
            "description",
            "cost_per_gb_month",
            "total_gb",
            "file_count",
            "monthly_cost",
            "top_owner",
        ]
        ws_sum.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws_sum.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        for i, ct in enumerate(result.centers, start=2):
            top_owner = ct.top_owners[0]["owner"] if ct.top_owners else ""
            ws_sum.cell(row=i, column=1, value=ct.name)
            ws_sum.cell(row=i, column=2, value=ct.description or "")
            # If the per-center rate is positive use it; otherwise fall back
            # to the Settings sheet's B2 via a formula so editing the global
            # rate updates this row too.
            if ct.cost_per_gb_month and ct.cost_per_gb_month > 0:
                ws_sum.cell(row=i, column=3, value=float(ct.cost_per_gb_month))
            else:
                ws_sum.cell(row=i, column=3, value="=Settings!$B$2")
            ws_sum.cell(row=i, column=4, value=round(ct.total_gb, 6))
            ws_sum.cell(row=i, column=5, value=int(ct.file_count))
            # CRITICAL: monthly_cost is a FORMULA, not a precomputed value.
            ws_sum.cell(row=i, column=6, value=f"=C{i}*D{i}")
            ws_sum.cell(row=i, column=7, value=top_owner)

        # Totals row at the end.
        if result.centers:
            total_row = len(result.centers) + 2
            ws_sum.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            ws_sum.cell(row=total_row, column=4,
                        value=f"=SUM(D2:D{total_row - 1})").font = Font(bold=True)
            ws_sum.cell(row=total_row, column=5,
                        value=f"=SUM(E2:E{total_row - 1})").font = Font(bold=True)
            ws_sum.cell(row=total_row, column=6,
                        value=f"=SUM(F2:F{total_row - 1})").font = Font(bold=True)

        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 40
        for col in ("C", "D", "E", "F"):
            ws_sum.column_dimensions[col].width = 18
        ws_sum.column_dimensions["G"].width = 32

        # ---- Detail sheet ---------------------------------------------------
        ws_det = wb.create_sheet("Detail")
        det_headers = [
            "cost_center",
            "owner",
            "total_gb",
            "file_count",
            "monthly_cost",
        ]
        ws_det.append(det_headers)
        for col_idx, _ in enumerate(det_headers, start=1):
            cell = ws_det.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        det_row = 2
        for ct in result.centers:
            for owner_row in ct.top_owners:
                gb = float(owner_row.get("total_bytes") or 0) / _BYTES_PER_GB
                ws_det.cell(row=det_row, column=1, value=ct.name)
                ws_det.cell(row=det_row, column=2, value=owner_row.get("owner") or "")
                ws_det.cell(row=det_row, column=3, value=round(gb, 6))
                ws_det.cell(row=det_row, column=4, value=int(owner_row.get("file_count") or 0))
                # Detail rows reference the GLOBAL Settings!$B$2 rate so a
                # single edit on Settings drives every owner total.
                ws_det.cell(
                    row=det_row, column=5, value=f"=Settings!$B$2*C{det_row}"
                )
                det_row += 1

        # Unmapped owners as their own block on Detail (so they show up in
        # the same audit document as everything else).
        if result.unmapped_owners:
            ws_det.cell(row=det_row, column=1, value="(unmapped)").font = Font(
                italic=True, color="888888"
            )
            det_row += 1
            for u in result.unmapped_owners:
                gb = float(u.get("total_bytes") or 0) / _BYTES_PER_GB
                ws_det.cell(row=det_row, column=1, value=UNMAPPED_BUCKET)
                ws_det.cell(row=det_row, column=2, value=u.get("owner") or "")
                ws_det.cell(row=det_row, column=3, value=round(gb, 6))
                ws_det.cell(row=det_row, column=4, value=int(u.get("file_count") or 0))
                ws_det.cell(
                    row=det_row, column=5, value=f"=Settings!$B$2*C{det_row}"
                )
                det_row += 1

        ws_det.column_dimensions["A"].width = 24
        ws_det.column_dimensions["B"].width = 32
        for col in ("C", "D", "E"):
            ws_det.column_dimensions[col].width = 16

        # Move Settings sheet to last so users land on Summary first.
        wb.move_sheet("Settings", offset=2)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # -- helpers ----------------------------------------------------------

    @staticmethod
    def _match_owner_idx(
        owner: str, pattern_table: Iterable[tuple]
    ) -> Optional[int]:
        """Return the index of the first cost center whose pattern matches
        ``owner`` (exact OR fnmatch glob)."""
        if not owner:
            return None
        for pat, idx in pattern_table:
            # Exact match wins outright; fnmatch covers globs.
            if pat == owner or fnmatch.fnmatchcase(owner, pat):
                return idx
        return None

    @staticmethod
    def _directory_of(path: str) -> str:
        if not path:
            return ""
        normalised = path.replace("\\", "/")
        if "/" not in normalised:
            return ""
        return normalised.rsplit("/", 1)[0]

    @staticmethod
    def _top_n(
        bytes_map: Dict[str, int],
        count_map: Dict[str, int],
        n: Optional[int],
        key: str,
    ) -> List[Dict[str, Any]]:
        items = [
            {
                key: k,
                "total_bytes": int(v),
                "total_gb": round(int(v) / _BYTES_PER_GB, 4),
                "file_count": int(count_map.get(k, 0)),
            }
            for k, v in bytes_map.items()
        ]
        items.sort(key=lambda r: r["total_bytes"], reverse=True)
        if n is not None:
            items = items[:n]
        return items
