"""FastAPI Web Dashboard.

Tum raporlama, kaynak yonetimi, arsiv ve zamanlama islemlerini
web arayuzu uzerinden sunar.

ONEMLI: Tum API endpoint'leri source_id (integer) kullanir,
source_name veya UNC path URL'de KULLANILMAZ (encoding sorunlari).
"""

import contextlib
import os
import json
import logging
import threading
import uuid
from datetime import datetime, timedelta

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from pydantic import BaseModel, Field, field_validator
from typing import Optional, List

# ── Arka plan export kuyrugu ──
_export_jobs = {}  # job_id -> {status, progress, file_path, error, created_at, ...}
_export_lock = threading.Lock()

logger = logging.getLogger("file_activity.dashboard")


# --- Issue #82 (Bug 1): /api/system/open-folder dual-behaviour helper ---
#
# The dashboard is typically served from a Windows file server while users
# browse it from their own workstations. Calling `subprocess.Popen(["explorer",
# ...])` server-side opens a window on the *server*, which is invisible to the
# remote user and makes the "Konuma Git" buttons look broken.
#
# This helper implements two modes:
#   * Local client (127.0.0.1 / ::1 / localhost): spawn Explorer natively and
#     return {"success": True, "mode": "native"}.
#   * Remote client: do NOT touch subprocess. Return HTTP 200 with
#     {"success": False, "mode": "remote_client", ...} so the frontend can
#     copy the path to the user's clipboard and surface a friendly hint.
#
# Path resolution (`os.path.realpath(os.path.normpath(...))`) and the
# `shell=False` argv-list Popen form are preserved for security. Missing paths
# still produce HTTP 404 via HTTPException.

_LOCAL_CLIENT_HOSTS = frozenset({"127.0.0.1", "::1", "localhost"})


# Issue #135 — coarse % estimate per scan phase. The MFT walk has no notion
# of "how much is left", so we fake a curve that conveys "we're moving"
# without misleading the user. Frontend uses this only as a visual hint;
# absolute KPI numbers (file_count, total_size) stay the source of truth.
def _phase_progress_pct(phase: str, file_count: int) -> int:
    """Map (phase, file_count) -> 0..100. Cheap, deterministic, no I/O."""
    if not phase:
        return 0
    if phase == "completed":
        return 100
    if phase in ("failed", "cancelled"):
        return 0
    if phase == "enumeration":
        # MFT walk: linear ramp up to 30% over the first 1M records, then
        # asymptote (we don't know total — pretend we're 30% of the way).
        return min(30, int(file_count / 33333))
    if phase == "insert":
        # Insert: 30..85%. Stretch over 5M records.
        return min(85, 30 + int(file_count / 100000))
    if phase == "analysis":
        return 95
    return 0


def open_folder_impl(body: dict, client_host: str, popen=None):
    """Run the open-folder decision logic.

    Pure-ish helper shared by the HTTP endpoint and the unit tests. Returns
    the JSON-serialisable response dict on success, or raises HTTPException
    for invalid input / missing paths. ``popen`` lets tests inject a stub in
    place of ``subprocess.Popen``.
    """
    if popen is None:
        import subprocess
        popen = subprocess.Popen

    folder = body.get("path", "") if isinstance(body, dict) else ""
    if not folder or not isinstance(folder, str):
        raise HTTPException(400, "path gerekli")

    # Guvenlik: normalize + gercek yol cozumleme (symlink/junction eskape koruma)
    folder = os.path.realpath(os.path.normpath(folder))
    if not (os.path.isdir(folder) or os.path.isfile(folder)):
        raise HTTPException(404, f"Dizin bulunamadi: {folder}")

    is_local = client_host in _LOCAL_CLIENT_HOSTS
    if not is_local:
        return {
            "success": False,
            "mode": "remote_client",
            "path": folder,
            "hint": (
                "Explorer cannot be opened on the server for a remote "
                "client. Use the copied path on your own machine."
            ),
        }

    # Yerel istemci: dosya/dizine gore Explorer'i acar.
    # shell=False ile argv listesi kullanilarak komut enjeksiyonu onlenir.
    if os.path.isdir(folder):
        popen(["explorer", folder], shell=False)
    else:  # os.path.isfile(folder)
        popen(["explorer", "/select,", folder], shell=False)
    return {"success": True, "mode": "native", "path": folder}


# --- Issue #82 (Bug 4) / Issue #105: /api/system/list-dir folder browser ---
#
# Powers the "loadFolderBrowser" modal in the dashboard so admins can pick a
# source / archive destination by clicking through the server's filesystem
# instead of typing UNC paths from memory. Localhost-only (mirrors the
# /api/system/open-folder pattern from PR #85): a remote client must NEVER be
# able to enumerate the file server's contents, so we return HTTP 403 in
# that case rather than leaking the directory structure.
#
# Performance / safety:
#   * scandir for one-syscall-per-entry traversal,
#   * 5000-entry cap (LIST_DIR_MAX_ENTRIES) to bound response size,
#   * hidden / system files filtered unless ?show_hidden=true,
#   * sort: directories first (alphabetical), then files (alphabetical),
#   * realpath + normpath path resolution to defeat symlink/junction tricks,
#   * empty `path` returns the logical roots (drives on Windows, "/" on Unix)
#     so the UI can boot the browser without knowing the platform.

LIST_DIR_MAX_ENTRIES = 5000


def _list_dir_logical_roots() -> list[dict]:
    """Return the platform-appropriate logical roots for the browser.

    On Windows we enumerate fixed drive letters that actually have a
    filesystem mounted (so we don't include phantom A:/B: floppy entries).
    On POSIX we return a single "/" root, which is enough for admins to
    navigate to /mnt, /media, /srv, etc.
    """
    import string

    roots: list[dict] = []
    if os.name == "nt":
        for letter in string.ascii_uppercase:
            drive = f"{letter}:\\"
            if os.path.isdir(drive):
                roots.append({
                    "name": drive,
                    "type": "dir",
                    "size": None,
                    "mtime": None,
                })
    else:
        roots.append({
            "name": "/",
            "type": "dir",
            "size": None,
            "mtime": None,
        })
    return roots


def _list_dir_is_hidden(name: str, full_path: str) -> bool:
    """Best-effort hidden / system flag detection.

    POSIX: dotfile convention.
    Windows: FILE_ATTRIBUTE_HIDDEN / FILE_ATTRIBUTE_SYSTEM via GetFileAttributesW
    when available; we do not raise on failure — a stat error simply means
    the entry is included (better to show it than hide it silently).
    """
    if name.startswith("."):
        return True
    if os.name == "nt":
        try:
            import ctypes

            attrs = ctypes.windll.kernel32.GetFileAttributesW(full_path)
            if attrs != -1:
                # FILE_ATTRIBUTE_HIDDEN = 0x2, FILE_ATTRIBUTE_SYSTEM = 0x4
                if attrs & 0x2 or attrs & 0x4:
                    return True
        except Exception:  # pragma: no cover - kernel32 absent / sandbox
            pass
    return False


def list_dir_impl(
    path: str,
    client_host: str,
    show_hidden: bool = False,
    max_entries: int = LIST_DIR_MAX_ENTRIES,
):
    """Pure helper backing the /api/system/list-dir endpoint.

    Returns ``{path, parent, entries}``; raises HTTPException(403) for a
    remote client and HTTPException(404) for a path that doesn't exist
    or isn't a directory.
    """
    if client_host not in _LOCAL_CLIENT_HOSTS:
        raise HTTPException(
            403,
            "Klasor tarayici sadece sunucudan calisirken kullanilabilir. "
            "Yolu manuel girin.",
        )

    # Empty path -> logical roots (drives on Windows, "/" on POSIX). We
    # surface ``parent=None`` so the UI hides the "up one level" affordance.
    if not path:
        return {
            "path": "",
            "parent": None,
            "entries": _list_dir_logical_roots(),
        }

    # Path resolution: normpath to collapse ``..``, then realpath to defeat
    # symlink/junction escape attempts. We don't bind to a specific allow-
    # list of roots because the operator legitimately needs to pick any
    # local or mounted UNC path; the localhost-only check above is what
    # keeps remote clients from walking the filesystem.
    resolved = os.path.realpath(os.path.normpath(path))
    if not os.path.isdir(resolved):
        raise HTTPException(404, f"Yol bulunamadi: {resolved}")

    entries: list[dict] = []
    truncated = False
    try:
        with os.scandir(resolved) as it:
            for entry in it:
                if not show_hidden and _list_dir_is_hidden(
                    entry.name, entry.path
                ):
                    continue
                try:
                    is_dir = entry.is_dir(follow_symlinks=False)
                except OSError:
                    is_dir = False
                size: int | None = None
                mtime: float | None = None
                if not is_dir:
                    try:
                        st = entry.stat(follow_symlinks=False)
                        size = st.st_size
                        mtime = st.st_mtime
                    except OSError:
                        size = None
                        mtime = None
                else:
                    try:
                        st = entry.stat(follow_symlinks=False)
                        mtime = st.st_mtime
                    except OSError:
                        mtime = None
                entries.append({
                    "name": entry.name,
                    "type": "dir" if is_dir else "file",
                    "size": size,
                    "mtime": mtime,
                })
                if len(entries) >= max_entries:
                    truncated = True
                    break
    except PermissionError:
        raise HTTPException(403, f"Erisim engellendi: {resolved}")
    except OSError as e:  # pragma: no cover - rare FS errors
        raise HTTPException(500, f"Dizin okunamadi: {e}")

    # Sort: dirs first (alphabetical, case-insensitive), then files.
    entries.sort(key=lambda e: (e["type"] != "dir", e["name"].lower()))

    # Compute parent. For a drive root ("C:\\") os.path.dirname returns
    # itself, so we surface ``parent=""`` to mean "go back to the logical
    # roots list". Same for POSIX "/".
    parent_raw = os.path.dirname(resolved.rstrip(os.sep) or resolved)
    if parent_raw == resolved or not parent_raw:
        parent: str | None = ""
    else:
        parent = parent_raw

    out = {
        "path": resolved,
        "parent": parent,
        "entries": entries,
    }
    if truncated:
        out["truncated"] = True
        out["max_entries"] = max_entries
    return out


# --- Pydantic Models ---

class SourceCreate(BaseModel):
    name: str
    unc_path: str
    archive_dest: Optional[str] = None

class PolicyCreate(BaseModel):
    name: str
    source_id: Optional[int] = None
    access_days: Optional[int] = None
    modify_days: Optional[int] = None
    min_size: Optional[int] = None
    max_size: Optional[int] = None
    extensions: Optional[List[str]] = None
    exclude_extensions: Optional[List[str]] = None

class ScheduleCreate(BaseModel):
    task_type: str
    source_id: int
    policy_name: Optional[str] = None
    cron_expression: str

    @field_validator("task_type")
    @classmethod
    def _validate_task_type(cls, v: str) -> str:
        # Issue #38: 'audit_export' added for scheduled WORM export.
        if v not in ("scan", "archive", "notify_users", "audit_export"):
            raise ValueError(
                "task_type must be 'scan', 'archive', 'notify_users' or 'audit_export'"
            )
        return v

    @field_validator("cron_expression")
    @classmethod
    def _validate_cron(cls, v: str) -> str:
        parts = v.strip().split()
        if len(parts) != 5:
            raise ValueError("cron_expression must have 5 fields (minute hour day month dow)")
        try:
            from apscheduler.triggers.cron import CronTrigger
            CronTrigger(
                minute=parts[0], hour=parts[1],
                day=parts[2], month=parts[3],
                day_of_week=parts[4],
            )
        except Exception as e:
            raise ValueError(f"invalid cron expression: {e}")
        return v.strip()

class ArchiveRequest(BaseModel):
    source_id: int
    policy_name: Optional[str] = None
    days: Optional[int] = None
    # Issue #158 (C-2) — explicit safety gate. ``dry_run`` defaults to
    # the value in ``archiving.dry_run`` (now true by default); the API
    # caller can override it to false ONLY when ``confirm=true`` is
    # also supplied. The endpoint refuses ambiguous requests with HTTP
    # 400 so a misclick / replay can't move files.
    dry_run: Optional[bool] = None
    confirm: Optional[bool] = False

class RestoreRequest(BaseModel):
    archive_id: Optional[int] = None
    original_path: Optional[str] = None


def _get_source(db, source_id: int):
    """Source ID'den kaynak bul, yoksa 404."""
    src = db.get_source_by_id(source_id)
    if not src:
        raise HTTPException(404, f"Kaynak bulunamadi (ID: {source_id})")
    return src


def _attach_cache_envelope(envelope: dict) -> dict:
    """Merge a ``{"results": ..., "cache": ...}`` envelope back into the
    flat report shape the frontend expects.

    Issue #123 spec: existing endpoints' response shape stays unchanged;
    the ``cache`` field is purely additive (frontend ignores it, ops
    debugging uses it). Implementation detail: the cache layer wraps the
    underlying ``ReportGenerator`` output in ``"results"`` so it can be
    treated as an opaque blob; we unwrap here and tag on the envelope.
    """
    results = envelope.get("results", {})
    if not isinstance(results, dict):
        return {"results": results, "cache": envelope.get("cache", {})}
    out = dict(results)
    out["cache"] = envelope.get("cache", {})
    return out


def _read_version() -> str:
    """Proje kok dizinindeki VERSION dosyasindan sürümü oku.

    Sürüm tek bir dosyada yasar (repo kok dizininde `VERSION`). Dashboard
    ve FastAPI title'i bu tek kaynaktan beslenir. Dosya yoksa 'unknown'
    doner — hardcoded bir yedek yok ki yanlis bir deger yayilmasin.

    Eger `.git/HEAD` bulunursa commit SHA'nin ilk 7 karakteri ekte
    gorunur (ornek: `1.8.0-dev+a1b2c3d`). Boylece VERSION dosyasi
    ayni kalsa bile kullanici gercek bir update olup olmadigini
    gorsel olarak ayirt edebilir. Setup-source.ps1 ZIP ile kurulum
    yaptiginda .git yoktur, o durumda sadece VERSION stringi doner.
    """
    base = "unknown"
    for candidate in (
        os.path.join(os.path.dirname(__file__), "..", "..", "VERSION"),
        os.path.join(os.getcwd(), "VERSION"),
    ):
        try:
            with open(candidate, "r", encoding="utf-8") as f:
                v = f.read().strip()
                if v:
                    base = v
                    break
        except (OSError, IOError):
            continue
    # Commit SHA eklemesi — .git varsa + ZIP'le gelen kurulumda olmayacak,
    # bu yuzden hata sessizce yutulur. Zip kurulumlarinda setup-source.ps1
    # indirme sirasinda "COMMIT_SHA" dosyasi yazabilir — ayni dizinde
    # bunun da varligini kontrol edelim.
    for candidate in (
        os.path.join(os.path.dirname(__file__), "..", "..", "COMMIT_SHA"),
        os.path.join(os.getcwd(), "COMMIT_SHA"),
    ):
        try:
            with open(candidate, "r", encoding="utf-8") as f:
                sha = f.read().strip()[:7]
                if sha and sha.isalnum():
                    return f"{base}+{sha}"
        except (OSError, IOError):
            continue
    try:
        head_candidates = (
            os.path.join(os.path.dirname(__file__), "..", "..", ".git", "HEAD"),
            os.path.join(os.getcwd(), ".git", "HEAD"),
        )
        for git_head in head_candidates:
            if not os.path.exists(git_head):
                continue
            with open(git_head, "r", encoding="utf-8") as f:
                head = f.read().strip()
            git_dir = os.path.dirname(git_head)
            if head.startswith("ref: "):
                ref_path = os.path.join(git_dir, head[5:])
                if os.path.exists(ref_path):
                    with open(ref_path, "r", encoding="utf-8") as f:
                        sha = f.read().strip()[:7]
                        if sha:
                            return f"{base}+{sha}"
            elif len(head) >= 7:
                return f"{base}+{head[:7]}"
    except (OSError, IOError):
        pass
    return base


APP_VERSION = _read_version()


def create_app(db, config, analytics=None, ad_lookup=None, email_notifier=None,
               operations_registry=None):
    """FastAPI uygulamasini olustur.

    analytics: Opsiyonel AnalyticsEngine. Verilmezse config.analytics'e gore
    olusturulur; DuckDB yoksa veya basarisiz olursa `available=False` ile
    doner ve endpoint'ler SQLite fallback'ine duser.

    ad_lookup: Opsiyonel ADLookup. Verilmezse config.active_directory'den
    olusturulur; ldap3 yoksa veya enabled=false ise available=False ile
    doner ve endpoint'ler cache degeri / None doner.

    email_notifier: Opsiyonel EmailNotifier. Verilmezse config.smtp'den
    olusturulur; smtp.enabled=false ise available=False ile doner ve
    e-posta endpoint'leri {"skipped": true} doner.
    """
    # ─────────────────────────────────────────────────────────────────
    # Issue #77 Phase 2 — auto-restore on SQLite corruption.
    #
    # Run BEFORE any heavy DB init (analytics ATTACH, scheduler boot,
    # backfill jobs). The probe is read-only via a transient
    # connection; if corruption is detected AND
    # ``backup.auto_restore_on_corruption`` is true we forensic-rename
    # the broken file and copy the latest snapshot in its place.
    #
    # main.py already called ``db.connect()`` before handing us the
    # Database, so we must close it first — the restore step needs the
    # WAL file unowned. After a successful restore we re-call connect()
    # so downstream init talks to the salvaged DB.
    # ─────────────────────────────────────────────────────────────────
    last_restore_result = None
    backup_cfg = (config or {}).get("backup") or {}
    if backup_cfg.get("enabled", True):
        try:
            from src.storage.backup_manager import BackupManager
            backup_mgr = BackupManager(db.db_path, config or {})
            # Drop any live thread-local connection so the broken/snap
            # swap below is safe. Database.close() is idempotent.
            try:
                db.close()
            except Exception:
                pass
            last_restore_result = backup_mgr.auto_restore_if_needed()
            if last_restore_result and last_restore_result.restored:
                logger.critical(
                    "DB auto-restored at startup: snapshot=%s broken=%s",
                    last_restore_result.snapshot_id,
                    last_restore_result.broken_path,
                )
            elif (
                last_restore_result
                and not last_restore_result.restored
                and last_restore_result.reason == "disabled_in_config"
            ):
                logger.critical(
                    "DB corruption detected but "
                    "auto_restore_on_corruption=false. Manual "
                    "intervention required. details=%s",
                    last_restore_result.details,
                )
            # Reopen the DB regardless of restore outcome — for the
            # not-corrupted path we just need the connection back; for
            # the restored path we want the salvaged file initialised
            # (table creation / WAL setup runs idempotently).
            try:
                db.connect()
            except Exception as e:
                logger.error(
                    "Post-auto-restore db.connect() failed: %s", e,
                )
        except Exception as e:  # pragma: no cover - defensive only
            logger.error("auto-restore probe failed: %s", e)
            # Make sure the DB is open even if we crashed mid-probe.
            try:
                if not getattr(db, "connected", False):
                    db.connect()
            except Exception:
                pass

    if analytics is None:
        from src.storage.analytics import AnalyticsEngine
        analytics = AnalyticsEngine(db.db_path, config.get("analytics", {}))
    if ad_lookup is None:
        from src.user_activity.ad_lookup import ADLookup
        ad_lookup = ADLookup(db, config)
    if email_notifier is None:
        from src.user_activity.email_notifier import EmailNotifier
        email_notifier = EmailNotifier(db, config)

    app = FastAPI(title="FILE ACTIVITY Dashboard", version=APP_VERSION)

    # ─────────────────────────────────────────────────────────────────
    # Issue #158 (C-1) — bearer-token gate.
    #
    # DashboardAuth defaults to enabled=true and bypasses localhost so
    # existing dev workflows are unaffected. Static assets are
    # whitelisted ahead of the token check so the frontend can render a
    # "set FILEACTIVITY_DASHBOARD_TOKEN" hint when the token is missing.
    # ─────────────────────────────────────────────────────────────────
    from src.security.dashboard_auth import DashboardAuth
    app.state.dashboard_auth = DashboardAuth(config)

    @app.middleware("http")
    async def auth_middleware(request: Request, call_next):
        # Whitelist: static files + favicon are not gated. Everything
        # else (including / which serves the dashboard SPA) is.
        path = request.url.path or ""
        if path.startswith("/static/") or path == "/favicon.ico":
            return await call_next(request)
        gate = getattr(app.state, "dashboard_auth", None)
        if gate is None or gate.check(request):
            return await call_next(request)
        return JSONResponse({"detail": "Unauthorized"}, status_code=401)

    app.state.analytics = analytics
    app.state.ad_lookup = ad_lookup
    app.state.email_notifier = email_notifier
    # Phase 2 banner state — frontend reads via /api/system/last-restore.
    app.state.last_auto_restore = (
        last_restore_result
        if (last_restore_result and last_restore_result.restored)
        else None
    )

    # Issue #114 Phase 1 — pluggable storage backend abstraction.
    # The manager holds the active backend (sqlite today, elasticsearch
    # in Phase 2). Endpoints continue to talk to ``db`` directly until
    # Phase 3 rewires the query layer; landing the abstraction now means
    # Phase 2 can drop in without churn here.
    from src.storage.backends.manager import StorageManager
    app.state.storage = StorageManager(db, config)

    # Issue #153 Lever A — surface the manual checkpointer for endpoints
    # that want a "now is a good time" hook (e.g. post-archive,
    # post-retention). The Database owns the lifecycle; we just expose
    # it. ``None`` here means init failed and the engine's own
    # auto-checkpoint is in charge — same fallback behaviour as before.
    app.state.checkpointer = getattr(db, "checkpointer", None)

    # Ransomware detector (#37) — watcher pushes events here. Construction is
    # cheap; safe to do unconditionally. The detector pulls its own config
    # block out of `config["security"]["ransomware"]`.
    try:
        from src.security.ransomware_detector import RansomwareDetector
        ransomware = RansomwareDetector(db, config)
        ransomware.email_notifier = email_notifier
        app.state.ransomware = ransomware
    except Exception as e:  # pragma: no cover - defensive only
        logger.warning("RansomwareDetector init failed: %s", e)
        app.state.ransomware = None

    # Syslog/CEF forwarder (#50) — bridges security events to a downstream
    # SIEM (Splunk / Elastic / Sentinel / QRadar). Always constructed; it
    # is a no-op when disabled in config.
    try:
        from src.integrations.syslog_forwarder import SyslogForwarder
        syslog = SyslogForwarder(config)
        app.state.syslog = syslog
        # Wire detector → syslog if both are available.
        if app.state.ransomware is not None and syslog.available:
            app.state.ransomware.set_external_emitter(
                syslog.emit_ransomware_alert
            )
        # Wire audit-chain integrity break → syslog.
        if syslog.available:
            try:
                db.set_audit_break_callback(syslog.emit_audit_break)
            except AttributeError:
                # Older Database without the hook — just skip.
                pass
    except Exception as e:  # pragma: no cover - defensive only
        logger.warning("SyslogForwarder init failed: %s", e)
        app.state.syslog = None

    # ─────────────────────────────────────────────────────────────────
    # Issue #118 Phase 1 — auto error reporter (GitHub Issues).
    # ─────────────────────────────────────────────────────────────────
    try:
        from src.telemetry.error_reporter import ErrorReporter
        app.state.error_reporter = ErrorReporter(config, APP_VERSION)
    except Exception as e:  # pragma: no cover - defensive only
        logger.warning("ErrorReporter init failed: %s", e)
        app.state.error_reporter = None

    @app.exception_handler(Exception)
    async def _telemetry_exception_handler(request, exc):
        reporter = getattr(app.state, "error_reporter", None)
        if reporter is not None:
            try:
                reporter.capture(exc, {
                    "path": str(request.url.path),
                    "method": request.method,
                })
            except Exception:
                pass
        return JSONResponse(
            status_code=500,
            content={"detail": "Internal Server Error"},
        )

    # ─────────────────────────────────────────────────────────────────
    # Security audit 2026-04-28, finding H-1 — CSP + hardening headers.
    #
    # Defence-in-depth against the stored-XSS surface in index.html:
    # paired with the frontend ``escapeHtml()`` sweep, the browser will
    # refuse to execute injected ``<script>`` even if a future regression
    # forgets to escape a leaf value. ``frame-ancestors 'none'`` plus
    # ``X-Frame-Options: DENY`` block clickjacking; ``nosniff`` keeps the
    # browser from MIME-sniffing JSON into HTML. ``Referrer-Policy:
    # no-referrer`` avoids leaking dashboard URLs (which sometimes carry
    # ids/owners) to outbound CDNs.
    #
    # ``script-src`` allows ``'unsafe-inline'`` because index.html embeds
    # ~6k lines of inline JS. Tightening to nonces is a Phase 3 follow-up.
    # ``cdn.jsdelivr.net`` is whitelisted because the dashboard pulls
    # d3 + chart.js from there; remove this line if you self-host them.
    # ─────────────────────────────────────────────────────────────────
    @app.middleware("http")
    async def _csp_middleware(request, call_next):
        response = await call_next(request)
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
            "style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data:; "
            "font-src 'self' data:; "
            "connect-src 'self'; "
            "frame-ancestors 'none'; "
            "base-uri 'self'; "
            "form-action 'self'"
        )
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Referrer-Policy"] = "no-referrer"
        return response

    # Issue #125 — process-local operations registry.
    if operations_registry is not None:
        app.state.operations = operations_registry
    else:
        try:
            from src.storage.operations_tracker import OperationsRegistry
            app.state.operations = OperationsRegistry()
        except Exception as e:  # pragma: no cover - defensive only
            logger.warning("OperationsRegistry init failed: %s", e)
            app.state.operations = None

    # Issue #132 — shared helper for endpoints that need to return a
    # "scan in progress" placeholder when no cached summary exists yet.
    # The shape is fixed: {has_data, scan_in_progress, scan_id,
    # progress_pct, message, reason}. progress_pct comes from the
    # in-memory scan progress dict (file_count vs. an estimate); when
    # we have no estimate the bar is rendered as indeterminate by the
    # frontend.
    def _partial_overview_response(source_id: int) -> Optional[dict]:
        """Issue #139 — try the rolling partial summary for an in-flight scan.

        Returns a dashboard-shaped dict when a running scan has at least
        one persisted partial-summary snapshot, otherwise None so the
        caller can fall through to the existing scan-in-progress
        placeholder. Wraps both calls in a try/except so legacy DBs
        without the ``partial_summary_json`` column don't break the
        endpoint — they just return None and the caller picks the
        original behaviour up.
        """
        from src.utils.size_formatter import format_size
        try:
            running_scan_id = db.is_scan_running(source_id)
        except Exception:
            return None
        if running_scan_id is None:
            return None
        try:
            partial = db.get_scan_partial_summary(running_scan_id)
        except Exception as e:  # pragma: no cover - defensive
            logger.debug("partial summary read failed: %s", e)
            return None
        if not partial:
            return None
        # Format sizes for the cards. ``partial`` was authored by
        # ``compute_partial_summary`` in ``src/analyzer/partial_summary``
        # which uses the canonical keys (total_files, total_size,
        # unique_owners, top_extensions, size_buckets, age_buckets).
        out = dict(partial)
        out["total_size_formatted"] = format_size(out.get("total_size", 0) or 0)
        for ext in out.get("top_extensions", []) or []:
            try:
                ext["size_formatted"] = format_size(ext.get("size", 0) or 0)
            except Exception:
                ext["size_formatted"] = "0 B"
        out["has_data"] = True
        out["is_partial"] = True
        out["scan_in_progress"] = True
        out["scan_id"] = running_scan_id
        return out

    def _scan_in_progress_response(source_id: int, running_scan_id: int,
                                   reason: str = "scan_in_progress") -> dict:
        from src.scanner.file_scanner import get_scan_progress
        progress = get_scan_progress(source_id) or {}
        if not isinstance(progress, dict):
            progress = {}
        # Heuristic %: file_count vs. last-known total from previous scan
        # if any. Falls back to None (frontend renders indeterminate).
        file_count = int(progress.get("file_count", 0) or 0)
        pct: Optional[int] = None
        try:
            with db.get_read_cursor() as cur:
                cur.execute(
                    "SELECT total_files FROM scan_runs "
                    "WHERE source_id = ? AND status = 'completed' "
                    "ORDER BY completed_at DESC LIMIT 1",
                    (source_id,),
                )
                row = cur.fetchone()
                last_total = int(row["total_files"]) if row and row.get("total_files") else 0
                if last_total > 0 and file_count > 0:
                    pct = max(0, min(99, int(file_count * 100 / last_total)))
        except Exception as e:  # pragma: no cover - defensive
            logger.debug("progress pct calc failed: %s", e)
        if pct is None:
            message = "Tarama devam ediyor"
        else:
            message = f"Tarama devam ediyor, %{pct} tamamlandi"
        return {
            "has_data": False,
            "scan_in_progress": True,
            "scan_id": running_scan_id,
            "progress_pct": pct,
            "file_count": file_count,
            "message": message,
            "reason": reason,
        }

    static_dir = os.path.join(os.path.dirname(__file__), "static")
    if os.path.exists(static_dir):
        app.mount("/static", StaticFiles(directory=static_dir), name="static")

    # --- HTML Endpoint ---

    @app.get("/", response_class=HTMLResponse)
    async def index():
        html_path = os.path.join(static_dir, "index.html")
        if os.path.exists(html_path):
            with open(html_path, "r", encoding="utf-8") as f:
                return f.read()
        return "<h1>FILE ACTIVITY Dashboard</h1><p>static/index.html bulunamadi.</p>"

    # --- SOURCE API (ID-based) ---

    @app.get("/api/sources")
    async def get_sources():
        sources = db.get_sources()
        return [s.__dict__ for s in sources]

    @app.get("/api/dashboard/init")
    async def dashboard_init():
        """Hizli baslangic - scan_runs + fallback scanned_files."""
        from src.utils.size_formatter import format_size
        sources = db.get_sources()
        source_list = [s.__dict__ for s in sources]

        summaries = {}
        for s in sources:
            try:
                with db.get_cursor() as cur:
                    # 1) Son tamamlanmis tarama (completed once priority)
                    cur.execute("""
                        SELECT id, total_files, total_size, started_at, completed_at, status
                        FROM scan_runs WHERE source_id=?
                        ORDER BY
                            CASE WHEN status='completed' THEN 0 ELSE 1 END,
                            started_at DESC
                        LIMIT 1
                    """, (s.id,))
                    scan = cur.fetchone()
                    if not scan:
                        summaries[s.id] = {"has_data": False, "scan_id": None}
                        continue

                    file_count = scan["total_files"] or 0
                    total_size = scan["total_size"] or 0

                    # 2) FALLBACK: total_files=0 ama dosyalar var olabilir
                    if file_count == 0:
                        cur.execute("""
                            SELECT COUNT(*) as cnt, COALESCE(SUM(file_size),0) as sz
                            FROM scanned_files WHERE source_id=? AND scan_id=?
                        """, (s.id, scan["id"]))
                        real = cur.fetchone()
                        if real["cnt"] > 0:
                            file_count = real["cnt"]
                            total_size = real["sz"]
                            # scan_runs'i da guncelle (bir dahaki sefere hizli gelsin)
                            cur.execute("""
                                UPDATE scan_runs SET total_files=?, total_size=?
                                WHERE id=? AND total_files=0
                            """, (file_count, total_size, scan["id"]))

                    # Pre-computed KPI summary (scan tamamlanirken
                    # yazilmisti). Varsa Overview bu satirdan render eder,
                    # scanned_files tablosuna hic dokunmaz.
                    kpi = db.get_scan_summary(scan["id"])

                    summaries[s.id] = {
                        "has_data": file_count > 0,
                        "scan_id": scan["id"],
                        "file_count": file_count,
                        "total_size": total_size,
                        "total_size_formatted": format_size(total_size),
                        "scan_status": {
                            "started_at": scan["started_at"],
                            "completed_at": scan["completed_at"],
                            "status": scan["status"],
                        },
                        "kpi": kpi,  # None ise frontend eski yola duser
                    }
            except Exception:
                summaries[s.id] = {"has_data": False, "scan_id": None}

        return {
            "sources": source_list,
            "summaries": summaries,
            "auto_select": source_list[0]["id"] if len(source_list) == 1 else (source_list[0]["id"] if source_list else None),
        }

    @app.post("/api/sources")
    async def add_source(data: SourceCreate):
        from src.storage.models import Source
        s = Source(name=data.name, unc_path=data.unc_path, archive_dest=data.archive_dest)
        try:
            sid = db.add_source(s)
            return {"id": sid, "message": f"Kaynak eklendi: {data.name}"}
        except Exception as e:
            raise HTTPException(400, str(e))

    @app.delete("/api/sources/{source_id}")
    async def remove_source(source_id: int):
        src = _get_source(db, source_id)
        if db.remove_source(src.name):
            return {"message": f"Kaynak silindi: {src.name}"}
        raise HTTPException(500, "Silme basarisiz")

    @app.post("/api/sources/{source_id}/test")
    async def test_source(source_id: int):
        from src.scanner.share_resolver import test_connectivity
        src = _get_source(db, source_id)
        ok, msg = test_connectivity(src.unc_path)
        return {"success": ok, "message": msg}

    # --- SCAN API (ID-based, async background) ---

    import threading
    _scan_threads = {}  # source_id -> thread
    _scan_results = {}  # source_id -> result

    # Issue #131 — keep a process-local handle to the active FileScanner
    # per source so the stop endpoint can flip its cancel_event without
    # racing with the worker thread's local references.
    _active_scanners: dict[int, "object"] = {}

    @app.post("/api/scan/{source_id}")
    async def run_scan(source_id: int):
        from src.scanner.file_scanner import (
            FileScanner,
            get_or_create_cancel_event,
            reset_cancel_event,
        )

        # Zaten tarama yapiliyor mu?
        if source_id in _scan_threads and _scan_threads[source_id].is_alive():
            return {"status": "already_running", "message": "Bu kaynak zaten taraniyor"}

        src = _get_source(db, source_id)

        # Fresh cancel_event for this scan (clears any stale ``set()`` left
        # over from a previous run that completed without a stop).
        reset_cancel_event(source_id)
        cancel_event = get_or_create_cancel_event(source_id)

        def _scan_worker():
            # Issue #125: register the scan in the operations tracker so the
            # "su an ne oluyor" banner reflects activity. Tracker calls are
            # wrapped — a failure here MUST NOT break the scan.
            registry = getattr(app.state, "operations", None)
            op_id = None
            try:
                if registry is not None:
                    op_id = registry.start(
                        "scan",
                        f"Tarama: {src.name}",
                        metadata={"source_id": src.id},
                    )
            except Exception as e:  # pragma: no cover - defensive only
                logger.debug("ops.start(scan) failed: %s", e)
            try:
                scanner = FileScanner(db, config)
                # Wire the shared cancel_event so /api/scan/{id}/stop can
                # break out of the main scan loop at the next batch.
                scanner.cancel_event = cancel_event
                # Issue #137 (replacing #135 ops_registry/op_id coupling) —
                # feed mid-walk MFT counters into the operations registry so
                # /api/scan/progress/{id} can surface a ``live_count`` while
                # the scanner is still in the (silent-to-the-DB) MFT collection
                # phase. Tracker outage MUST NOT break the scan — wrap the
                # whole callback in try/except.
                if registry is not None and op_id:
                    def _mft_progress(stage: str, processed: int) -> None:
                        try:
                            registry.progress(
                                op_id,
                                label=f"MFT okuma: {processed:,} kayit",
                                processed=processed,
                            )
                        except Exception as e:  # pragma: no cover
                            logger.debug("ops.progress(scan) failed: %s", e)
                    scanner.progress_callback = _mft_progress
                _active_scanners[source_id] = scanner
                result = scanner.scan_source(src.id, src.name, src.unc_path)
                _scan_results[source_id] = result
            finally:
                _active_scanners.pop(source_id, None)
                try:
                    if registry is not None and op_id:
                        registry.finish(op_id)
                except Exception as e:  # pragma: no cover - defensive only
                    logger.debug("ops.finish(scan) failed: %s", e)

        t = threading.Thread(target=_scan_worker, daemon=True)
        _scan_threads[source_id] = t
        _scan_results.pop(source_id, None)
        t.start()

        return {"status": "started", "message": f"Tarama baslatildi: {src.name}"}

    @app.post("/api/scan/{source_id}/stop")
    async def stop_scan(source_id: int):
        """Issue #131 — kullanici tetikli iptal.

        Resolves the active scan_run for ``source_id``, sets the shared
        cancel_event, waits up to 30s for the worker thread to exit
        cleanly, and otherwise force-marks the scan_run as
        ``status='cancelled'``. Always returns 200 with a JSON body
        describing the outcome (cancelled, partial_files, scan_id).

        - 200 with cancelled=False if no scan was running.
        - 200 with cancelled=True if cancel succeeded (clean or forced).

        Audit: writes a single ``scan_cancelled`` audit event so admins
        can later trace user-initiated stops in the chain. Failures to
        write audit are non-fatal (best-effort).
        """
        from src.scanner.file_scanner import (
            get_or_create_cancel_event,
            get_scan_progress,
        )

        thread = _scan_threads.get(source_id)
        is_running = thread is not None and thread.is_alive()

        # Look up the active scan_run id (status='running').
        active_scan_id: Optional[int] = None
        try:
            incomplete = db.get_incomplete_scan(source_id)
            if incomplete:
                active_scan_id = incomplete.get("scan_id")
        except Exception as e:
            logger.warning("stop_scan: get_incomplete_scan failed: %s", e)

        if not is_running and active_scan_id is None:
            return {
                "cancelled": False,
                "scan_id": None,
                "partial_files": 0,
                "reason": "no_active_scan",
            }

        # Capture partial file count from progress before the worker exits.
        progress = get_scan_progress(source_id) or {}
        partial_before = int(progress.get("file_count", 0) or 0) \
            if isinstance(progress, dict) else 0

        # Signal cancellation. Shared event — worker thread checks at
        # every batch boundary (default 1000 files).
        cancel_event = get_or_create_cancel_event(source_id)
        cancel_event.set()

        # Wait up to 30s for the worker to exit on its own. The scan
        # thread flushes pending rows, marks the scan_run as cancelled,
        # and returns; we don't want to block the request indefinitely.
        if is_running and thread is not None:
            thread.join(timeout=30.0)

        forced = False
        if is_running and thread is not None and thread.is_alive():
            # Worker still running after 30s — log + force the DB row to
            # ``cancelled`` so the dashboard doesn't keep showing a
            # 'running' scan that's actually a zombie. The thread will
            # eventually finish on its own (or the process restarts).
            logger.warning(
                "stop_scan: thread still alive after 30s for source_id=%d, "
                "force-marking scan_run as cancelled",
                source_id,
            )
            forced = True
            if active_scan_id is not None:
                try:
                    db.complete_scan_run(
                        active_scan_id,
                        partial_before,
                        int(progress.get("total_size", 0) or 0)
                            if isinstance(progress, dict) else 0,
                        int(progress.get("errors", 0) or 0)
                            if isinstance(progress, dict) else 0,
                        "cancelled",
                    )
                except Exception as e:
                    logger.warning("stop_scan: force-mark cancelled failed: %s", e)

        # Final partial count: prefer the worker's post-exit progress,
        # which reflects the last flushed batch.
        progress_after = get_scan_progress(source_id) or {}
        partial_after = int(progress_after.get("file_count", 0) or 0) \
            if isinstance(progress_after, dict) else partial_before

        # Issue #129 ops registry: best-effort cleanup if worker didn't
        # exit cleanly (the worker's finally-block normally calls
        # finish; this is a safety net for the forced path).
        if forced:
            registry = getattr(app.state, "operations", None)
            if registry is not None:
                try:
                    for op in list(registry.list_active()):
                        if op.type == "scan" and \
                                op.metadata.get("source_id") == source_id:
                            registry.finish(op.op_id)
                except Exception as e:  # pragma: no cover - defensive
                    logger.debug("ops registry cleanup failed: %s", e)

        # Audit (best-effort).
        try:
            db.insert_audit_event_simple(
                source_id=source_id,
                event_type="scan_cancelled",
                username="admin",
                file_path=None,
                details=(
                    f"scan_id={active_scan_id};partial_files={partial_after};"
                    f"forced={forced}"
                ),
                detected_by="dashboard",
            )
        except Exception as e:  # pragma: no cover - audit best-effort
            logger.warning("scan_cancelled audit yazilamadi: %s", e)

        return {
            "cancelled": True,
            "scan_id": active_scan_id,
            "partial_files": partial_after,
            "forced": forced,
        }

    @app.get("/api/scan/progress/{source_id}")
    async def scan_progress(source_id: int):
        # Issue #135 — endpoint now exposes ``phase``,
        # ``phase_pct`` (best-effort), ``scan_id`` and ``total_size_bytes``
        # so the frontend can render a granular label
        # ("MFT okunuyor" / "DB'ye yaziliyor" / "Analiz calisiyor").
        # Falls back to scan_runs.current_phase when the in-memory
        # progress dict is missing (e.g. dashboard load mid-scan).
        from src.scanner.file_scanner import get_scan_progress
        progress = get_scan_progress(source_id)

        # Thread durumunu kontrol et
        is_running = source_id in _scan_threads and _scan_threads[source_id].is_alive()
        result = _scan_results.get(source_id)

        # Issue #137 — surface the live row counter from the operations
        # registry so the Sources page card and DOSYA KPI can stay in
        # sync with the ops banner during the MFT collection phase.
        # During that phase the DB scan_runs row + the in-memory
        # ``progress["file_count"]`` are still 0 (the scanner hasn't
        # iterated MFT records into batches yet) — but the MFT backend
        # already reports a structured ``processed`` counter to the
        # registry. Expose it as ``live_count`` and let the frontend
        # prefer whichever number is larger.
        live_count: Optional[int] = None
        ops = getattr(app.state, "operations", None)
        if ops is not None:
            try:
                live_op = ops.find_active_op_by_metadata(source_id=source_id)
                if live_op is not None:
                    val = (live_op.metadata or {}).get("processed")
                    if val is not None:
                        try:
                            live_count = int(val)
                        except (TypeError, ValueError):
                            live_count = None
            except Exception as e:  # pragma: no cover - defensive only
                logger.debug("live_count lookup failed: %s", e)

        if result and not is_running:
            # Tarama bitti, sonucu dondur
            payload = {
                **result,
                "status": "completed",
                "phase": "completed",
                "phase_pct": 100,
                "finished": True,
                "live_count": live_count,
            }
            if "total_size" in result and "total_size_bytes" not in payload:
                payload["total_size_bytes"] = result.get("total_size")
            return payload

        if progress:
            phase = progress.get("phase") or "enumeration"
            phase_pct = _phase_progress_pct(phase, progress.get("file_count", 0))
            scan_id = None
            try:
                # Pull live scan_id from the incomplete-scan helper so the
                # frontend can deep-link if it wants to.
                inc = db.get_incomplete_scan(source_id)
                if inc:
                    scan_id = inc.get("scan_id")
            except Exception:
                pass
            return {
                **progress,
                "phase": phase,
                "phase_pct": phase_pct,
                "scan_id": scan_id,
                "total_size_bytes": progress.get("total_size", 0),
                "finished": False,
                "live_count": live_count,
            }

        return {
            "status": "idle",
            "phase": None,
            "phase_pct": 0,
            "file_count": 0,
            "total_size": 0,
            "total_size_bytes": 0,
            "finished": False,
            "live_count": live_count,
        }

    # --- COMPATIBILITY REPORT API ---

    @app.get("/api/scan/compatibility/{source_id}")
    async def scan_compatibility(source_id: int):
        """Son taramanin dosya adi uyumluluk raporu."""
        result = _scan_results.get(source_id)
        if result and "compatibility" in result:
            return result["compatibility"]
        # Progress'ten al
        from src.scanner.file_scanner import get_scan_progress
        progress = get_scan_progress(source_id)
        if progress and "compatibility" in progress:
            return progress["compatibility"]
        return {"total_files_analyzed": 0, "health_score": 100, "issues": [], "summary": {}}

    # --- REPORT API (ID-based) ---

    @app.get("/api/reports/status/{source_id}")
    async def report_status(source_id: int):
        from src.analyzer.report_generator import ReportGenerator
        src = _get_source(db, source_id)
        gen = ReportGenerator(db, config)
        return gen.generate_status_report(src.id)

    # Issue #125 — context manager that wraps a block in start/finish on
    # the operations registry. Tracker outage MUST NOT break the work,
    # so all tracker calls are individually try/except'd.
    @contextlib.contextmanager
    def _track_op(op_type: str, label: str, metadata: Optional[dict] = None):
        registry = getattr(app.state, "operations", None)
        op_id = None
        try:
            if registry is not None:
                op_id = registry.start(op_type, label, metadata=metadata)
        except Exception as e:  # pragma: no cover - defensive only
            logger.debug("ops.start(%s) failed: %s", op_type, e)
        try:
            yield
        finally:
            try:
                if registry is not None and op_id:
                    registry.finish(op_id)
            except Exception as e:  # pragma: no cover - defensive only
                logger.debug("ops.finish(%s) failed: %s", op_type, e)

    @app.get("/api/reports/frequency/{source_id}")
    async def report_frequency(source_id: int, days: Optional[str] = None):
        from src.analyzer.report_generator import ReportGenerator
        from src.analyzer import cache as analyzer_cache
        src = _get_source(db, source_id)
        # Once v2 summary_json'daki age_buckets'a bak — scanned_files'i
        # hic tarama. Custom days verildiyse klasik hesaplamaya dus.
        if not days:
            scan_id = db.get_latest_scan_id(src.id, include_running=True)
            if scan_id:
                summary = db.get_scan_summary(scan_id)
                if summary and isinstance(summary, dict) and "age_buckets" in summary:
                    from datetime import datetime
                    return {
                        "source": {"id": source_id, "name": src.name},
                        "scan_id": scan_id,
                        "age_buckets": summary.get("age_buckets"),
                        "frequency": {
                            "age_buckets": summary.get("age_buckets"),
                            "total_files": summary.get("total_files"),
                            "total_size": summary.get("total_size"),
                        },
                        "from_summary": True,
                        "generated_at": datetime.now().isoformat(),
                    }
        gen = ReportGenerator(db, config)
        custom = [int(d) for d in days.split(",")] if days else None
        # Custom-days variants get a distinct cache slot per bucket spec
        # (otherwise different callers would shadow each other).
        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if scan_id is None:
            return gen.generate_frequency_report(src.id, custom)
        analyzer_name = "frequency" if not custom else f"frequency:{','.join(map(str, custom))}"
        with _track_op(
            "analysis",
            f"Erisim sikligi analizi: {src.name}",
            metadata={"source_id": src.id},
        ):
            envelope = analyzer_cache.get_or_compute(
                db, analyzer_name, scan_id,
                lambda: gen.generate_frequency_report(src.id, custom),
            )
            return _attach_cache_envelope(envelope)

    @app.get("/api/reports/types/{source_id}")
    async def report_types(source_id: int):
        from src.analyzer.report_generator import ReportGenerator
        from src.analyzer import cache as analyzer_cache
        src = _get_source(db, source_id)
        with _track_op(
            "analysis",
            f"Tur analizi: {src.name}",
            metadata={"source_id": src.id},
        ):
            gen = ReportGenerator(db, config)
            scan_id = db.get_latest_scan_id(src.id, include_running=True)
            if scan_id is None:
                return gen.generate_type_report(src.id)
            envelope = analyzer_cache.get_or_compute(
                db, "types", scan_id,
                lambda: gen.generate_type_report(src.id),
            )
            return _attach_cache_envelope(envelope)

    @app.get("/api/reports/sizes/{source_id}")
    async def report_sizes(source_id: int):
        from src.analyzer.report_generator import ReportGenerator
        from src.analyzer import cache as analyzer_cache
        src = _get_source(db, source_id)
        with _track_op(
            "analysis",
            f"Boyut analizi: {src.name}",
            metadata={"source_id": src.id},
        ):
            gen = ReportGenerator(db, config)
            scan_id = db.get_latest_scan_id(src.id, include_running=True)
            if scan_id is None:
                return gen.generate_size_report(src.id)
            envelope = analyzer_cache.get_or_compute(
                db, "sizes", scan_id,
                lambda: gen.generate_size_report(src.id),
            )
            return _attach_cache_envelope(envelope)

    @app.get("/api/reports/full/{source_id}")
    async def report_full(source_id: int):
        """Issue #132: when no completed scan AND a scan is running,
        return the scan-in-progress banner shape rather than forcing
        ReportGenerator to compute over an empty/partial table.

        Issue #139: when a scan is running and a partial-summary
        snapshot is available, return it (with ``is_partial: True``) so
        the Reports page renders rolling KPIs instead of the empty
        placeholder.
        """
        from src.analyzer.report_generator import ReportGenerator
        src = _get_source(db, source_id)
        completed_scan_id = db.get_latest_scan_id(src.id, include_running=False)
        if completed_scan_id is None:
            partial_resp = _partial_overview_response(src.id)
            if partial_resp:
                return partial_resp
            running_scan_id = db.is_scan_running(src.id)
            if running_scan_id is not None:
                return _scan_in_progress_response(src.id, running_scan_id,
                                                  reason="no_completed_scan")
        with _track_op(
            "analysis",
            f"Tam rapor: {src.name}",
            metadata={"source_id": src.id},
        ):
            gen = ReportGenerator(db, config)
            return gen.generate_full_report(src.id)

    # --- ARCHIVE API ---

    @app.post("/api/archive/run")
    async def run_archive(data: ArchiveRequest):
        from src.archiver.archive_policy import ArchivePolicyEngine
        from src.archiver.archive_engine import ArchiveEngine

        src = _get_source(db, data.source_id)
        if not src.archive_dest:
            raise HTTPException(400, "Arsiv hedefi tanimli degil")

        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if not scan_id:
            raise HTTPException(400, "Tarama verisi bulunamadi")

        # Issue #158 (C-2) — confirm gate.
        # dry_run resolution: if the caller passed dry_run explicitly
        # we honour it; otherwise we use the config default
        # (archiving.dry_run, now true by default).
        config_dry_run = bool(
            (config or {}).get("archiving", {}).get("dry_run", True)
        )
        effective_dry_run = (
            bool(data.dry_run) if data.dry_run is not None else config_dry_run
        )
        if (not effective_dry_run) and not bool(data.confirm):
            raise HTTPException(
                400,
                "Real archive run requires confirm=true (issue #158 C-2). "
                "Re-submit with {\"dry_run\": false, \"confirm\": true} "
                "after reviewing a dry-run report, or set dry_run=true to "
                "preview without moving files.",
            )

        policy_engine = ArchivePolicyEngine(db)

        if data.policy_name:
            files = policy_engine.get_files_by_policy(src.id, scan_id, data.policy_name)
            archived_by = data.policy_name
        elif data.days:
            files = policy_engine.get_files_by_days(src.id, scan_id, data.days)
            archived_by = f"manual:{data.days}days"
        else:
            raise HTTPException(400, "policy_name veya days belirtmelisiniz")

        if not files:
            return {"archived": 0, "message": "Arsivlenecek dosya yok"}

        engine = ArchiveEngine(db, config)
        # Plumb the resolved dry_run override into the engine so the
        # API contract is honoured even if the engine's default
        # (config.archiving.dry_run) drifts later.
        return engine.archive_files(
            files, src.archive_dest, src.unc_path, src.id, archived_by,
            dry_run=effective_dry_run,
        )

    @app.post("/api/archive/dry-run")
    async def archive_dry_run(data: ArchiveRequest):
        from src.archiver.archive_policy import ArchivePolicyEngine
        from src.utils.size_formatter import format_size

        src = _get_source(db, data.source_id)
        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if not scan_id:
            raise HTTPException(400, "Tarama verisi bulunamadi")

        policy_engine = ArchivePolicyEngine(db)
        days = data.days or 365
        files = policy_engine.get_files_by_days(src.id, scan_id, days)

        total_size = sum(f["file_size"] for f in files) if files else 0
        return {
            "file_count": len(files) if files else 0,
            "total_size": total_size,
            "total_size_formatted": format_size(total_size),
            "sample": files[:20] if files else []
        }

    @app.get("/api/archive/search")
    async def archive_search(q: str, extension: Optional[str] = None,
                              page: int = Query(1, ge=1, le=10000)):
        return db.search_archived_files(q, extension=extension, page=page)

    @app.get("/api/archive/stats")
    async def archive_stats():
        return db.get_archive_stats()

    @app.post("/api/archive/restore")
    async def restore_file(data: RestoreRequest):
        from src.archiver.restore_engine import RestoreEngine
        engine = RestoreEngine(db)
        if data.archive_id:
            return engine.restore_by_id(data.archive_id)
        elif data.original_path:
            return engine.restore_by_path(data.original_path)
        raise HTTPException(400, "archive_id veya original_path gerekli")

    # --- POLICY API ---

    @app.get("/api/policies")
    async def get_policies():
        return db.get_policies()

    @app.post("/api/policies")
    async def add_policy(data: PolicyCreate):
        from src.archiver.archive_policy import ArchivePolicyEngine
        from src.storage.models import ArchivePolicy

        engine = ArchivePolicyEngine(db)
        rules = engine.create_policy_rules(
            access_days=data.access_days, modify_days=data.modify_days,
            min_size=data.min_size, max_size=data.max_size,
            extensions=data.extensions, exclude_extensions=data.exclude_extensions
        )
        pol = ArchivePolicy(name=data.name, source_id=data.source_id, rules_json=rules)
        pid = db.add_policy(pol)
        return {"id": pid, "message": f"Politika olusturuldu: {data.name}"}

    @app.delete("/api/policies/{policy_id}")
    async def remove_policy(policy_id: int):
        pol = db.get_policy_by_id(policy_id)
        if not pol:
            raise HTTPException(404, "Politika bulunamadi")
        if db.remove_policy(pol["name"]):
            return {"message": f"Politika silindi: {pol['name']}"}
        raise HTTPException(500, "Silme basarisiz")

    # --- SCHEDULE API ---

    @app.get("/api/schedules")
    async def get_schedules():
        return db.get_scheduled_tasks()

    @app.post("/api/schedules")
    async def add_schedule(data: ScheduleCreate):
        from src.storage.models import ScheduledTask
        # audit_export tasks are global — no source. Accept source_id=0
        # as the "ignored" sentinel; require a real source for everything else.
        if data.task_type == "audit_export":
            source_id_val = data.source_id  # stored but ignored by runner
        else:
            src = _get_source(db, data.source_id)
            source_id_val = src.id

        policy_id = None
        if data.policy_name:
            pol = db.get_policy_by_name(data.policy_name)
            if not pol:
                raise HTTPException(404, "Politika bulunamadi")
            policy_id = pol["id"]

        task = ScheduledTask(
            task_type=data.task_type, source_id=source_id_val,
            policy_id=policy_id, cron_expression=data.cron_expression
        )
        tid = db.add_scheduled_task(task)
        return {"id": tid, "message": "Zamanlanmis gorev olusturuldu"}

    @app.delete("/api/schedules/{task_id}")
    async def remove_schedule(task_id: int):
        if db.remove_scheduled_task(task_id):
            return {"message": "Gorev silindi"}
        raise HTTPException(404, "Gorev bulunamadi")

    @app.post("/api/schedules/notify-users/run-now/{source_id}")
    async def notify_users_run_now(source_id: int):
        """notify_users gorevi zamanlayiciyi beklemeden hemen calistir.

        EmailNotifier ve ADLookup app.state'ten alinir; ikisi de yoksa
        scheduler'in _run_notify_users'i 'skipped' sonucu doner (hata
        atmaz). Gonderim sonucu: her kullanici icin sayisal ozet.
        """
        from src.scheduler.task_scheduler import TaskScheduler
        src = _get_source(db, source_id)
        ad = getattr(app.state, "ad_lookup", None)
        notifier = getattr(app.state, "email_notifier", None)
        ts = TaskScheduler(db, config, ad_lookup=ad, email_notifier=notifier)
        fake_task = {"id": 0, "source_id": src.id, "task_type": "notify_users"}
        return ts._run_notify_users(fake_task)

    # --- REPORT EXPORT API ---

    @app.get("/api/reports/export/{source_id}")
    async def report_export(source_id: int):
        """HTML rapor dosyasi olustur ve indir."""
        from src.analyzer.report_generator import ReportGenerator
        from src.analyzer.report_exporter import ReportExporter
        src = _get_source(db, source_id)
        gen = ReportGenerator(db, config)
        data = gen.generate_full_report(src.id)
        if "error" in data:
            raise HTTPException(400, data["error"])
        exporter = ReportExporter(config)
        paths = exporter.export_full_report(data, src.name)
        if paths and os.path.exists(paths.get("html_path", "")):
            return FileResponse(
                paths["html_path"],
                filename=os.path.basename(paths["html_path"]),
                media_type="text/html"
            )
        raise HTTPException(500, "Rapor olusturulamadi")

    # --- DRILL-DOWN API ---

    def _run_drilldown(duckdb_fn, sqlite_fn):
        """DuckDB varsa oradan calistir, yoksa SQLite'a dus.

        duckdb_fn: AnalyticsEngine uzerinde cagrilacak bound method
        sqlite_fn: db (Database) uzerinde cagrilacak bound method
        Her ikisi de (*args) -> {"total": int, "files": list} doner.
        """
        def call(*args):
            if analytics.available:
                try:
                    return duckdb_fn(*args)
                except Exception as e:
                    logger.warning("DuckDB drilldown basarisiz, SQLite fallback: %s", e)
            return sqlite_fn(*args)
        return call

    @app.get("/api/drilldown/frequency/{source_id}")
    async def drilldown_frequency(source_id: int, min_days: int = 0,
                                   max_days: Optional[int] = None,
                                   page: int = Query(1, ge=1, le=10000),
                                   limit: int = Query(100, ge=1, le=500)):
        src = _get_source(db, source_id)
        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if not scan_id:
            raise HTTPException(400, "Tarama verisi bulunamadi")
        offset = (page - 1) * limit
        run = _run_drilldown(analytics.get_files_by_frequency, db.get_files_by_frequency)
        result = run(src.id, scan_id, min_days, max_days, limit, offset)
        result["page"] = page
        result["limit"] = limit
        return result

    @app.get("/api/drilldown/type/{source_id}")
    async def drilldown_type(source_id: int, extension: str = "",
                              page: int = Query(1, ge=1, le=10000),
                              limit: int = Query(100, ge=1, le=500)):
        src = _get_source(db, source_id)
        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if not scan_id:
            raise HTTPException(400, "Tarama verisi bulunamadi")
        offset = (page - 1) * limit
        run = _run_drilldown(analytics.get_files_by_extension, db.get_files_by_extension)
        result = run(src.id, scan_id, extension, limit, offset)
        result["page"] = page
        result["limit"] = limit
        return result

    @app.get("/api/drilldown/size/{source_id}")
    async def drilldown_size(source_id: int, min_bytes: int = 0,
                              max_bytes: Optional[int] = None,
                              page: int = Query(1, ge=1, le=10000),
                              limit: int = Query(100, ge=1, le=500)):
        src = _get_source(db, source_id)
        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if not scan_id:
            raise HTTPException(400, "Tarama verisi bulunamadi")
        offset = (page - 1) * limit
        run = _run_drilldown(analytics.get_files_by_size_range, db.get_files_by_size_range)
        result = run(src.id, scan_id, min_bytes, max_bytes, limit, offset)
        result["page"] = page
        result["limit"] = limit
        return result

    @app.get("/api/drilldown/owner/{source_id}")
    async def drilldown_owner(source_id: int, owner: str = "",
                               page: int = Query(1, ge=1, le=10000),
                               limit: int = Query(100, ge=1, le=500)):
        src = _get_source(db, source_id)
        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if not scan_id:
            raise HTTPException(400, "Tarama verisi bulunamadi")
        offset = (page - 1) * limit
        run = _run_drilldown(analytics.get_files_by_owner, db.get_files_by_owner)
        result = run(src.id, scan_id, owner, limit, offset)
        result["page"] = page
        result["limit"] = limit
        return result

    class DrilldownArchiveRequest(BaseModel):
        source_id: int
        filter_type: str  # "frequency", "type", "size", "owner"
        min_days: Optional[int] = None
        max_days: Optional[int] = None
        extension: Optional[str] = None
        min_bytes: Optional[int] = None
        max_bytes: Optional[int] = None
        owner: Optional[str] = None

    @app.post("/api/drilldown/archive")
    async def drilldown_archive(data: DrilldownArchiveRequest):
        from src.archiver.archive_engine import ArchiveEngine
        src = _get_source(db, data.source_id)
        if not src.archive_dest:
            raise HTTPException(400, "Arsiv hedefi tanimli degil")
        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if not scan_id:
            raise HTTPException(400, "Tarama verisi bulunamadi")

        # Collect files based on filter_type
        if data.filter_type == "frequency":
            result = db.get_files_by_frequency(src.id, scan_id, data.min_days or 0, data.max_days, 10000, 0)
        elif data.filter_type == "type":
            result = db.get_files_by_extension(src.id, scan_id, data.extension or "", 10000, 0)
        elif data.filter_type == "size":
            result = db.get_files_by_size_range(src.id, scan_id, data.min_bytes or 0, data.max_bytes, 10000, 0)
        elif data.filter_type == "owner":
            result = db.get_files_by_owner(src.id, scan_id, data.owner or "", 10000, 0)
        else:
            raise HTTPException(400, "Gecersiz filter_type")

        files = result.get("files", [])
        if not files:
            return {"archived": 0, "message": "Arsivlenecek dosya yok"}

        engine = ArchiveEngine(db, config)
        archived_by = f"drilldown:{data.filter_type}"
        return engine.archive_files(files, src.archive_dest, src.unc_path, src.id, archived_by)

    # --- EXPORT API (XLS / PDF) ---

    @app.get("/api/export/xls/{source_id}")
    async def export_xls(source_id: int):
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        src = _get_source(db, source_id)
        try:
            from src.analyzer.report_exporter_v2 import XLSExporter
            exporter = XLSExporter(db, config)
            data = exporter.export_full_report(src.id)
            return StreamingResponse(
                BytesIO(data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=FileActivity_{src.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"}
            )
        except ImportError:
            raise HTTPException(500, "openpyxl kurulu degil. pip install openpyxl")
        except Exception as e:
            logger.error("XLS export error: %s", e)
            raise HTTPException(500, str(e))

    @app.get("/api/export/pdf/{source_id}")
    async def export_pdf(source_id: int):
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        src = _get_source(db, source_id)
        try:
            from src.analyzer.report_exporter_v2 import PDFExporter
            exporter = PDFExporter(db, config)
            data = exporter.export_full_report(src.id)
            return StreamingResponse(
                BytesIO(data),
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=FileActivity_{src.name}_{datetime.now().strftime('%Y%m%d')}.pdf"}
            )
        except ImportError:
            raise HTTPException(500, "reportlab kurulu degil. pip install reportlab")
        except Exception as e:
            logger.error("PDF export error: %s", e)
            raise HTTPException(500, str(e))

    # --- WATCHER CHANGES API ---

    @app.get("/api/watcher/{source_id}/changes")
    async def watcher_changes(source_id: int):
        from src.scanner.file_watcher import _watchers
        if source_id in _watchers:
            w = _watchers[source_id]
            return {
                "changes": w.stats.get("last_changes", []),
                "total_changes": w.stats.get("total_changes", 0),
                "running": w._running,
            }
        return {"changes": [], "total_changes": 0, "running": False}

    # --- USER ACTIVITY API ---

    @app.get("/api/users/overview")
    async def users_overview(source_id: Optional[int] = None, days: int = 30):
        # Check if we have event log data; if not, fallback to file ownership
        has_logs = db.has_access_log_data()
        if has_logs:
            return {
                "source": "event_log",
                "top_users": db.get_top_users(source_id=source_id, days=days),
                "department_stats": db.get_department_stats(days=days),
                "access_timeline": db.get_access_timeline(source_id=source_id, days=days),
                "heatmap": db.get_hourly_heatmap(source_id=source_id, days=min(days, 7))
            }
        else:
            # Fallback to file ownership from scanned_files
            owners = []
            if source_id:
                owners = db.get_file_owners_stats(source_id)
            else:
                # Try all sources
                sources = db.get_sources()
                for s in sources:
                    sid = db.get_latest_scan_id(s.id)
                    if sid:
                        owners = db.get_file_owners_stats(s.id, sid)
                        break
            return {
                "source": "file_ownership",
                "owners": owners,
                "top_users": [],
                "department_stats": [],
                "access_timeline": [],
                "heatmap": []
            }

    @app.get("/api/users/heatmap")
    async def users_heatmap(source_id: Optional[int] = None, days: int = 7):
        """Haftalik saatlik erisim heatmap'i - ayri endpoint."""
        raw = db.get_hourly_heatmap(source_id=source_id, days=days)
        # HTML expects: {matrix: [[24 vals]*7], max_value: N, days: [str*7]}
        day_names = ["Pazartesi", "Sali", "Carsamba", "Persembe", "Cuma", "Cumartesi", "Pazar"]
        matrix = [[0]*24 for _ in range(7)]
        max_val = 0
        if isinstance(raw, list):
            for entry in raw:
                d = entry.get("dow", 0)
                h = entry.get("hour", 0)
                c = entry.get("count", 0)
                if 0 <= d < 7 and 0 <= h < 24:
                    matrix[d][h] = c
                    if c > max_val:
                        max_val = c
        elif isinstance(raw, dict):
            matrix = raw.get("matrix", matrix)
            max_val = raw.get("max_value", max_val)
        return {"matrix": matrix, "max_value": max_val or 1, "days": day_names}

    @app.get("/api/users/{username}/activity")
    async def user_activity(username: str, days: int = 30):
        return db.get_user_activity(username, days=days)

    @app.get("/api/users/{username}/efficiency")
    async def user_efficiency(username: str,
                               source_id: Optional[int] = None,
                               scan_id: Optional[int] = None):
        """Kullanici verimlilik skoru + uyumsuzluk raporu.

        0-100 arasi skor, faktorler, somut oneriler. source_id ve scan_id
        opsiyonel; hic biri verilmezse son tamamlanmis tarama kullanilir.
        """
        from src.user_activity.efficiency_score import compute_user_score
        return compute_user_score(db, username, source_id=source_id, scan_id=scan_id)

    @app.get("/api/users/{username}/detail")
    async def user_detail(username: str, days: int = 30):
        """Kullanici detay raporu - HTML dashboard icin genisletilmis format."""
        from src.utils.size_formatter import format_size
        activity = db.get_user_activity(username, days=days)

        # hourly distribution - list of 24 values
        hourly_dict = activity.get("hourly_distribution", {})
        hourly = [0] * 24
        if isinstance(hourly_dict, dict):
            for h_str, cnt in hourly_dict.items():
                try:
                    hourly[int(h_str)] = cnt
                except (ValueError, IndexError):
                    pass
        elif isinstance(hourly_dict, list):
            hourly = hourly_dict[:24] if len(hourly_dict) >= 24 else hourly_dict + [0]*(24-len(hourly_dict))

        # summary - fields are nested inside activity["summary"]
        summary = activity.get("summary", {}) or {}
        total_access = summary.get("total_access", 0) or 0
        unique_files = summary.get("unique_files", 0) or 0
        active_days = summary.get("active_days", 0) or 0
        total_bytes = summary.get("total_data", 0) or 0
        reads = summary.get("reads", 0) or 0
        writes = summary.get("writes", 0) or 0
        deletes = summary.get("deletes", 0) or 0

        # risk score - compute from hourly_distribution and summary
        risk = {"score": 0, "level": "normal", "factors": []}

        # top extensions
        top_ext = activity.get("top_extensions", [])

        # action_summary - build from summary data
        action_summary = {
            "read": reads, "write": writes, "delete": deletes,
            "modify": 0, "create": 0, "copy": 0, "rename": 0, "permission_change": 0
        }

        # recent_activity - not available from get_user_activity
        recent = []

        # AD lookup: e-posta + display name (yoksa null alanlar)
        ad_info = ad_lookup.lookup(username) or {
            "email": None, "display_name": None, "found": False, "source": None
        }

        return {
            "username": username,
            "ad": {
                "email": ad_info.get("email"),
                "display_name": ad_info.get("display_name"),
                "found": ad_info.get("found", False),
                "source": ad_info.get("source"),
            },
            "summary": {
                "total_access": total_access,
                "unique_files": unique_files,
                "active_days": active_days,
                "total_data_formatted": format_size(total_bytes),
                "reads": reads,
                "writes": writes,
                "deletes": deletes
            },
            "risk_score": risk,
            "hourly": hourly,
            "top_extensions": top_ext,
            "action_summary": action_summary,
            "recent_activity": recent
        }

    # --- ANOMALY API ---

    @app.get("/api/anomalies")
    async def get_anomalies(severity: Optional[str] = None, days: int = 7):
        """Anomali listesi - HTML duz array bekliyor."""
        alerts = db.get_anomalies(severity=severity, days=days)
        # HTML expects flat array, not {alerts:[], summary:{}}
        if isinstance(alerts, list):
            return alerts
        return []

    @app.post("/api/anomalies/{anomaly_id}/acknowledge")
    async def acknowledge_anomaly(anomaly_id: int, by_user: str = "admin"):
        db.acknowledge_anomaly(anomaly_id, by_user)
        return {"message": "Anomali onaylandi"}

    # --- FILE WATCHER API ---

    @app.post("/api/watcher/{source_id}/start")
    async def start_watcher(source_id: int,
                             interval: Optional[int] = Query(
                                 default=None, ge=10, le=3600,
                                 description="Polling araligi (saniye). "
                                             "Bos birakilirsa config.watcher.poll_interval_seconds "
                                             "veya 60 saniye kullanilir."
                             )):
        from src.scanner.file_watcher import FileWatcher, DEFAULT_POLL_INTERVAL
        src = _get_source(db, source_id)
        if interval is None:
            interval = int(config.get("watcher", {}).get("poll_interval_seconds",
                                                          DEFAULT_POLL_INTERVAL))
        ransomware = getattr(app.state, "ransomware", None)
        watcher = FileWatcher(db, src.id, src.unc_path, interval,
                                ransomware_detector=ransomware)
        watcher.start()
        return {"status": "started", "interval": watcher.interval}

    @app.post("/api/watcher/{source_id}/stop")
    async def stop_watcher(source_id: int):
        from src.scanner.file_watcher import _watchers
        if source_id in _watchers:
            _watchers[source_id].stop()
            return {"status": "stopped"}
        return {"status": "not_running"}

    @app.get("/api/watcher/status")
    async def watcher_status(source_id: int = None):
        from src.scanner.file_watcher import get_watcher_status
        return get_watcher_status(source_id)

    # --- AUDIT API ---

    @app.get("/api/audit/events")
    async def audit_events(source_id: int = None, event_type: str = None,
                           username: str = None,
                           days: int = Query(7, ge=1, le=365),
                           page: int = Query(1, ge=1, le=10000)):
        return db.get_audit_events(source_id, event_type, username, days, page)

    @app.get("/api/audit/summary")
    async def audit_summary(source_id: int = None, days: int = 7):
        return db.get_audit_summary(source_id, days)

    # --- AUDIT CHAIN API (issue #38) ---

    @app.get("/api/audit/verify")
    async def audit_verify(since_seq: int = Query(1, ge=1),
                           end_seq: Optional[int] = None):
        """Verify the tamper-evident audit chain from since_seq onward.

        Returns ``{verified, total, broken_at, broken_reason}``.
        """
        return db.verify_audit_chain(start_seq=since_seq, end_seq=end_seq)

    @app.get("/api/audit/chain")
    async def audit_chain(page: int = Query(1, ge=1),
                          page_size: int = Query(100, ge=1, le=1000)):
        """Paginated chain rows joined with file_audit_events (newest first)."""
        return db.get_audit_chain_page(page=page, page_size=page_size)

    @app.post("/api/audit/export")
    async def audit_export(start_date: Optional[str] = None,
                           end_date: Optional[str] = None):
        """Trigger a WORM JSONL export of the chain in [start_date, end_date]."""
        from src.storage.audit_export import AuditExporter
        exporter = AuditExporter(db, config)
        return exporter.export_range(start_date, end_date)

    # --- INSIGHTS API ---

    @app.get("/api/insights/{source_id}")
    async def get_insights(source_id: int, refresh: bool = False):
        """AI Insights: son scan icin cached sonucu doner, yoksa hesaplar.

        refresh=true ile yeniden hesaplamayi force ederek cache'i tazeler.
        Cache'te yoksa ilk cagri agir; sonraki cagrilar anlik.

        Issue #132: when no cache and a scan is running, return the
        scan-in-progress banner shape instead of forcing a heavy
        InsightsEngine compute that would contend with the writer.
        """
        from src.analyzer.ai_insights import InsightsEngine
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not refresh and scan_id:
            cached = db.get_scan_insights(scan_id)
            if cached:
                cached["from_cache"] = True
                return cached

        # No cache. If a scan is running, return the partial snapshot
        # (issue #139) when available so the Insights page renders top-
        # extension / size-bucket / age-bucket cards instead of an
        # empty placeholder. Fall back to the in-progress banner if no
        # partial snapshot has been computed yet.
        if not refresh:
            partial_resp = _partial_overview_response(source_id)
            if partial_resp:
                return partial_resp
            running_scan_id = db.is_scan_running(source_id)
            if running_scan_id is not None:
                return _scan_in_progress_response(source_id, running_scan_id,
                                                  reason="insights_not_cached")

        engine = InsightsEngine(db)
        result = engine.generate_insights(source_id)
        # Sonraki acilislar anlik olsun diye scan_id varsa cache'e yaz
        if scan_id:
            try:
                db.save_scan_insights(scan_id, result)
            except Exception as e:
                logger.warning("insights cache yazilamadi: %s", e)
        result["from_cache"] = False
        return result

    @app.post("/api/insights/{source_id}/recompute")
    async def insights_recompute(source_id: int):
        """Insights'i yeniden hesapla ve cache'i tazele."""
        from src.analyzer.ai_insights import InsightsEngine
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, "Tamamlanmis scan yok")
        result = InsightsEngine(db).generate_insights(source_id)
        db.save_scan_insights(scan_id, result)
        return {"status": "ok", "scan_id": scan_id,
                "insights_count": len(result.get("insights", []))}

    # --- RISK SCORE API ---

    @app.get("/api/reports/mit-naming/{source_id}")
    async def mit_naming_report(source_id: int):
        """MIT Libraries dosya adlandirma standartlarina uyum analizi."""
        from src.scanner.file_scanner import MITNamingAnalyzer

        scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            raise HTTPException(404, "Tarama bulunamadi")

        analyzer = MITNamingAnalyzer()
        with db.get_cursor() as cur:
            cur.execute("""
                SELECT file_path, file_name FROM scanned_files
                WHERE scan_id=?
            """, (scan_id,))
            for row in cur:
                analyzer.analyze(row["file_path"], row["file_name"])

        return analyzer.get_report()

    @app.get("/api/reports/mit-naming/{source_id}/files")
    async def mit_naming_files(source_id: int, code: str = "R1",
                                page: int = Query(1, ge=1, le=10000),
                                page_size: int = Query(100, ge=1, le=500)):
        """MIT ihlal koduna gore dosya listesi (R1,R2,R3,R4,B1,B2,B3,B4,B5,B6)."""
        import re as re_mod
        from src.utils.size_formatter import format_size

        scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            raise HTTPException(404, "Tarama bulunamadi")

        # Kural tanimlar
        checks = {
            "R1": lambda p, n: bool(re_mod.search(r'\s', n)),
            "R2": lambda p, n: bool(n) and not re_mod.match(r'^[a-zA-Z]', n),
            "R3": lambda p, n: bool(n) and '.' in n and not re_mod.match(r'^[a-zA-Z0-9._-]+$', n[:n.rfind('.')]),
            "R4": lambda p, n: '.' not in n or not n.rsplit('.', 1)[-1].isalpha(),
            "B1": lambda p, n: len(n) > 31,
            "B2": lambda p, n: len(p) > 256,
            "B3": lambda p, n: '.' in n and n[:n.rfind('.')].count('.') > 0,
            "B4": lambda p, n: bool(re_mod.search(r'[A-Z]', n[:n.rfind('.')] if '.' in n else n)),
            "B5": lambda p, n: len(n) > 10 and '_' not in n and '-' not in n,
            "B6": lambda p, n: any('.' in part and part not in ('', '.', '..') for part in p.replace('\\', '/').split('/')),
        }

        check_fn = checks.get(code.upper())
        if not check_fn:
            raise HTTPException(400, f"Gecersiz kod: {code}. Gecerli: {', '.join(checks.keys())}")

        # Dosyalari tara ve ihlal edenleri topla
        matching = []
        with db.get_cursor() as cur:
            cur.execute("""
                SELECT id, file_path, file_name, file_size, owner, last_modify_time
                FROM scanned_files WHERE scan_id=?
            """, (scan_id,))
            for row in cur:
                if check_fn(row["file_path"], row["file_name"]):
                    matching.append(dict(row))

        total = len(matching)
        offset = (page - 1) * page_size
        page_files = matching[offset:offset + page_size]
        for f in page_files:
            f["file_size_formatted"] = format_size(f.get("file_size", 0))
            f["directory"] = f["file_path"].rsplit('\\', 1)[0] if '\\' in f["file_path"] else f["file_path"].rsplit('/', 1)[0]

        return {
            "code": code.upper(),
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, -(-total // page_size)),
            "files": page_files
        }

    @app.get("/api/reports/mit-naming/{source_id}/export")
    async def mit_naming_export(source_id: int):
        """MIT ihlal raporunu Excel olarak export et."""
        import re as re_mod
        from fastapi.responses import StreamingResponse
        import io

        scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            raise HTTPException(404, "Tarama bulunamadi")

        checks = {
            "R1": ("Bosluk Iceren", lambda p, n: bool(re_mod.search(r'\s', n))),
            "R2": ("Ilk Karakter Harf Degil", lambda p, n: bool(n) and not re_mod.match(r'^[a-zA-Z]', n)),
            "R3": ("Yasak Karakter", lambda p, n: bool(n) and '.' in n and not re_mod.match(r'^[a-zA-Z0-9._-]+$', n[:n.rfind('.')])),
            "R4": ("Uzanti Sorunu", lambda p, n: '.' not in n or not n.rsplit('.', 1)[-1].isalpha()),
            "B1": ("Uzun Ad (>31)", lambda p, n: len(n) > 31),
            "B3": ("Base Nokta", lambda p, n: '.' in n and n[:n.rfind('.')].count('.') > 0),
            "B4": ("Buyuk Harf", lambda p, n: bool(re_mod.search(r'[A-Z]', n[:n.rfind('.')] if '.' in n else n))),
            "B5": ("Ayirici Yok", lambda p, n: len(n) > 10 and '_' not in n and '-' not in n),
        }

        # Tum dosyalari tara
        violations = {code: [] for code in checks}
        with db.get_cursor() as cur:
            cur.execute("""
                SELECT file_path, file_name, file_size, owner, last_modify_time
                FROM scanned_files WHERE source_id=? AND scan_id=?
            """, (source_id, scan_id))
            for row in cur:
                for code, (label, fn) in checks.items():
                    if fn(row["file_path"], row["file_name"]):
                        violations[code].append(dict(row))

        # CSV olustur (Excel uyumlu)
        output = io.StringIO()
        output.write('\ufeff')  # BOM for Excel UTF-8
        output.write('Ihlal Kodu,Ihlal Turu,Dosya Adi,Tam Yol,Boyut,Sahip,Son Degisiklik\n')
        for code, (label, _) in checks.items():
            for f in violations[code][:5000]:  # Her kod icin max 5000
                path = f["file_path"].replace('"', '""')
                name = f["file_name"].replace('"', '""')
                owner = (f.get("owner") or "").replace('"', '""')
                output.write(f'{code},{label},"{name}","{path}",{f.get("file_size",0)},"{owner}",{f.get("last_modify_time","")}\n')

        output.seek(0)
        from datetime import datetime
        filename = f"MIT_Naming_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @app.get("/api/reports/mit-naming/{source_id}/export.xlsx")
    async def export_mit_naming_xlsx(
        source_id: int,
        ids: Optional[str] = Query(
            None,
            description="Comma-separated scanned_files.id values; empty = all violating rows",
        ),
        format: Optional[str] = Query(
            None,
            description="Set to 'csv' to bypass Excel's 1,048,576-row limit (issue #122)",
        ),
    ):
        """Export MIT-naming violations as XLSX (issue #80; #122 streaming).

        Columns: file_path, owner, last_modify_time, file_size, rule, severity.
        One row per (file × violated rule); a single file violating R1+B4
        produces two rows. ``ids`` filters to scanned_files.id values; empty
        means every violating row in the latest scan.

        For scans that produce more than ~1M violation rows the workbook is
        split across ``Data_1``, ``Data_2``, ... sheets via
        ``write_large_workbook`` (issue #122). Pass ``?format=csv`` to opt
        into a single-file CSV instead — no row cap at all.
        """
        import re as re_mod
        from fastapi.responses import StreamingResponse
        import io
        from datetime import datetime

        from src.utils.xlsx_writer import write_large_workbook, stream_csv

        scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            raise HTTPException(404, "Tarama bulunamadi")

        # Rule definitions: code -> (label, severity, predicate(path, name))
        # severity matches MITNamingAnalyzer.get_report() conventions.
        rules = [
            ("R1", "Bosluk Iceren", "critical",
             lambda p, n: bool(re_mod.search(r"\s", n))),
            ("R2", "Ilk Karakter Harf Degil", "critical",
             lambda p, n: bool(n) and not re_mod.match(r"^[a-zA-Z]", n)),
            ("R3", "Yasak Karakter", "critical",
             lambda p, n: bool(n) and "." in n
             and not re_mod.match(r"^[a-zA-Z0-9._-]+$", n[: n.rfind(".")])),
            ("R4", "Uzanti Sorunu", "critical",
             lambda p, n: "." not in n or not n.rsplit(".", 1)[-1].isalpha()),
            ("B1", "Uzun Ad (>31)", "warning",
             lambda p, n: len(n) > 31),
            ("B2", "Uzun Yol (>256)", "warning",
             lambda p, n: len(p) > 256),
            ("B3", "Base'de Nokta", "warning",
             lambda p, n: "." in n and n[: n.rfind(".")].count(".") > 0),
            ("B4", "Buyuk Harf", "info",
             lambda p, n: bool(re_mod.search(
                 r"[A-Z]", n[: n.rfind(".")] if "." in n else n))),
            ("B5", "Ayirici Yok", "info",
             lambda p, n: len(n) > 10 and "_" not in n and "-" not in n),
            ("B6", "Dizin Adinda Nokta", "info",
             lambda p, n: any(
                 "." in part and part not in ("", ".", "..")
                 for part in p.replace("\\", "/").split("/")
             )),
        ]

        # Parse the optional ids filter once.
        id_filter: Optional[set] = None
        if ids:
            id_filter = set()
            for tok in ids.split(","):
                tok = tok.strip()
                if not tok:
                    continue
                try:
                    id_filter.add(int(tok))
                except ValueError:
                    # Skip non-numeric tokens silently — exporting is best-effort.
                    continue

        columns = [
            {"key": "file_path", "header": "file_path", "width": 60},
            {"key": "owner", "header": "owner", "width": 24},
            {"key": "last_modify_time", "header": "last_modify_time", "width": 22},
            {"key": "file_size", "header": "file_size", "width": 14},
            {"key": "rule", "header": "rule", "width": 28},
            {"key": "severity", "header": "severity", "width": 12},
        ]

        # Generator that streams (file × violating-rule) rows out of the DB
        # cursor. ``write_large_workbook`` and ``stream_csv`` both pull from
        # this lazily — peak memory stays in the MB range even for the 5M-row
        # internal probe (issue #122).
        def _violation_rows():
            with db.get_cursor() as cur:
                cur.execute(
                    """SELECT id, file_path, file_name, owner,
                              last_modify_time, file_size
                       FROM scanned_files
                       WHERE source_id = ? AND scan_id = ?""",
                    (source_id, scan_id),
                )
                for r in cur:
                    if id_filter is not None and r["id"] not in id_filter:
                        continue
                    path = r["file_path"] or ""
                    name = r["file_name"] or ""
                    for code, label, severity, fn in rules:
                        try:
                            if fn(path, name):
                                yield {
                                    "file_path": path,
                                    "owner": r["owner"] or "",
                                    "last_modify_time": r["last_modify_time"] or "",
                                    "file_size": r["file_size"] or 0,
                                    "rule": f"{code} - {label}",
                                    "severity": severity,
                                }
                        except Exception:  # pragma: no cover - defensive predicate guard
                            continue

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        # ---- CSV fallback (issue #122) — bypasses Excel's row cap ---------
        if (format or "").lower() == "csv":
            filename_csv = (
                f"MIT_Naming_Report_source{source_id}_scan{scan_id}_{ts}.csv"
            )
            return StreamingResponse(
                stream_csv(_violation_rows(), columns),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f"attachment; filename={filename_csv}",
                    "X-Format-Fallback": "csv",
                },
            )

        # ---- XLSX (default) — write_only streaming + auto sheet split ----
        buf = io.BytesIO()
        meta = write_large_workbook(_violation_rows(), columns, buf)
        buf.seek(0)

        filename = (
            f"MIT_Naming_Report_source{source_id}_scan{scan_id}_{ts}.xlsx"
        )
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Total-Rows": str(meta["total_rows"]),
                "X-Sheet-Count": str(meta["sheet_count"]),
            },
        )

    @app.get("/api/insights/{source_id}/files")
    async def insight_files(source_id: int, insight_type: str = "stale_1year",
                            page: int = Query(1, ge=1, le=10000),
                            page_size: int = Query(100, ge=1, le=500)):
        """AI insight tipine gore dosya listesi."""
        from src.utils.size_formatter import format_size
        from src.analyzer.ai_insights import get_insight_files

        scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            raise HTTPException(404, "Tarama bulunamadi")

        # Bilinmeyen insight_type artik sessizce bos liste donmuyor —
        # ValueError yukseliyor (issue #82, Bug 2). HTTP 400'e cevirip
        # frontend'in mesaji modalda gostermesini sagliyoruz.
        try:
            files = get_insight_files(db, scan_id, insight_type)
        except ValueError as e:
            raise HTTPException(400, str(e))
        total = len(files)
        offset = (page - 1) * page_size
        page_files = files[offset:offset + page_size]

        for f in page_files:
            f["file_size_formatted"] = format_size(f.get("file_size", 0))
            f["directory"] = f["file_path"].rsplit('\\', 1)[0] if '\\' in f["file_path"] else f["file_path"].rsplit('/', 1)[0]

        return {
            "insight_type": insight_type,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, -(-total // page_size)),
            "files": page_files
        }

    @app.get("/api/overview/{source_id}")
    async def overview(source_id: int):
        """Instant Overview: pre-computed summary'den okur, scanned_files
        tablosuna dokunmaz. Scan tamamlaniyorken kaydedilen JSON'dan gelir.

        Summary henuz hesaplanmamissa (ornek: scan calisiyor, eski scan)
        frontend kendi fallback'ine duser (has_data=false).

        Issue #132: when no cached summary exists AND a scan is currently
        running, return ``{has_data: false, scan_in_progress: true, ...}``
        so the frontend can render a banner ("Tarama devam ediyor, %35
        tamamlandi") inside the page instead of an empty state.
        """
        from src.utils.size_formatter import format_size
        _get_source(db, source_id)
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if scan_id:
            kpi = db.get_scan_summary(scan_id)
            if kpi:
                # Boyutlari formatla
                kpi["total_size_formatted"] = format_size(kpi.get("total_size", 0))
                kpi["stale_size_formatted"] = format_size(kpi.get("stale_size", 0))
                kpi["large_size_formatted"] = format_size(kpi.get("large_size", 0))
                kpi["duplicate_waste_formatted"] = format_size(kpi.get("duplicate_waste_size", 0))
                for ext in kpi.get("top_extensions", []):
                    ext["size_formatted"] = format_size(ext.get("size", 0))
                for owner in kpi.get("top_owners", []):
                    owner["size_formatted"] = format_size(owner.get("size", 0))
                kpi["has_data"] = True
                kpi["is_partial"] = False
                return kpi

        # Issue #139 — no completed-scan summary available. Try the
        # rolling partial-summary snapshot from the active scan first;
        # only fall back to the in-progress placeholder if even the
        # partial snapshot has not been written yet.
        partial_resp = _partial_overview_response(source_id)
        if partial_resp:
            return partial_resp

        running_scan_id = db.is_scan_running(source_id)
        if running_scan_id is not None:
            reason = ("no_completed_scan" if not scan_id
                      else "summary_not_computed")
            return _scan_in_progress_response(source_id, running_scan_id,
                                              reason=reason)
        if scan_id:
            return {"has_data": False, "scan_id": scan_id,
                    "reason": "summary_not_computed"}
        return {"has_data": False, "reason": "no_completed_scan"}

    @app.post("/api/overview/{source_id}/recompute")
    async def overview_recompute(source_id: int):
        """Manuel: son scan icin summary'yi yeniden hesapla.

        Kullanim: scan bittigi sirada summary yazilmadi (crash vs.) ise
        veya KPI algoritmasi guncellendikten sonra.
        """
        _get_source(db, source_id)
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, "Tamamlanmis scan yok")
        summary = db.compute_scan_summary(scan_id)
        return {"status": "ok", "scan_id": scan_id,
                "total_files": summary.get("total_files")}

    @app.get("/api/risk-score/{source_id}")
    async def risk_score(source_id: int):
        """Supervisor risk score - TEK optimized sorgu (6 yerine 2)."""
        scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            return {"risk_score": 0, "kpis": {}}

        # Hizli yol: pre-computed summary varsa oradan hesapla
        kpi = db.get_scan_summary(scan_id)
        if kpi:
            total = kpi.get("total_files", 0) or 1
            risky = kpi.get("risky_count", 0)
            stale = kpi.get("stale_count", 0)
            dup_waste = kpi.get("duplicate_waste_size", 0)
            total_size = kpi.get("total_size", 0) or 1
            # Basit formul: risky % + stale % + (dup_waste / total_size) %
            score = int(min(100,
                (risky / total) * 40 +
                (stale / total) * 30 +
                (dup_waste / total_size) * 30 * 100
            ))
            return {
                "risk_score": score,
                "kpis": {
                    "total_files": kpi.get("total_files"),
                    "risky_files": risky,
                    "stale_files": stale,
                    "duplicate_groups": kpi.get("duplicate_groups"),
                    "total_size": kpi.get("total_size"),
                    "stale_size": kpi.get("stale_size"),
                    "risky_pct": round(risky * 100 / total, 1),
                    "stale_pct": round(stale * 100 / total, 1),
                },
                "source": "precomputed",
            }

        risky_exts = ('exe','bat','ps1','vbs','cmd','msi','scr','com','js','wsf')
        placeholders = ','.join(['?'] * len(risky_exts))
        stale_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')

        with db.get_cursor() as cur:
            # TEK SORGU: 5 metrigi bir seferde hesapla (6 ayri sorgu yerine)
            cur.execute(f"""
                SELECT
                    COUNT(*) as total_files,
                    COALESCE(SUM(file_size), 0) as total_size,
                    SUM(CASE WHEN last_access_time < ? THEN 1 ELSE 0 END) as stale_count,
                    COALESCE(SUM(CASE WHEN last_access_time < ? THEN file_size ELSE 0 END), 0) as stale_size,
                    SUM(CASE WHEN LOWER(extension) IN ({placeholders}) THEN 1 ELSE 0 END) as risky_count,
                    COUNT(DISTINCT CASE WHEN owner IS NOT NULL AND owner != '' THEN owner END) as owner_count,
                    SUM(CASE WHEN file_size > 104857600 THEN 1 ELSE 0 END) as large_count,
                    COALESCE(SUM(CASE WHEN file_size > 104857600 THEN file_size ELSE 0 END), 0) as large_size
                FROM scanned_files
                WHERE source_id=? AND scan_id=?
            """, (stale_date, stale_date) + risky_exts + (source_id, scan_id))
            r = cur.fetchone()
            total_files = r["total_files"]
            total_size = r["total_size"]
            stale_count = r["stale_count"]
            stale_size = r["stale_size"]
            risky_count = r["risky_count"]
            owner_count = r["owner_count"]
            large_count = r["large_count"]
            large_size = r["large_size"]
            stale_pct = round(stale_count * 100 / max(total_files, 1), 1)

            # Duplikasyon sorgusu (ayri cunku GROUP BY gerekli)
            cur.execute("""
                SELECT COUNT(*) as cnt FROM (
                    SELECT file_name, file_size FROM scanned_files
                    WHERE source_id=? AND scan_id=? AND file_size > 1048576
                    GROUP BY file_name, file_size HAVING COUNT(*) > 1
                )
            """, (source_id, scan_id))
            dup_groups = cur.fetchone()["cnt"]

        # Risk skoru hesapla (0-100)
        risk = 0
        risk += min(stale_pct * 0.4, 30)
        risk += min(risky_count / max(total_files, 1) * 1000, 20)
        risk += min(large_size / max(total_size, 1) * 100, 15)
        risk += min(dup_groups * 2, 15)
        risk += 10 if owner_count <= 1 else 0
        risk += 10 if stale_pct > 50 else 0
        risk_score_val = min(round(risk), 100)

        # Watcher
        try:
            from src.scanner.file_watcher import get_watcher_status
            watcher = get_watcher_status(source_id)
            changes_24h = watcher.get("total_changes", 0) if isinstance(watcher, dict) else 0
        except Exception:
            changes_24h = 0

        return {
            "risk_score": risk_score_val,
            "risk_level": "critical" if risk_score_val >= 70 else "warning" if risk_score_val >= 40 else "good",
            "total_files": total_files,
            "total_size": total_size,
            "kpis": {
                "stale_pct": stale_pct,
                "stale_count": stale_count,
                "stale_size": stale_size,
                "risky_files": risky_count,
                "owner_count": owner_count,
                "large_files": large_count,
                "large_size": large_size,
                "dup_groups": dup_groups,
                "changes_24h": changes_24h,
            }
        }

    # --- SCAN TREND API ---

    @app.get("/api/trend/{source_id}")
    async def scan_trend(source_id: int):
        """Storage growth trend from scan history."""
        with db.get_cursor() as cur:
            cur.execute("""
                SELECT id, started_at, completed_at, total_files, total_size, status
                FROM scan_runs WHERE source_id = ? ORDER BY started_at DESC LIMIT 20
            """, (source_id,))
            scans = [dict(r) for r in cur.fetchall()]

        growth = None
        if len(scans) >= 2:
            latest = scans[0]
            previous = scans[1]
            if latest["total_files"] and previous["total_files"]:
                growth = {
                    "file_diff": (latest["total_files"] or 0) - (previous["total_files"] or 0),
                    "size_diff": (latest["total_size"] or 0) - (previous["total_size"] or 0),
                }

        return {"scans": scans[::-1], "growth": growth}

    # --- DUPLIKE RAPOR ve SECICI ARSIV ---

    @app.get("/api/reports/duplicates/{source_id}")
    async def duplicate_report(source_id: int,
                                page: int = Query(1, ge=1, le=10000),
                                page_size: int = Query(50, ge=1, le=500),
                                min_size: int = 0):
        """Kopya dosya raporu - gruplandirmali.

        DuckDB kurulu ve ATTACH basarili ise tek-gecis CTE ile calisir,
        aksi halde SQLite yoluna duser.
        """
        from src.utils.size_formatter import format_size
        result = None
        if analytics.available:
            try:
                scan_id = db.get_latest_scan_id(source_id, include_running=False)
                if scan_id:
                    result = analytics.get_duplicate_groups(
                        scan_id, min_size, page, page_size
                    )
            except Exception as e:
                logger.warning("DuckDB duplicate sorgusu basarisiz, SQLite fallback: %s", e)
                result = None
        if result is None:
            result = db.get_duplicate_groups(
                source_id, min_size=min_size, page=page, page_size=page_size
            )
        # Boyut formatlama
        result["total_waste_size_formatted"] = format_size(result.get("total_waste_size", 0))
        for g in result.get("groups", []):
            g["file_size_formatted"] = format_size(g.get("file_size", 0))
            g["waste_size_formatted"] = format_size(g.get("waste_size", 0))
        return result

    @app.get("/api/export/duplicates/{source_id}")
    async def export_duplicates(source_id: int):
        """Kopya dosyalari CSV olarak export et."""
        from src.utils.size_formatter import format_size
        from fastapi.responses import StreamingResponse
        from datetime import datetime
        import io

        scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            raise HTTPException(404, "Tarama bulunamadi")

        # Tum duplike gruplari getir (max 10000 grup)
        result = db.get_duplicate_groups(source_id, scan_id=scan_id, page=1, page_size=10000)

        output = io.StringIO()
        output.write('\ufeff')  # BOM for Excel
        output.write('Grup,Dosya Adi,Boyut,Boyut (Okunur),Kopya Sayisi,Dosya Yolu,Sahip,Son Erisim,Son Degisiklik\n')

        for idx, g in enumerate(result.get("groups", []), 1):
            for f in g.get("files", []):
                path = (f.get("file_path") or "").replace('"', '""')
                name = (f.get("file_name") or g.get("file_name", "")).replace('"', '""')
                owner = (f.get("owner") or "").replace('"', '""')
                output.write(f'{idx},"{name}",{g.get("file_size",0)},"{format_size(g.get("file_size",0))}",{g.get("count",0)},"{path}","{owner}",{f.get("last_access_time","")},{f.get("last_modify_time","")}\n')

        output.seek(0)
        filename = f"Duplicates_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    # --- CONTENT-HASH DUPLICATE DETECTION (issue #35) ---
    # Tiered pipeline: size -> 4 KB prefix hash -> full SHA-256.
    # Recompute endpoint'i hash'leri yeniden hesaplar ve persist eder;
    # GET endpoint'i cached sonuclari (duplicate_hash_groups) okur.

    @app.post("/api/duplicates/content/{source_id}/compute")
    async def content_duplicates_compute(source_id: int):
        """Icerik-tabanli kopya tespitini calistir ve sonuclari persist et.

        Son tamamlanmis scan_id bulunur, `ContentDuplicateEngine.compute`
        tetiklenir. Donus degeri pipeline istatistikleri.
        """
        from src.analyzer.content_duplicates import ContentDuplicateEngine
        from src.utils.size_formatter import format_size

        _get_source(db, source_id)
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, "Tamamlanmis scan yok")

        engine = ContentDuplicateEngine(db, config)
        stats = engine.compute(scan_id)
        stats["scan_id"] = scan_id
        stats["bytes_hashed_formatted"] = format_size(stats.get("bytes_hashed", 0))
        return stats

    @app.get("/api/duplicates/content/{source_id}")
    async def content_duplicates_report(
        source_id: int,
        page: int = Query(1, ge=1, le=10000),
        page_size: int = Query(50, ge=1, le=500),
    ):
        """Cached icerik-hash kopya raporunu oku (waste_size DESC sirali).

        Hesaplama yapmaz — `POST .../compute` endpoint'i ile uretilir.
        """
        from src.analyzer.content_duplicates import ContentDuplicateEngine
        from src.utils.size_formatter import format_size

        _get_source(db, source_id)
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            return {
                "has_data": False,
                "reason": "no_completed_scan",
                "scan_id": None,
                "total_groups": 0,
                "groups": [],
                "page": page,
                "page_size": page_size,
                "total_pages": 1,
            }

        engine = ContentDuplicateEngine(db, config)
        result = engine.get_report(scan_id, page=page, page_size=page_size)
        result["has_data"] = True
        result["total_waste_size_formatted"] = format_size(result.get("total_waste_size", 0))
        for g in result.get("groups", []):
            g["file_size_formatted"] = format_size(g.get("file_size", 0))
            g["waste_size_formatted"] = format_size(g.get("waste_size", 0))
        return result

    # --- DUPLICATE QUARANTINE (issue #83 Phase 1) ---
    # Quarantine-only delete. Files MOVE to data/quarantine/<YYYYMMDD>/<hash>/,
    # they are NEVER os.remove()'d. Hard delete + auto-cleanup are Phase 2.
    # Both endpoints share the same DuplicateCleaner so config (kill-switch,
    # cap, token requirement) is read once.

    @app.post("/api/reports/duplicates/{source_id}/quarantine/preview")
    async def duplicates_quarantine_preview(source_id: int, request: Request):
        """Dry-run: would_move / skipped_held / skipped_last_copy + size."""
        from src.archiver.duplicate_cleaner import DuplicateCleaner
        _get_source(db, source_id)
        body = await request.json()
        file_ids = body.get("file_ids") or []
        if not isinstance(file_ids, list):
            raise HTTPException(400, "file_ids must be a list")
        cleaner = DuplicateCleaner(db, config)
        try:
            preview = cleaner.preview([int(i) for i in file_ids])
        except (ValueError, TypeError) as e:
            raise HTTPException(400, str(e))
        return preview.to_dict()

    @app.post("/api/reports/duplicates/{source_id}/quarantine")
    async def duplicates_quarantine(source_id: int, request: Request):
        """Move selected files to the quarantine root. Requires confirm
        AND safety_token == "QUARANTINE". Returns full QuarantineResult
        with before/after/delta and the persisted gain_report_id."""
        from src.archiver.duplicate_cleaner import DuplicateCleaner
        _get_source(db, source_id)
        body = await request.json()
        file_ids = body.get("file_ids") or []
        confirm = bool(body.get("confirm", False))
        safety_token = body.get("safety_token", "")
        moved_by = body.get("moved_by") or "system"
        if not isinstance(file_ids, list):
            raise HTTPException(400, "file_ids must be a list")
        cleaner = DuplicateCleaner(db, config)
        with _track_op(
            "archive",
            f"Karantina: {len(file_ids)} dosya",
            metadata={"source_id": source_id, "file_count": len(file_ids)},
        ):
            try:
                result = cleaner.quarantine(
                    file_ids=[int(i) for i in file_ids],
                    confirm=confirm,
                    safety_token=safety_token,
                    moved_by=moved_by,
                    source_id=source_id,
                )
            except ValueError as e:
                # Confirm/token/cap failures are 400 — not 500.
                raise HTTPException(400, str(e))
            except RuntimeError as e:
                # Kill-switch off → 403 so the UI can show a clear hint.
                raise HTTPException(403, str(e))
        return result.to_dict()

    # --- QUARANTINE BROWSER + LIFECYCLE (issue #110 Phase 2) ---
    # Read-only listing + manual purge + restore endpoints. The daily
    # auto-purge job runs from task_scheduler.py — these handlers just
    # let operators inspect / act on individual rows from the UI.

    @app.get("/api/quarantine")
    async def quarantine_list(
        status: Optional[str] = Query(None),
        limit: int = Query(500, ge=1, le=5000),
    ):
        """List quarantine_log rows with derived ``status`` and
        ``will_purge_at`` fields. ``status`` filter is one of
        ``quarantined`` | ``restored`` | ``purged`` | ``all``."""
        dup_cfg = ((config or {}).get("duplicates") or {}).get(
            "quarantine"
        ) or {}
        try:
            qdays = max(1, int(dup_cfg.get("quarantine_days") or 30))
        except (TypeError, ValueError):
            qdays = 30
        sql = (
            "SELECT id, file_id, original_path, quarantine_path, sha256, "
            "file_size, moved_at, moved_by, gain_report_id, "
            "purged_at, restored_at "
            "FROM quarantine_log "
        )
        where: list = []
        params: list = []
        if status == "quarantined":
            where.append("purged_at IS NULL AND restored_at IS NULL")
        elif status == "restored":
            where.append("restored_at IS NOT NULL")
        elif status == "purged":
            where.append("purged_at IS NOT NULL")
        if where:
            sql += "WHERE " + " AND ".join(where) + " "
        sql += "ORDER BY moved_at DESC LIMIT ?"
        params.append(int(limit))
        rows: list = []
        with db.get_cursor() as cur:
            cur.execute(sql, params)
            for r in cur.fetchall():
                d = dict(r)
                # Derive status + will_purge_at for the UI.
                if d.get("purged_at"):
                    d["status"] = "purged"
                elif d.get("restored_at"):
                    d["status"] = "restored"
                else:
                    d["status"] = "quarantined"
                moved_at = d.get("moved_at")
                will_purge: Optional[str] = None
                if moved_at:
                    try:
                        if isinstance(moved_at, str):
                            dt = datetime.fromisoformat(
                                moved_at.replace("Z", "")
                            )
                        else:
                            dt = moved_at
                        will_purge = (dt + timedelta(days=qdays)).strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                    except Exception:
                        will_purge = None
                d["will_purge_at"] = will_purge
                rows.append(d)
        return {
            "enabled": bool(dup_cfg.get("enabled", True)),
            "quarantine_days": qdays,
            "rows": rows,
        }

    @app.post("/api/quarantine/{quarantine_log_id}/purge")
    async def quarantine_purge(quarantine_log_id: int, request: Request):
        """Manual hard-delete. Body: {confirm: true, safety_token: "PURGE"}.

        SHA-256 mismatch = forensic preserve, never delete (returns 409).
        """
        from src.archiver.duplicate_cleaner import (
            DuplicateCleaner, PURGE_SAFETY_TOKEN_VALUE,
        )
        body = await request.json()
        confirm = bool(body.get("confirm", False))
        safety_token = body.get("safety_token", "")
        purged_by = body.get("purged_by") or "operator"
        if not confirm:
            raise HTTPException(400, "confirm=True required to purge")
        # Placeholder for future #112 approval-list wiring — for now we
        # only enforce the safety_token. The token MUST equal "PURGE".
        if safety_token != PURGE_SAFETY_TOKEN_VALUE:
            raise HTTPException(
                400,
                f"safety_token must equal {PURGE_SAFETY_TOKEN_VALUE!r}",
            )
        cleaner = DuplicateCleaner(db, config)
        result = cleaner.purge_one(int(quarantine_log_id), purged_by=purged_by)
        if result.status in ("purged", "skipped_missing"):
            return result.to_dict()
        if result.status == "skipped_not_found":
            raise HTTPException(404, result.reason or "not found")
        if result.status == "abort_sha_mismatch":
            # 409 Conflict: file integrity mismatch, refused — operator
            # must triage.
            raise HTTPException(
                409,
                {
                    "error": "sha_mismatch_forensic_preserve",
                    "detail": result.to_dict(),
                },
            )
        if result.status in (
            "skipped_already_purged", "skipped_restored",
        ):
            raise HTTPException(409, result.reason or result.status)
        # status == "error" — bubble as 500 with audit-friendly body.
        raise HTTPException(
            500,
            {"error": result.status, "detail": result.to_dict()},
        )

    @app.post("/api/quarantine/{quarantine_log_id}/restore")
    async def quarantine_restore(quarantine_log_id: int, request: Request):
        """Restore from quarantine. Body: {confirm: true}."""
        from src.archiver.duplicate_cleaner import DuplicateCleaner
        body = await request.json()
        confirm = bool(body.get("confirm", False))
        restored_by = body.get("restored_by") or "operator"
        if not confirm:
            raise HTTPException(400, "confirm=True required to restore")
        cleaner = DuplicateCleaner(db, config)
        result = cleaner.restore(
            int(quarantine_log_id), restored_by=restored_by
        )
        if result.status == "restored":
            return result.to_dict()
        if result.status == "skipped_not_found":
            raise HTTPException(404, result.reason or "not found")
        if result.status in (
            "skipped_collision",
            "skipped_already_restored",
            "skipped_already_purged",
            "skipped_missing",
        ):
            raise HTTPException(409, result.reason or result.status)
        raise HTTPException(
            500, {"error": result.status, "detail": result.to_dict()},
        )

    # --- OPERATIONS HISTORY (gain reports) ---
    # Read-only listing of gain_reports rows (any operation). Used by the
    # "Islem Gecmisi" page to render before/after/delta panels.

    @app.get("/api/operations/history")
    async def operations_history(limit: int = Query(50, ge=1, le=500),
                                  operation: Optional[str] = None):
        from src.storage.gain_reporter import GainReporter
        reporter = GainReporter(db, config)
        return {"reports": reporter.list_reports(
            limit=limit, operation=operation
        )}

    @app.get("/api/operations/{op_id}")
    async def operation_detail(op_id: int):
        from src.storage.gain_reporter import GainReporter
        reporter = GainReporter(db, config)
        report = reporter.get_report(op_id)
        if report is None:
            raise HTTPException(404, "Operation report not found")
        return report

    @app.post("/api/archive/selective")
    async def archive_selective(request: Request):
        """Secili dosyalari arsivle (duplicate cleanup icin)."""
        from src.archiver.archive_engine import ArchiveEngine
        from src.utils.size_formatter import format_size

        body = await request.json()
        source_id = body.get("source_id")
        file_ids = body.get("file_ids", [])

        if not source_id or not file_ids:
            raise HTTPException(400, "source_id ve file_ids gerekli")

        # Issue #158 (C-2) — confirm gate (mirrors /api/archive/run).
        config_dry_run = bool(
            (config or {}).get("archiving", {}).get("dry_run", True)
        )
        body_dry_run = body.get("dry_run")
        effective_dry_run = (
            bool(body_dry_run) if body_dry_run is not None else config_dry_run
        )
        confirm = bool(body.get("confirm", False))
        if (not effective_dry_run) and not confirm:
            raise HTTPException(
                400,
                "Real archive run requires confirm=true (issue #158 C-2). "
                "Re-submit with dry_run=false + confirm=true after "
                "reviewing the selection.",
            )

        # Kaynak bilgisi
        # NOTE: ``db.get_source(...)`` did not exist — historical typo.
        # ``get_source_by_id`` returns a ``Source`` dataclass (attribute
        # access, not dict access). See latent-bug fix bundled with
        # security-audit-2026-04-28.
        source = db.get_source_by_id(source_id)
        if not source:
            raise HTTPException(404, "Kaynak bulunamadi")

        archive_dest = source.archive_dest
        if not archive_dest:
            raise HTTPException(400, "Arsiv hedefi tanimli degil")

        # Secili dosyalari veritabanindan al
        files = []
        with db.get_cursor() as cur:
            placeholders = ','.join('?' * len(file_ids))
            cur.execute(f"""
                SELECT * FROM scanned_files WHERE id IN ({placeholders}) AND source_id=?
            """, file_ids + [source_id])
            files = [dict(r) for r in cur.fetchall()]

        if not files:
            raise HTTPException(404, "Secili dosyalar bulunamadi")

        engine = ArchiveEngine(db, config)
        result = engine.archive_files(
            files, archive_dest, source.unc_path, source_id,
            archived_by="duplicate_cleanup",
            trigger_type="manual",
            trigger_detail="duplicate_cleanup",
            dry_run=effective_dry_run,
        )
        return result

    # --- BUYUME ANALIZI ---

    @app.get("/api/growth/{source_id}")
    async def growth_stats(source_id: int):
        """Yillik/aylik/gunluk buyume istatistikleri."""
        stats = None
        if analytics.available:
            try:
                stats = analytics.get_growth_stats(source_id)
            except Exception as e:
                logger.warning("DuckDB growth sorgusu basarisiz, SQLite fallback: %s", e)
        if stats is None:
            stats = db.get_growth_stats(source_id)
        creators = db.get_top_file_creators(source_id)
        stats["top_creators"] = creators
        return stats

    @app.get("/api/reports/top-creators/{source_id}")
    async def top_creators(source_id: int, limit: int = 20):
        """En cok dosya olusturan kullanicilar."""
        return db.get_top_file_creators(source_id, limit=limit)

    # --- ARCHIVE BY INSIGHT API ---

    @app.post("/api/archive/by-insight")
    async def archive_by_insight(request):
        """Archive files based on AI insight recommendation."""
        from starlette.requests import Request
        body = await request.json()
        insight_type = body.get("type")  # "stale_1year", "stale_3year", "temp_files", "large_files", "duplicates"
        source_id = body.get("source_id")
        confirm = body.get("confirm", False)

        if not insight_type or not source_id:
            raise HTTPException(400, "type ve source_id zorunlu")

        src = _get_source(db, source_id)
        if not src.archive_dest:
            raise HTTPException(400, "Arsiv hedefi tanimli degil")

        scan_id = db.get_latest_scan_id(src.id, include_running=True)
        if not scan_id:
            raise HTTPException(400, "Tarama verisi bulunamadi")

        from src.utils.size_formatter import format_size

        # Build query based on insight_type
        with db.get_cursor() as cur:
            if insight_type == "stale_1year":
                cur.execute("""
                    SELECT * FROM scanned_files WHERE source_id=? AND scan_id=?
                    AND julianday('now') - julianday(last_access_time) > 365
                    ORDER BY last_access_time ASC LIMIT 10000
                """, (source_id, scan_id))
            elif insight_type == "stale_3year":
                cur.execute("""
                    SELECT * FROM scanned_files WHERE source_id=? AND scan_id=?
                    AND julianday('now') - julianday(last_access_time) > 1095
                    ORDER BY last_access_time ASC LIMIT 10000
                """, (source_id, scan_id))
            elif insight_type == "temp_files":
                cur.execute("""
                    SELECT * FROM scanned_files WHERE source_id=? AND scan_id=?
                    AND (LOWER(extension) IN ('tmp','temp','bak','old','log','cache')
                         OR file_name LIKE '~$%' OR file_name LIKE '%.tmp')
                    ORDER BY file_size DESC LIMIT 10000
                """, (source_id, scan_id))
            elif insight_type == "large_files":
                cur.execute("""
                    SELECT * FROM scanned_files WHERE source_id=? AND scan_id=?
                    AND file_size > 104857600
                    AND julianday('now') - julianday(last_access_time) > 180
                    ORDER BY file_size DESC LIMIT 10000
                """, (source_id, scan_id))
            elif insight_type == "duplicates":
                # Get duplicate groups first
                cur.execute("""
                    SELECT file_name, file_size FROM scanned_files
                    WHERE source_id=? AND scan_id=? AND file_size > 1048576
                    GROUP BY file_name, file_size HAVING COUNT(*) > 1
                """, (source_id, scan_id))
                dup_groups = [dict(r) for r in cur.fetchall()]
                # Get all but newest of each duplicate group
                files = []
                for grp in dup_groups[:100]:
                    cur.execute("""
                        SELECT * FROM scanned_files
                        WHERE source_id=? AND scan_id=? AND file_name=? AND file_size=?
                        ORDER BY last_modify_time DESC
                    """, (source_id, scan_id, grp["file_name"], grp["file_size"]))
                    group_files = [dict(r) for r in cur.fetchall()]
                    if len(group_files) > 1:
                        files.extend(group_files[1:])  # Keep newest, archive rest
                if not confirm:
                    total_size = sum(f["file_size"] for f in files)
                    return {
                        "preview": True,
                        "file_count": len(files),
                        "total_size": total_size,
                        "total_size_formatted": format_size(total_size),
                        "sample": [{"file_path": f["file_path"], "file_name": f["file_name"], "file_size": f["file_size"]} for f in files[:20]]
                    }
                # For confirm, use collected files
                from src.archiver.archive_engine import ArchiveEngine
                engine = ArchiveEngine(db, config)
                return engine.archive_files(
                    files, src.archive_dest, src.unc_path, src.id,
                    archived_by=f"ai_insight:{insight_type}",
                    trigger_type='ai_insight', trigger_detail=insight_type
                )
            else:
                raise HTTPException(400, f"Gecersiz insight type: {insight_type}")

            files = [dict(r) for r in cur.fetchall()] if insight_type != "duplicates" else []

        if insight_type == "duplicates":
            return {"preview": True, "file_count": 0, "total_size": 0, "total_size_formatted": "0 B", "sample": []}

        if not files:
            return {"preview": True, "file_count": 0, "total_size": 0, "total_size_formatted": "0 B", "sample": []}

        total_size = sum(f["file_size"] for f in files)

        if not confirm:
            return {
                "preview": True,
                "file_count": len(files),
                "total_size": total_size,
                "total_size_formatted": format_size(total_size),
                "sample": [{"file_path": f["file_path"], "file_name": f["file_name"], "file_size": f["file_size"]} for f in files[:20]]
            }

        # Confirmed - execute archive
        from src.archiver.archive_engine import ArchiveEngine
        engine = ArchiveEngine(db, config)
        return engine.archive_files(
            files, src.archive_dest, src.unc_path, src.id,
            archived_by=f"ai_insight:{insight_type}",
            trigger_type='ai_insight', trigger_detail=insight_type
        )

    # --- ARCHIVE OPERATIONS API ---

    @app.get("/api/archive/operations")
    async def get_operations(source_id: int = None, limit: int = 50):
        return db.get_archive_operations(source_id, limit)

    @app.get("/api/archive/operations/{op_id}")
    async def get_operation_detail(op_id: int):
        result = db.get_archive_operation_detail(op_id)
        if not result:
            raise HTTPException(404, "Islem bulunamadi")
        return result

    @app.post("/api/restore/by-operation/{op_id}")
    async def restore_by_operation(op_id: int):
        """Bir arsiv operasyonundaki TUM dosyalari geri yukle."""
        from src.archiver.restore_engine import RestoreEngine

        operation = db.get_archive_operation_detail(op_id)
        if not operation:
            raise HTTPException(404, "Islem bulunamadi")

        if operation["operation_type"] != "archive":
            raise HTTPException(400, "Sadece arsiv islemleri geri yuklenebilir")

        # Restore operation kaydi olustur
        restore_op_id = None
        try:
            restore_op_id = db.create_archive_operation(
                'restore', operation.get("source_id"),
                'manual', f'Restore of operation #{op_id}'
            )
        except Exception:
            pass

        engine = RestoreEngine(db)
        restored = 0
        failed = 0
        total_size = 0
        errors = []

        # Islem zamanindaki arsivlenmis dosyalari getir
        files = operation.get("files", [])
        if not files:
            # Fallback: files_json'dan al
            import json
            fj = operation.get("files_json")
            if fj and isinstance(fj, str):
                try:
                    files = json.loads(fj)
                except Exception:
                    pass

        for f in files:
            archive_id = f.get("id")
            if not archive_id:
                continue
            if f.get("restored_at"):
                continue
            result = engine.restore_by_id(archive_id, op_id=restore_op_id)
            if result.get("success"):
                restored += 1
                total_size += f.get("file_size", 0)
            else:
                failed += 1
                errors.append({"id": archive_id, "error": result.get("error", "")})

        # Tamamla
        if restore_op_id:
            try:
                status = 'completed' if failed == 0 else 'partial'
                db.complete_archive_operation(restore_op_id, restored, total_size, status)
            except Exception:
                pass

        return {
            "restored": restored,
            "failed": failed,
            "total_size": total_size,
            "errors": errors[:10],
            "operation_id": restore_op_id
        }

    # --- TOPLU GERI YUKLEME ---

    @app.post("/api/restore/bulk")
    async def bulk_restore(request):
        """Toplu geri yukleme - onizleme veya gercek."""
        from src.archiver.restore_engine import RestoreEngine
        from src.utils.size_formatter import format_size

        body = await request.json()
        archive_ids = body.get("archive_ids", [])
        confirm = body.get("confirm", False)

        if not archive_ids:
            raise HTTPException(400, "archive_ids gerekli")

        engine = RestoreEngine(db)

        if not confirm:
            # Onizleme
            preview = engine.preview_restore(archive_ids)
            preview["total_size_formatted"] = format_size(preview.get("total_size", 0))
            return {"preview": True, **preview}
        else:
            # Gercek geri yukleme
            # source_id al (ilk dosyadan)
            first = db.get_archived_file_by_id(archive_ids[0])
            source_id = first.get("source_id") if first else None
            result = engine.bulk_restore(archive_ids, source_id)
            result["total_size_formatted"] = format_size(result.get("total_size", 0))
            return result

    @app.get("/api/archive/browse")
    async def browse_archived(source_id: int = None,
                               page: int = Query(1, ge=1, le=10000),
                               page_size: int = Query(50, ge=1, le=500)):
        """Arsivlenmis dosyalara goz at (geri yukleme UI icin)."""
        from src.utils.size_formatter import format_size
        result = db.search_archived_files("", page=page, page_size=page_size)
        for f in result.get("results", []):
            f["file_size_formatted"] = format_size(f.get("file_size", 0))
        return result

    # --- ARSIV GECMISI ---

    @app.get("/api/archive/history")
    async def archive_history(source_id: int = None,
                              page: int = Query(1, ge=1, le=10000),
                              page_size: int = Query(20, ge=1, le=500),
                              date_from: str = None,
                              date_to: str = None, op_type: str = None):
        """Sayfalanmis arsiv islem gecmisi."""
        from src.utils.size_formatter import format_size
        result = db.get_archive_history(source_id, page, page_size,
                                        date_from, date_to, op_type)
        for op in result["operations"]:
            op["total_size_formatted"] = format_size(op.get("total_size") or 0)
        return result

    @app.get("/api/archive/operations/{op_id}/files")
    async def operation_files(op_id: int,
                               page: int = Query(1, ge=1, le=10000),
                               page_size: int = Query(100, ge=1, le=500)):
        """Arsiv islemindeki dosyalari sayfalanmis getir."""
        from src.utils.size_formatter import format_size
        result = db.get_archive_operation_files(op_id, page, page_size)
        for f in result.get("files", []):
            f["file_size_formatted"] = format_size(f.get("file_size") or 0)
        return result

    # --- SYSTEM API ---

    @app.post("/api/system/open-folder")
    async def open_folder(request: Request):
        """Dizini Windows Explorer'da ac (yalnizca yerel istemci icin)."""
        body = await request.json()
        client_host = (request.client.host if request.client else "")
        return open_folder_impl(body, client_host)

    @app.get("/api/system/list-dir")
    async def list_dir(
        request: Request,
        path: str = "",
        show_hidden: bool = False,
    ):
        """Klasor tarayici icin dizin icerigini listele (sadece localhost).

        Bos ``path`` -> mantiksal kokleri dondurur (Windows surucu harfleri
        veya POSIX "/"). 5000 girisle sinirlanmis, 'dir'>'file' siralanmis.
        Uzaktan istemci icin 403; var olmayan yol icin 404.
        """
        client_host = (request.client.host if request.client else "")
        return list_dir_impl(path, client_host, show_hidden=show_hidden)

    @app.get("/api/system/health")
    async def health():
        # WAL ve toplam disk durumu
        wal_warning = None
        try:
            wal_path = db.db_path + "-wal"
            wal_size = os.path.getsize(wal_path) if os.path.exists(wal_path) else 0
            if wal_size > 500_000_000:
                wal_warning = {
                    "wal_size_bytes": wal_size,
                    "wal_size_formatted": f"{wal_size / 1073741824:.2f} GB"
                                          if wal_size > 1073741824
                                          else f"{wal_size / 1048576:.0f} MB",
                    "severity": "critical" if wal_size > 5_000_000_000 else "warning",
                    "recommendation": "Kaynaklar -> Veritabani Bakimi -> Optimize Et butonuna basarak "
                                       "WAL'i temizleyin.",
                }
        except Exception:
            pass

        # Issue #64: surface which PII regex backend is in effect
        # (hyperscan when available + opted-in, else stdlib re).
        try:
            from src.compliance._pii_backends import (
                hyperscan_available, hyperscan_version,
            )
            pii_cfg = (
                ((config or {}).get("compliance", {}) or {}).get("pii", {}) or {}
            )
            pii_engine_pref = pii_cfg.get("engine", "auto")
            hs_ok = hyperscan_available()
            if pii_engine_pref == "re":
                effective = "re"
            elif pii_engine_pref == "hyperscan":
                effective = "hyperscan" if hs_ok else "re"
            else:  # auto
                effective = "hyperscan" if hs_ok else "re"
            pii_backend_info = {
                "backend": effective,
                "configured": pii_engine_pref,
                "hyperscan_available": hs_ok,
                "version": hyperscan_version() if effective == "hyperscan" else None,
            }
        except Exception:
            pii_backend_info = {"backend": "re", "configured": "auto",
                                 "hyperscan_available": False, "version": None}

        # Issue #113: surface total disk capacity so the forecast page can
        # compute the default 85% threshold without a second roundtrip.
        # shutil.disk_usage() works on the volume hosting the SQLite DB —
        # which is the same volume holding ``data/`` (scans, snapshots,
        # quarantine), i.e. the resource we care about for capacity alarms.
        disk_info = None
        try:
            import shutil as _shutil
            target = os.path.dirname(os.path.abspath(db.db_path)) or "."
            usage = _shutil.disk_usage(target)
            disk_info = {
                "path": target,
                "total_bytes": int(usage.total),
                "used_bytes": int(usage.used),
                "free_bytes": int(usage.free),
            }
        except Exception as e:  # pragma: no cover - non-critical
            logger.debug("disk_usage probe failed: %s", e)

        return {
            "status": "ok",
            "time": datetime.now().isoformat(),
            "version": APP_VERSION,
            "database": db.health_check(),
            "analytics": analytics.health(),
            "email": email_notifier.health(),
            "wal_warning": wal_warning,
            "pii_backend": pii_backend_info,
            "disk": disk_info,
        }

    @app.get("/api/system/analytics")
    async def analytics_status():
        """DuckDB analitik motor durumunu dondur."""
        return analytics.health()

    # --- NOTIFICATIONS API ---

    class TestEmailRequest(BaseModel):
        to: str
        username: Optional[str] = "test"
        display_name: Optional[str] = None

    @app.get("/api/notifications/status")
    async def notifications_status():
        """SMTP konfigurasyon ozeti + canli bind testi."""
        info = email_notifier.health()
        if email_notifier.available:
            info["probe"] = email_notifier.test_connection()
        return info

    @app.post("/api/notifications/test")
    async def notifications_test(payload: TestEmailRequest):
        """Belirtilen adrese kucuk bir dogrulama e-postasi gonder.

        Kullanim: SMTP ayarlarini dogrulamak icin. Gercek skor degil,
        sabit ornek skor verisi gonderilir. Admin CC varsa CC'ye eklenir.
        """
        if not email_notifier.available:
            raise HTTPException(400, f"SMTP kullanilamaz: {email_notifier._init_error}")
        fake_score = {
            "score": 85, "grade": "B", "total_penalty": 15,
            "factors": [
                {"name": "stale_files", "label": "Test: 1+ yildir erisilmeyen dosyalar",
                 "count": 42, "penalty": 8, "max": 30},
                {"name": "oversized_files", "label": "Test: 100 MB'dan buyuk dosyalar",
                 "count": 7, "penalty": 7, "max": 15},
            ],
            "non_compliance": {"stale_files": 42, "oversized_files": 7,
                                "naming_violations": 0, "duplicate_files": 0,
                                "dormant": False},
            "suggestions": [
                "Bu bir TEST mesajidir. SMTP ayarlariniz dogru calisiyor.",
                "Gercek rapor zamanli bildirim zamanlayicisi (PR D) ile gelecek.",
            ],
            "total_files": 100, "total_size": 5 * 1024 * 1024 * 1024,
        }
        result = email_notifier.send_user_report(
            username=payload.username or "test",
            email=payload.to,
            score_result=fake_score,
            display_name=payload.display_name,
        )
        if not result.get("ok"):
            raise HTTPException(500, result.get("error") or "Gonderim basarisiz")
        return result

    @app.post("/api/notifications/send-to/{username}")
    async def notifications_send_to(username: str):
        """Tek kullaniciya gercek verimlilik raporu gonder.

        Kullanici e-postasi AD'den cozulur; yoksa 400 doner.
        """
        if not email_notifier.available:
            raise HTTPException(400, f"SMTP kullanilamaz: {email_notifier._init_error}")

        # AD lookup opsiyonel; ADLookup mevcutsa kullan
        ad_info = None
        try:
            ad = getattr(app.state, "ad_lookup", None)
            if ad and ad.available:
                ad_info = ad.lookup(username)
        except Exception:
            ad_info = None

        if not ad_info or not ad_info.get("email"):
            raise HTTPException(404, f"{username} icin e-posta bulunamadi (AD lookup sonucu yok)")

        from src.user_activity.efficiency_score import compute_user_score
        score = compute_user_score(db, username)
        result = email_notifier.send_user_report(
            username=username,
            email=ad_info["email"],
            score_result=score,
            display_name=ad_info.get("display_name"),
        )
        if not result.get("ok"):
            raise HTTPException(500, result.get("error") or "Gonderim basarisiz")
        return {"sent_to": ad_info["email"], "cc": result.get("cc"), "score": score["score"]}

    @app.get("/api/notifications/log")
    async def notifications_log(username: Optional[str] = None,
                                 status: Optional[str] = None,
                                 page: int = Query(1, ge=1, le=10000),
                                 page_size: int = Query(50, ge=1, le=500)):
        """Gonderim log listesi (sayfalanmis)."""
        offset = (page - 1) * page_size
        conditions = []
        params: list = []
        if username:
            conditions.append("username = ?")
            params.append(username)
        if status:
            conditions.append("status = ?")
            params.append(status)
        where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
        with db.get_cursor() as cur:
            cur.execute(f"SELECT COUNT(*) AS cnt FROM notification_log{where}", params)
            total = cur.fetchone()["cnt"]
            cur.execute(
                f"""SELECT id, username, email, cc, subject, status, error, sent_at
                    FROM notification_log{where}
                    ORDER BY sent_at DESC LIMIT ? OFFSET ?""",
                params + [page_size, offset],
            )
            rows = [dict(r) for r in cur.fetchall()]
        return {"total": total, "rows": rows, "page": page, "page_size": page_size}

    @app.get("/api/system/version")
    async def version():
        """Calisan sürüm (repo kokundeki VERSION dosyasindan)."""
        return {"version": APP_VERSION}

    @app.get("/api/system/status")
    async def system_status():
        """Issue #125 — list currently running background operations.

        Always 200; returns ``{"operations": []}`` when nothing is
        active. Reads in-memory only — no DB access — so the dashboard
        can poll every 5 seconds without measurable cost. If the
        registry is missing (older app.state, hot-reload edge case)
        the endpoint silently returns an empty list rather than 500.
        """
        registry = getattr(app.state, "operations", None)
        if registry is None:
            return {"operations": []}
        try:
            ops = [op.to_public_dict() for op in registry.list_active()]
        except Exception as e:  # pragma: no cover - defensive only
            logger.warning("operations registry list_active failed: %s", e)
            ops = []
        return {"operations": ops}

    @app.get("/api/system/ad/status")
    async def ad_status():
        """AD/LDAP baglanti durumu + bind testi."""
        return ad_lookup.health()

    @app.get("/api/users/{username}/ad-info")
    async def user_ad_info(username: str, refresh: bool = False):
        """Kullanici adindan e-posta + display name cozumle.

        AD devre disi veya erisilmez ise cache'ten doner; cache yoksa
        null alanlarla sessizce doner.
        """
        info = ad_lookup.lookup(username, force_refresh=refresh)
        if info is None:
            return {
                "username": username,
                "email": None,
                "display_name": None,
                "found": False,
                "source": None,
            }
        return info

    def _parse_ver(v: str):
        """Semver-ish karsilastirma icin tuple parse et.

        '1.6.0'      -> (1,6,0, False) — stabil
        '1.7.0-dev'  -> (1,7,0, True)  — pre-release (master'da, release'ten onde)
        """
        v = v.strip().lstrip("v")
        main, _, suffix = v.partition("-")
        parts = main.split(".")
        try:
            nums = tuple(int(p) for p in parts)
        except ValueError:
            nums = (0,)
        return nums + (bool(suffix),)

    @app.get("/api/system/version-check")
    async def version_check():
        """GitHub latest release ile yerel sürümü karsilastir.

        Kurumsal aglarda GitHub API'sine erisim yoksa hata yerine
        `error` alani ile temiz doner — UI banner'i sessizce gizlenir.
        """
        import urllib.request
        import urllib.error
        import ssl

        api_url = "https://api.github.com/repos/deepdarbe/FILE_ACTIVITY/releases/latest"
        local = APP_VERSION
        try:
            req = urllib.request.Request(api_url, headers={"User-Agent": "file-activity-dashboard"})
            # TLS dogrulama: kurumsal proxy arkasinda ilerleyebilsin diye
            # sistem CA store'u kullan; basarisiz olursa sessizce dus
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            remote = (data.get("tag_name") or "").lstrip("v")
            if not remote:
                return {"local": local, "remote": None, "update_available": False,
                        "error": "no tag_name"}
            local_tuple = _parse_ver(local)
            remote_tuple = _parse_ver(remote)
            # Pre-release (suffix=True) kendi base'iyle esitse ilerdedir, update yok
            local_core = local_tuple[:-1]
            remote_core = remote_tuple[:-1]
            is_pre = local_tuple[-1]
            update_available = (local_core < remote_core) and not (is_pre and local_core == remote_core)
            return {
                "local": local,
                "remote": remote,
                "update_available": update_available,
                "release_url": data.get("html_url"),
                "published_at": data.get("published_at"),
            }
        except urllib.error.URLError as e:
            return {"local": local, "remote": None, "update_available": False,
                    "error": f"github API erisilemedi: {e.reason}"}
        except (ssl.SSLError, TimeoutError, OSError) as e:
            return {"local": local, "remote": None, "update_available": False,
                    "error": f"ag hatasi: {e}"}
        except Exception as e:
            logger.warning("version-check beklenmeyen hata: %s", e)
            return {"local": local, "remote": None, "update_available": False,
                    "error": "beklenmeyen hata"}

    @app.post("/api/system/update")
    async def update():
        """Yerel update.cmd'yi detached olarak baslat.

        update.cmd setup-source.ps1'i calistirir, bu da calisan python
        process'ini durdurur ve master'in en son halini indirir. Dashboard
        30 saniye sonra yeniden baslatilir.
        """
        import subprocess
        import sys

        if sys.platform != "win32":
            raise HTTPException(400, "Guncelleme sadece Windows uzerinde desteklenir")

        # update.cmd'yi cwd veya kurulum dizininde ara
        candidates = [
            os.path.join(os.getcwd(), "update.cmd"),
            r"C:\FileActivity\update.cmd",
        ]
        update_cmd = next((p for p in candidates if os.path.exists(p)), None)
        if not update_cmd:
            raise HTTPException(
                404,
                f"update.cmd bulunamadi. Kontrol edilen yollar: {candidates}. "
                "Manuel guncelleme icin setup-source.ps1'i tekrar calistirin."
            )

        # Detached process: dashboard'u hemen donduruyor, update.cmd
        # kendi console'unda devam ediyor. CREATE_NEW_CONSOLE ile kullanici
        # ilerlemeyi gorebilsin.
        DETACHED_PROCESS = 0x00000008
        CREATE_NEW_CONSOLE = 0x00000010
        CREATE_NEW_PROCESS_GROUP = 0x00000200

        subprocess.Popen(
            ["cmd.exe", "/c", update_cmd],
            creationflags=DETACHED_PROCESS | CREATE_NEW_CONSOLE | CREATE_NEW_PROCESS_GROUP,
            close_fds=True,
            cwd=os.path.dirname(update_cmd),
        )
        logger.info("Guncelleme tetiklendi: %s", update_cmd)
        return {
            "status": "started",
            "message": "Guncelleme baslatildi. 30-60 saniye sonra dashboard'u yenileyin.",
            "update_cmd": update_cmd,
        }

    @app.get("/api/db/stats")
    async def db_stats():
        """Veritabani istatistikleri."""
        if analytics.available:
            try:
                wal_path = db.db_path + "-wal"
                shm_path = db.db_path + "-shm"
                tables = ["scanned_files", "scan_runs", "archived_files",
                          "user_access_logs", "sources"]
                return analytics.get_db_stats(tables, db.db_path, wal_path, shm_path)
            except Exception as e:
                logger.warning("DuckDB db_stats basarisiz, SQLite fallback: %s", e)
        return db.get_db_stats()

    @app.post("/api/db/cleanup")
    async def db_cleanup(
        request: Request,
        keep_last: Optional[int] = Query(
            default=None, ge=0, le=100,
            description="Son N taramayi koru. 0 = hepsini sil.",
        ),
        keep_last_n_scans: Optional[int] = Query(
            default=None, ge=0, le=100,
            description="Alias of keep_last (config naming).",
        ),
    ):
        """Eski tarama verilerini temizle. Son N taramayi korur.

        Issue #133: customer hit ``?keep_last=0`` and got 422 because
        the previous handler used ``ge=1``. The endpoint now accepts:

        - ``keep_last`` (default 5, range 0..100) — query param, original shape.
        - ``keep_last_n_scans`` — query param alias matching the config / DB
          method parameter name. If both are passed,
          ``keep_last_n_scans`` wins (it's the newer, more explicit
          spelling).
        - JSON body ``{"keep_last_n_scans": N, "confirm": true}`` — M-3
          pattern; ``confirm: true`` is required when using the body form.

        ``keep_last=0`` is a valid request: it deletes every scan_run
        for every source. The audit chain and orphan cleanup branches
        in :py:meth:`Database.cleanup_old_scans` cope with N=0 because
        they use ``OFFSET ?`` which yields the full set when N=0.
        """
        # ── Optional JSON body (M-3 pattern) ─────────────────────────────
        body: dict = {}
        try:
            body = await request.json()
            if not isinstance(body, dict):
                body = {}
        except (ValueError, UnicodeDecodeError):
            # No body, empty body, or non-JSON content-type — fall back to
            # query-param-only mode. ValueError covers json.JSONDecodeError.
            body = {}

        # If a JSON body was provided, enforce the confirm gate (M-3).
        if body:
            if not bool(body.get("confirm", False)):
                raise HTTPException(
                    400, "confirm=true required in request body to run cleanup"
                )
            body_keep = body.get("keep_last_n_scans")
            if body_keep is not None:
                try:
                    body_keep = int(body_keep)
                except (TypeError, ValueError):
                    raise HTTPException(
                        422, "keep_last_n_scans must be an integer (0..100)"
                    )
                if not (0 <= body_keep <= 100):
                    raise HTTPException(
                        422, "keep_last_n_scans must be between 0 and 100"
                    )
                # Body value wins over query params when both present.
                keep_last_n_scans = body_keep

        # ── Resolve effective keep value ─────────────────────────────────
        effective: int
        if keep_last_n_scans is not None:
            effective = int(keep_last_n_scans)
        elif keep_last is not None:
            effective = int(keep_last)
        else:
            effective = 5  # legacy default

        result = db.cleanup_old_scans(keep_last_n=effective)

        # ── Audit log (best-effort, chain-routed per #160) ───────────────
        try:
            db.insert_audit_event_simple(
                source_id=None,
                event_type="db_cleanup",
                username="admin",
                file_path="",
                details=f"keep_last_n={effective}",
                detected_by="dashboard",
            )
        except Exception as _audit_err:  # pragma: no cover - best-effort
            logger.warning("db_cleanup audit yazılamadı: %s", _audit_err)

        return result

    @app.post("/api/db/optimize")
    async def db_optimize():
        """VACUUM + ANALYZE ile veritabanini optimize et."""
        result = db.optimize_database()
        return result

    # ── AD-HOC SQL SORGU PANELI (issue #48) ──
    # Whitelist-guarded read-only DuckDB executor. Frontend "Sorgu" sekmesi
    # buradan beslenir; SQL'i once SqlQueryGuard.validate suzgecinden gecirir,
    # ardindan audit_event yazip DuckDB uzerinden SQLite'a salt-okunur calistirir.

    class QueryRequest(BaseModel):
        sql: str
        max_rows: int = Field(default=1000, ge=1, le=10000)

    @app.post("/api/analytics/query")
    async def analytics_query(req: QueryRequest, request: Request):
        from src.dashboard.sql_query import SqlQueryGuard
        panel_cfg = (config.get("analytics", {}) or {}).get("query_panel", {}) or {}
        if not panel_cfg.get("enabled", True):
            raise HTTPException(403, "Sorgu paneli devre disi")
        max_rows = min(int(req.max_rows), int(panel_cfg.get("max_rows", 10000)))
        guard = SqlQueryGuard(
            max_rows=max_rows,
            timeout_seconds=int(panel_cfg.get("timeout_seconds", 30)),
        )
        ok, reason = guard.validate(req.sql)
        if not ok:
            raise HTTPException(400, f"Sorgu reddedildi: {reason}")
        try:
            db.insert_audit_event_simple(
                source_id=None, event_type="sql_query", username="admin",
                file_path=None, details=req.sql[:500], detected_by="dashboard",
            )
        except Exception as e:  # pragma: no cover - audit best-effort
            logger.warning("sql_query audit yazilamadi: %s", e)
        try:
            return guard.execute(db, req.sql)
        except Exception as e:
            raise HTTPException(500, f"Sorgu hatasi: {e}")

    # ── ARKA PLAN EXPORT SISTEMI ──

    def _export_worker(job_id: str, report_type: str, source_id: int, scan_id: int, params: dict):
        """Arka planda XLS olusturma worker'i."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from src.utils.size_formatter import format_size
        import re as re_mod

        try:
            with _export_lock:
                _export_jobs[job_id]["status"] = "running"
                _export_jobs[job_id]["progress"] = 0

            export_dir = os.path.join(os.path.dirname(db.db_path), "exports")
            os.makedirs(export_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            wb = openpyxl.Workbook()
            ws = wb.active

            # Stiller
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
            alt_fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
            border = Border(
                left=Side(style='thin', color='D0D5DD'),
                right=Side(style='thin', color='D0D5DD'),
                top=Side(style='thin', color='D0D5DD'),
                bottom=Side(style='thin', color='D0D5DD')
            )

            if report_type == "mit_naming":
                ws.title = "MIT Adlandirma Uyumu"
                filename = f"MIT_Naming_{timestamp}.xlsx"
                headers = ["Ihlal Kodu", "Ihlal Turu", "Ciddiyet", "Dosya Adi", "Tam Yol", "Boyut", "Boyut (Okunan)", "Sahip", "Son Degisiklik"]
                checks = {
                    "R1": ("Bosluk Iceren", "Zorunlu", lambda p, n: bool(re_mod.search(r'\s', n))),
                    "R2": ("Ilk Karakter Harf Degil", "Zorunlu", lambda p, n: bool(n) and not re_mod.match(r'^[a-zA-Z]', n)),
                    "R3": ("Yasak Karakter", "Zorunlu", lambda p, n: bool(n) and '.' in n and not re_mod.match(r'^[a-zA-Z0-9._-]+$', n[:n.rfind('.')])),
                    "R4": ("Uzanti Sorunu", "Zorunlu", lambda p, n: '.' not in n or not n.rsplit('.', 1)[-1].isalpha()),
                    "B1": ("Uzun Ad (>31)", "Oneri", lambda p, n: len(n) > 31),
                    "B3": ("Base Nokta", "Oneri", lambda p, n: '.' in n and n[:n.rfind('.')].count('.') > 0),
                    "B4": ("Buyuk Harf", "Oneri", lambda p, n: bool(re_mod.search(r'[A-Z]', n[:n.rfind('.')] if '.' in n else n))),
                    "B5": ("Ayirici Yok", "Oneri", lambda p, n: len(n) > 10 and '_' not in n and '-' not in n),
                }
                for i, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=i, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                row = 2
                with db.get_cursor() as cur:
                    cur.execute("SELECT COUNT(*) as cnt FROM scanned_files WHERE source_id=? AND scan_id=?", (source_id, scan_id))
                    total = cur.fetchone()["cnt"]
                    cur.execute("SELECT file_path, file_name, file_size, owner, last_modify_time FROM scanned_files WHERE source_id=? AND scan_id=?", (source_id, scan_id))
                    processed = 0
                    for rec in cur:
                        processed += 1
                        if processed % 5000 == 0:
                            with _export_lock:
                                _export_jobs[job_id]["progress"] = int(processed * 100 / max(total, 1))
                        for code, (label, severity, fn) in checks.items():
                            if fn(rec["file_path"], rec["file_name"]):
                                ws.cell(row=row, column=1, value=code)
                                ws.cell(row=row, column=2, value=label)
                                ws.cell(row=row, column=3, value=severity)
                                ws.cell(row=row, column=4, value=rec["file_name"])
                                ws.cell(row=row, column=5, value=rec["file_path"])
                                ws.cell(row=row, column=6, value=rec.get("file_size", 0))
                                ws.cell(row=row, column=7, value=format_size(rec.get("file_size", 0)))
                                ws.cell(row=row, column=8, value=rec.get("owner", ""))
                                ws.cell(row=row, column=9, value=rec.get("last_modify_time", ""))
                                if row % 2 == 0:
                                    for c in range(1, 10):
                                        ws.cell(row=row, column=c).fill = alt_fill
                                row += 1

                # Ozet sayfasi
                ws_sum = wb.create_sheet("Ozet", 0)
                ws_sum.cell(row=1, column=1, value="MIT Libraries File Naming Scheme - Uyum Raporu").font = Font(bold=True, size=14)
                ws_sum.cell(row=2, column=1, value=f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                ws_sum.cell(row=3, column=1, value=f"Toplam dosya: {total:,}")
                ws_sum.cell(row=4, column=1, value=f"Ihlal satiri: {row-2:,}")

            elif report_type == "duplicates":
                ws.title = "Kopya Dosyalar"
                filename = f"Duplicates_{timestamp}.xlsx"
                headers = ["Grup", "Dosya Adi", "Boyut", "Boyut (Okunan)", "Tam Yol", "Sahip", "Son Degisiklik"]
                for i, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=i, value=h)
                    cell.font = header_font
                    cell.fill = header_fill

                row = 2
                with db.get_cursor() as cur:
                    cur.execute("""
                        SELECT file_name, file_size, file_path, owner, last_modify_time
                        FROM scanned_files WHERE source_id=? AND scan_id=? AND file_size > 1048576
                        ORDER BY file_name, file_size, file_path
                    """, (source_id, scan_id))
                    prev_key = None
                    group_num = 0
                    for rec in cur:
                        key = (rec["file_name"], rec["file_size"])
                        if key != prev_key:
                            group_num += 1
                            prev_key = key
                        ws.cell(row=row, column=1, value=group_num)
                        ws.cell(row=row, column=2, value=rec["file_name"])
                        ws.cell(row=row, column=3, value=rec["file_size"])
                        ws.cell(row=row, column=4, value=format_size(rec["file_size"]))
                        ws.cell(row=row, column=5, value=rec["file_path"])
                        ws.cell(row=row, column=6, value=rec.get("owner", ""))
                        ws.cell(row=row, column=7, value=rec.get("last_modify_time", ""))
                        row += 1

            elif report_type == "full":
                ws.title = "Tum Dosyalar"
                filename = f"FullReport_{timestamp}.xlsx"
                headers = ["Dosya Adi", "Uzanti", "Boyut", "Boyut (Okunan)", "Tam Yol", "Sahip", "Olusturma", "Son Erisim", "Son Degisiklik"]
                for i, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=i, value=h)
                    cell.font = header_font
                    cell.fill = header_fill

                row = 2
                with db.get_cursor() as cur:
                    cur.execute("SELECT COUNT(*) as cnt FROM scanned_files WHERE source_id=? AND scan_id=?", (source_id, scan_id))
                    total = cur.fetchone()["cnt"]
                    cur.execute("SELECT * FROM scanned_files WHERE source_id=? AND scan_id=? ORDER BY file_path", (source_id, scan_id))
                    processed = 0
                    for rec in cur:
                        processed += 1
                        if processed % 5000 == 0:
                            with _export_lock:
                                _export_jobs[job_id]["progress"] = int(processed * 100 / max(total, 1))
                        ws.cell(row=row, column=1, value=rec["file_name"])
                        ws.cell(row=row, column=2, value=rec.get("extension", ""))
                        ws.cell(row=row, column=3, value=rec.get("file_size", 0))
                        ws.cell(row=row, column=4, value=format_size(rec.get("file_size", 0)))
                        ws.cell(row=row, column=5, value=rec["file_path"])
                        ws.cell(row=row, column=6, value=rec.get("owner", ""))
                        ws.cell(row=row, column=7, value=rec.get("creation_time", ""))
                        ws.cell(row=row, column=8, value=rec.get("last_access_time", ""))
                        ws.cell(row=row, column=9, value=rec.get("last_modify_time", ""))
                        if row % 2 == 0:
                            for c in range(1, 10):
                                ws.cell(row=row, column=c).fill = alt_fill
                        row += 1
            else:
                raise ValueError(f"Bilinmeyen rapor tipi: {report_type}")

            # Sutun genisliklerini ayarla
            for ws_sheet in wb.worksheets:
                for col in range(1, ws_sheet.max_column + 1):
                    max_len = max(len(str(ws_sheet.cell(row=r, column=col).value or "")) for r in range(1, min(ws_sheet.max_row + 1, 100)))
                    ws_sheet.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)

            file_path = os.path.join(export_dir, filename)
            wb.save(file_path)

            with _export_lock:
                _export_jobs[job_id]["status"] = "completed"
                _export_jobs[job_id]["progress"] = 100
                _export_jobs[job_id]["file_path"] = file_path
                _export_jobs[job_id]["file_name"] = filename
                _export_jobs[job_id]["file_size"] = os.path.getsize(file_path)

        except Exception as e:
            logging.getLogger("file_activity").error(f"Export error: {e}")
            with _export_lock:
                _export_jobs[job_id]["status"] = "error"
                _export_jobs[job_id]["error"] = str(e)

    @app.post("/api/export/start")
    async def start_export(report_type: str = Query(...), source_id: int = Query(...)):
        """Arka planda XLS export baslat. Tarama devam ederken bile calisir."""
        # Son tamamlanmis taramayi kullan (aktif tarama kilitlemesin)
        with db.get_cursor() as cur:
            cur.execute("""
                SELECT id FROM scan_runs WHERE source_id=? AND status='completed'
                ORDER BY started_at DESC LIMIT 1
            """, (source_id,))
            completed = cur.fetchone()
            if completed:
                scan_id = completed["id"]
            else:
                # Tamamlanmis yoksa son taramayi al
                scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            raise HTTPException(404, "Tarama bulunamadi")

        job_id = str(uuid.uuid4())[:8]
        with _export_lock:
            _export_jobs[job_id] = {
                "status": "queued",
                "progress": 0,
                "report_type": report_type,
                "source_id": source_id,
                "scan_id": scan_id,
                "created_at": datetime.now().isoformat(),
                "file_path": None,
                "file_name": None,
                "error": None,
            }

        thread = threading.Thread(
            target=_export_worker,
            args=(job_id, report_type, source_id, scan_id, {}),
            daemon=True
        )
        thread.start()
        return {"job_id": job_id, "status": "queued"}

    @app.get("/api/export/status/{job_id}")
    async def export_status(job_id: str):
        """Export is durumu sorgula."""
        with _export_lock:
            job = _export_jobs.get(job_id)
        if not job:
            raise HTTPException(404, "Export isi bulunamadi")
        return {
            "job_id": job_id,
            "status": job["status"],
            "progress": job["progress"],
            "file_name": job.get("file_name"),
            "file_size": job.get("file_size"),
            "error": job.get("error"),
        }

    @app.get("/api/export/download/{job_id}")
    async def export_download(job_id: str):
        """Tamamlanmis export dosyasini indir."""
        with _export_lock:
            job = _export_jobs.get(job_id)
        if not job:
            raise HTTPException(404, "Export isi bulunamadi")
        if job["status"] != "completed" or not job.get("file_path"):
            raise HTTPException(400, "Export henuz tamamlanmadi")
        return FileResponse(
            job["file_path"],
            filename=job["file_name"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.get("/api/export/jobs")
    async def export_jobs():
        """Tum export islerini listele."""
        with _export_lock:
            return [{"job_id": k, **{kk: vv for kk, vv in v.items() if kk != "file_path"}} for k, v in _export_jobs.items()]

    # --- RANSOMWARE DETECTOR API (issue #37) ---

    def _get_detector():
        det = getattr(app.state, "ransomware", None)
        if det is None:
            raise HTTPException(503, "Ransomware detector kullanilamiyor")
        return det

    @app.get("/api/security/ransomware/alerts")
    async def list_ransomware_alerts(since_minutes: int = Query(60, ge=1, le=10080)):
        """Son N dakikadaki ransomware uyarilarini listele (yeni once)."""
        det = _get_detector()
        return det.get_active_alerts(since_minutes=since_minutes)

    @app.post("/api/security/ransomware/alerts/{alert_id}/acknowledge")
    async def acknowledge_ransomware_alert(alert_id: int, by_user: str = "admin"):
        """Bir uyariyi onayla — acknowledged_at + acknowledged_by yazilir."""
        with db.get_cursor() as cur:
            cur.execute(
                """UPDATE ransomware_alerts
                   SET acknowledged_at = datetime('now','localtime'),
                       acknowledged_by = ?
                   WHERE id = ?""",
                (by_user, alert_id),
            )
            if cur.rowcount == 0:
                raise HTTPException(404, f"Uyari bulunamadi (ID: {alert_id})")
        return {"acknowledged": True, "id": alert_id, "by": by_user}

    @app.post("/api/security/ransomware/canaries/{source_id}/deploy")
    async def deploy_ransomware_canaries(source_id: int):
        """Kaynagin paylasim koküne canary dosyalarini birak."""
        src = _get_source(db, source_id)
        det = _get_detector()
        placed = det.deploy_canaries(src.id, src.unc_path)
        return {
            "source_id": src.id,
            "share_root": src.unc_path,
            "placed": placed,
            "canary_names": sorted(det.canary_names),
        }

    @app.post("/api/security/ransomware/test")
    async def ransomware_test(source_id: Optional[int] = None,
                                username: str = "test_user",
                                rule: str = "rename_velocity"):
        """Sentetik olay enjeksiyonu — kural + e-posta + SMB kill (dry-run)
        boru hattini hizlica dogrulamak icin. Production'da kullanilmamali.

        rule: 'rename_velocity' | 'risky_extension' | 'mass_deletion' | 'canary_access'
        """
        det = _get_detector()
        if source_id is not None:
            _get_source(db, source_id)

        rule = (rule or "").strip().lower()
        last_alert = None
        if rule == "rename_velocity":
            for i in range(det.rename_threshold + 1):
                last_alert = det.consume_event({
                    "source_id": source_id,
                    "username": username,
                    "file_path": f"/test/synth_{i}.bin",
                    "old_path": f"/test/synth_{i}.txt",
                    "event_type": "rename",
                }) or last_alert
        elif rule == "risky_extension":
            last_alert = det.consume_event({
                "source_id": source_id,
                "username": username,
                "file_path": "/test/secret.docx.encrypted",
                "event_type": "modify",
            })
        elif rule == "mass_deletion":
            for i in range(det.delete_threshold + 1):
                last_alert = det.consume_event({
                    "source_id": source_id,
                    "username": username,
                    "file_path": f"/test/del_{i}.bin",
                    "event_type": "delete",
                }) or last_alert
        elif rule == "canary_access":
            canary = sorted(det.canary_names)[0] if det.canary_names else "_AAAA_canary_DO_NOT_DELETE.txt"
            last_alert = det.consume_event({
                "source_id": source_id,
                "username": username,
                "file_path": f"/test/{canary}",
                "event_type": "access",
            })
        else:
            raise HTTPException(400, f"Bilinmeyen kural: {rule}")

        # SMB kill her zaman dry-run pipeline'i ile dogrulanir.
        smb_dry = None
        try:
            from src.security.smb_session import kill_user_session
            smb_dry = kill_user_session(username, dry_run=True)
        except Exception as e:  # pragma: no cover - defensive
            smb_dry = {"error": f"smb_session_unavailable: {e}", "killed": 0}

        return {
            "rule": rule,
            "alert": last_alert,
            "smb_dry_run": smb_dry,
            "auto_kill_session": det.auto_kill_session,
            "notification_email": det.notification_email or None,
        }

    # ─────────────────────────────────────────────────────────────────
    # NTFS ACL / effective-permissions analyzer (#49)
    # ─────────────────────────────────────────────────────────────────

    def _get_acl_analyzer():
        existing = getattr(app.state, "acl_analyzer", None)
        if existing is not None:
            return existing
        from src.security.acl_analyzer import AclAnalyzer
        analyzer = AclAnalyzer(db, config, ad_lookup=ad_lookup)
        app.state.acl_analyzer = analyzer
        return analyzer

    @app.get("/api/security/acl")
    async def get_effective_acl(path: str = Query(..., min_length=1)):
        """Live effective DACL read for one path. Windows-only."""
        analyzer = _get_acl_analyzer()
        if not analyzer.is_supported():
            raise HTTPException(501, "ACL live read requires Windows + pywin32")
        try:
            return analyzer.get_effective_acl(path)
        except FileNotFoundError as e:
            raise HTTPException(404, f"Path not found: {path}") from e
        except PermissionError as e:
            raise HTTPException(403, f"Access denied reading ACL: {e}") from e
        except Exception as e:
            raise HTTPException(500, f"ACL read failed: {e}") from e

    @app.get("/api/security/acl/trustee/{sid}/paths")
    async def acl_paths_for_trustee(sid: str,
                                    limit: int = Query(100, ge=1, le=10000)):
        analyzer = _get_acl_analyzer()
        return {
            "trustee_sid": sid,
            "limit": limit,
            "paths": analyzer.find_paths_for_trustee(sid, limit=limit),
        }

    @app.get("/api/security/acl/sprawl")
    async def acl_sprawl(scan_id: Optional[int] = None,
                         severity_threshold: Optional[int] = None):
        analyzer = _get_acl_analyzer()
        thr = severity_threshold if severity_threshold is not None else analyzer.sprawl_threshold_mask
        return {
            "scan_id": scan_id,
            "severity_threshold": int(thr),
            "trustees": analyzer.detect_sprawl(scan_id=scan_id,
                                                severity_threshold=int(thr)),
        }

    @app.post("/api/security/acl/scan/{source_id}")
    async def acl_snapshot(source_id: int,
                           max_files: Optional[int] = None):
        analyzer = _get_acl_analyzer()
        if not analyzer.is_supported():
            raise HTTPException(501, "ACL snapshot requires Windows + pywin32")
        src = _get_source(db, source_id)
        with db.get_cursor() as cur:
            cur.execute(
                """SELECT id FROM scan_runs WHERE source_id=?
                   ORDER BY CASE WHEN status='completed' THEN 0 ELSE 1 END,
                            started_at DESC LIMIT 1""",
                (src.id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(409, f"No scan_runs found for source {source_id}")
        scan_id = row["id"]
        import asyncio

        def _do_snapshot():
            return analyzer.snapshot_source(src.id, scan_id, max_files=max_files)

        result = await asyncio.get_event_loop().run_in_executor(None, _do_snapshot)
        result["source_id"] = src.id
        result["scan_id"] = scan_id
        return result

    # ─────────────────────────────────────────────────────────────────
    # Orphaned-SID report + bulk reassignment (#56)
    # ─────────────────────────────────────────────────────────────────

    def _get_orphan_analyzer():
        existing = getattr(app.state, "orphan_sid_analyzer", None)
        if existing is not None:
            return existing
        from src.security.orphan_sid import OrphanSidAnalyzer
        analyzer = OrphanSidAnalyzer(db, config, ad_lookup=ad_lookup)
        app.state.orphan_sid_analyzer = analyzer
        return analyzer

    @app.get("/api/security/orphan-sids/{source_id}")
    async def orphan_sids_report(source_id: int,
                                 max_unique_sids: Optional[int] = None):
        """Detect orphan owner SIDs in the latest scan for ``source_id``."""
        analyzer = _get_orphan_analyzer()
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, f"No scan_runs found for source {source_id}")
        cap = max_unique_sids if max_unique_sids is not None else analyzer.max_unique_sids_default
        result = analyzer.detect_orphans(scan_id, max_unique_sids=int(cap))
        result["source_id"] = source_id
        return result

    @app.get("/api/security/orphan-sids/{source_id}/files")
    async def orphan_sid_files(source_id: int,
                               sid: str = Query(..., min_length=1),
                               page: int = Query(1, ge=1),
                               page_size: int = Query(100, ge=1, le=1000)):
        analyzer = _get_orphan_analyzer()
        return analyzer.get_orphan_files(source_id, sid, page=page, page_size=page_size)

    class OrphanReassignRequest(BaseModel):
        source_id: int
        sid: str
        new_owner: str
        dry_run: bool = True
        max_files: Optional[int] = None

    @app.post("/api/security/orphan-sids/reassign")
    async def orphan_sid_reassign(req: OrphanReassignRequest):
        analyzer = _get_orphan_analyzer()
        # Honour the opt-in dual-approval rule: refuse non-dry-run runs
        # unless the caller explicitly asked for it. (Dual-approval UX
        # itself is out of scope for this PR — this just blocks an
        # accidental single-button live reassignment.)
        if (not req.dry_run) and analyzer.require_dual_approval_for_reassign:
            raise HTTPException(
                403,
                "Live reassignment requires dual approval; submit via the "
                "approval workflow or set dry_run=true to preview.",
            )
        if (not req.dry_run) and not analyzer.is_supported():
            raise HTTPException(501, "Live reassignment requires Windows + pywin32")
        try:
            return analyzer.reassign_owner(
                req.source_id, req.sid, req.new_owner,
                dry_run=req.dry_run, max_files=req.max_files,
            )
        except NotImplementedError as e:
            raise HTTPException(501, str(e)) from e
        except ValueError as e:
            raise HTTPException(400, str(e)) from e

    @app.get("/api/security/orphan-sids/{source_id}/export.csv")
    async def orphan_sid_export_csv(source_id: int):
        """Streaming CSV download of orphan files for offline review."""
        from fastapi.responses import StreamingResponse
        import io

        analyzer = _get_orphan_analyzer()
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, f"No scan_runs found for source {source_id}")

        # Detect, then stream the file rows. We re-implement the row
        # generation here (instead of writing to a temp file) so the
        # response can stream incrementally on million-row shares.
        report = analyzer.detect_orphans(scan_id)
        orphan_sids = [row["sid"] for row in report.get("orphan_sids", [])]

        def _iter():
            buf = io.StringIO()
            writer = __import__("csv").writer(buf)
            writer.writerow([
                "path", "owner_sid", "file_size",
                "last_modify_time", "owner_resolved",
            ])
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)

            if not orphan_sids:
                return

            placeholders = ",".join(["?"] * len(orphan_sids))
            params: list = [source_id, scan_id, *orphan_sids]
            with db.get_cursor() as cur:
                cur.execute(
                    f"""SELECT file_path, owner, file_size, last_modify_time
                        FROM scanned_files
                        WHERE source_id = ? AND scan_id = ?
                          AND owner IN ({placeholders})
                        ORDER BY file_path""",
                    tuple(params),
                )
                for r in cur.fetchall():
                    writer.writerow([
                        r["file_path"], r["owner"], r["file_size"],
                        r["last_modify_time"] or "", "false",
                    ])
                    yield buf.getvalue()
                    buf.seek(0)
                    buf.truncate(0)

        filename = f"orphan_sids_source{source_id}_scan{scan_id}.csv"
        return StreamingResponse(
            _iter(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # ─────────────────────────────────────────────────────────────────
    # Security dashboard pages (#81): XLSX exports + ack-all
    #
    # The three security pages (Orphan SIDs, Ransomware Alerts, ACL
    # Analyzer) each get an `*/export.xlsx` endpoint and the ransomware
    # page gets a `/acknowledge-all` bulk action. We deliberately keep
    # the workbook construction inline rather than routing through the
    # background `_export_worker` queue: these tables are bounded
    # (top-N orphan SIDs, recent alerts, top trustees), so a synchronous
    # StreamingResponse is the simpler answer and matches `mit-naming`.
    # ─────────────────────────────────────────────────────────────────

    def _xlsx_response(rows, headers, filename: str, sheet_title: str):
        """Build a one-sheet XLSX from ``rows`` and return a
        ``StreamingResponse``. ``rows`` is a list of lists/tuples; the
        first column header maps to the first item in each row, etc.
        Raises HTTPException(500) if openpyxl is unavailable.
        """
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(500, "openpyxl kurulu degil. pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31] or "Sheet1"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B5797", end_color="2B5797",
                                   fill_type="solid")
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r_idx, r in enumerate(rows, start=2):
            for c_idx, val in enumerate(r, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @app.get("/api/security/feature-flags")
    async def security_feature_flags():
        """Read-only view of the three security feature flags so the
        dashboard can render a "kapali" banner when a page is disabled
        in config.yaml. Cheap; no DB access.
        """
        sec = (config or {}).get("security", {}) or {}
        ransom = sec.get("ransomware", {}) or {}
        orphan = sec.get("orphan_sid", {}) or {}
        return {
            "ransomware": {
                "enabled": bool(ransom.get("enabled", False)),
            },
            "orphan_sid": {
                "enabled": bool(orphan.get("enabled", False)),
                "require_dual_approval_for_reassign": bool(
                    orphan.get("require_dual_approval_for_reassign", False)
                ),
            },
            # ACL analyzer has no enabled flag in config.yaml today; it
            # is always available (read-side is DB-only).
            "acl": {"enabled": True},
        }

    @app.get("/api/security/orphan-sids/{source_id}/export.xlsx")
    async def orphan_sid_export_xlsx(source_id: int):
        """XLSX of orphan-SID summary rows (one row per orphan SID)."""
        analyzer = _get_orphan_analyzer()
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, f"No scan_runs found for source {source_id}")
        report = analyzer.detect_orphans(scan_id)
        rows = []
        for r in report.get("orphan_sids", []):
            rows.append([
                r.get("sid", ""),
                r.get("file_count", 0),
                r.get("total_size", 0),
                ", ".join(r.get("sample_paths", [])[:3]),
            ])
        filename = f"orphan_sids_source{source_id}_scan{scan_id}.xlsx"
        return _xlsx_response(
            rows,
            headers=["SID / Owner", "File Count", "Total Size (bytes)",
                     "Sample Paths"],
            filename=filename,
            sheet_title="Orphan SIDs",
        )

    @app.get("/api/security/ransomware/alerts/export.xlsx")
    async def ransomware_alerts_export_xlsx(
        since_minutes: int = Query(1440, ge=1, le=10080),
    ):
        """XLSX of ransomware alerts in the last N minutes."""
        det = _get_detector()
        alerts = det.get_active_alerts(since_minutes=since_minutes) or []
        rows = []
        for a in alerts:
            rows.append([
                a.get("triggered_at", ""),
                a.get("rule_name", ""),
                a.get("severity", ""),
                a.get("source_id", ""),
                a.get("username", ""),
                a.get("file_count", 0),
                ", ".join((a.get("sample_paths") or [])[:3]),
                a.get("acknowledged_at") or "",
                a.get("acknowledged_by") or "",
            ])
        filename = (
            f"ransomware_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return _xlsx_response(
            rows,
            headers=["Triggered At", "Rule", "Severity", "Source ID",
                     "Username", "File Count", "Sample Paths",
                     "Acknowledged At", "Acknowledged By"],
            filename=filename,
            sheet_title="Ransomware Alerts",
        )

    @app.post("/api/security/ransomware/alerts/acknowledge-all")
    async def acknowledge_all_ransomware_alerts(
        by_user: str = "admin",
        since_minutes: int = Query(1440, ge=1, le=10080),
    ):
        """Bulk acknowledge every unacknowledged alert in the window.

        Mirrors the per-row acknowledge endpoint but applies to all
        rows that match ``triggered_at > now - since_minutes`` AND have
        ``acknowledged_at IS NULL``. Returns the count of rows touched.
        """
        with db.get_cursor() as cur:
            cur.execute(
                """UPDATE ransomware_alerts
                   SET acknowledged_at = datetime('now','localtime'),
                       acknowledged_by = ?
                   WHERE acknowledged_at IS NULL
                     AND triggered_at > datetime('now', ? || ' minutes')""",
                (by_user, f"-{int(since_minutes)}"),
            )
            rowcount = cur.rowcount or 0
        return {
            "acknowledged": True,
            "rows_updated": int(rowcount),
            "by": by_user,
            "since_minutes": int(since_minutes),
        }

    @app.get("/api/security/acl/sprawl/export.xlsx")
    async def acl_sprawl_export_xlsx(
        scan_id: Optional[int] = None,
        severity_threshold: Optional[int] = None,
    ):
        """XLSX of the ACL-sprawl trustee report."""
        analyzer = _get_acl_analyzer()
        thr = (severity_threshold if severity_threshold is not None
               else analyzer.sprawl_threshold_mask)
        trustees = analyzer.detect_sprawl(
            scan_id=scan_id, severity_threshold=int(thr),
        )
        rows = []
        for t in trustees:
            rows.append([
                t.get("trustee_sid", ""),
                t.get("trustee_name", "") or "",
                t.get("file_count", 0),
                t.get("max_mask", 0),
                t.get("sample_permission_name", "") or "",
            ])
        filename = (
            f"acl_sprawl_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return _xlsx_response(
            rows,
            headers=["Trustee SID", "Trustee Name", "File Count",
                     "Max Permission Mask", "Sample Permission"],
            filename=filename,
            sheet_title="ACL Sprawl",
        )

    @app.get("/api/security/acl/trustee/{sid}/paths/export.xlsx")
    async def acl_trustee_paths_export_xlsx(
        sid: str, limit: int = Query(1000, ge=1, le=10000),
    ):
        """XLSX of every path a given trustee has access to."""
        analyzer = _get_acl_analyzer()
        paths = analyzer.find_paths_for_trustee(sid, limit=limit)
        rows = []
        for p in paths:
            rows.append([
                p.get("file_path", ""),
                p.get("permission_name", "") or "",
                p.get("ace_type", "") or "",
                p.get("permissions_mask", 0),
                int(p.get("is_inherited") or 0),
            ])
        # filename can't include backslashes from a SID, so sanitize
        safe_sid = sid.replace("\\", "_").replace("/", "_")
        filename = f"acl_trustee_{safe_sid}_paths.xlsx"
        return _xlsx_response(
            rows,
            headers=["File Path", "Permission", "ACE Type", "Mask",
                     "Inherited"],
            filename=filename,
            sheet_title="Trustee Paths",
        )

    # ─────────────────────────────────────────────────────────────────
    # Syslog/CEF integration (#50)
    # ─────────────────────────────────────────────────────────────────

    @app.get("/api/integrations/syslog/status")
    async def syslog_status():
        forwarder = getattr(app.state, "syslog", None)
        if forwarder is None:
            return {"available": False, "configured": False,
                    "reason": "forwarder_not_initialized"}
        return forwarder.health()

    @app.post("/api/integrations/syslog/test")
    async def syslog_test():
        forwarder = getattr(app.state, "syslog", None)
        if forwarder is None:
            return {"sent": False, "error": "forwarder_not_initialized"}
        if not forwarder.available:
            return {"sent": False, "error": "forwarder_disabled_or_unconfigured"}
        ok = forwarder.emit(
            "info",
            "test_event",
            {"msg": "FILE ACTIVITY syslog test event",
             "source": "dashboard", "version": APP_VERSION},
        )
        if not ok:
            return {"sent": False, "error": forwarder.health().get("last_error")}
        return {"sent": True}

    # ─────────────────────────────────────────────────────────────────
    # MCP Server discovery (#67) — read-only doc page
    # ─────────────────────────────────────────────────────────────────

    @app.get("/api/system/mcp/info")
    async def mcp_info():
        """Read-only MCP server discovery info for the dashboard.

        The MCP server is a separate process (``python -m src.mcp_server``);
        this endpoint just exposes its tool list + recommended install
        command so operators can wire it into Claude Desktop / Code without
        leaving the dashboard. We import the tool registry lazily so the
        dashboard process does not pull ``httpx`` etc. at startup if the
        operator never opens this page.
        """
        tools_info: list[dict] = []
        configured = False
        try:
            from src.mcp_server.tools import TOOLS  # type: ignore
            for t in TOOLS:
                tools_info.append({
                    "name": t.name,
                    "description": t.description,
                    "is_write": bool(getattr(t, "is_write", False)),
                })
            configured = True
        except Exception as e:  # pragma: no cover - defensive
            logger.warning("mcp_info: TOOLS import failed: %s", e)
            # Hardcoded fallback (the 15 names from PR #67) so the page
            # still renders something useful even if the optional MCP
            # extras aren't installed.
            for name, desc in [
                ("scan_list_sources", "List configured file-share sources."),
                ("scan_run", "Start a background scan (write — confirm required)."),
                ("scan_status", "Live scan progress for a source."),
                ("report_summary", "Latest-scan KPI summary."),
                ("report_duplicates", "Paged list of duplicate-content groups."),
                ("report_orphan_sids", "Owner SIDs that no longer resolve in AD."),
                ("pii_list_findings", "Browse persisted PII findings."),
                ("pii_subject_export", "GDPR subject export."),
                ("archive_dry_run", "Preview archive candidates."),
                ("archive_run", "Run archive workflow (write — confirm required)."),
                ("hold_list_active", "List active legal holds."),
                ("hold_add", "Create a legal hold (write — confirm required)."),
                ("hold_release", "Release a legal hold (write — confirm required)."),
                ("audit_query", "Query the tamper-evident audit log."),
                ("audit_verify_chain", "Verify the SHA-256 hash chain."),
            ]:
                tools_info.append({"name": name, "description": desc,
                                   "is_write": name.endswith("_run")
                                   or name in {"hold_add", "hold_release"}})
        return {
            "configured": configured,
            "tools_count": len(tools_info),
            "tools": tools_info,
            "transports": ["stdio"],
            "install_command":
                "claude mcp add file-activity -- python -m src.mcp_server",
        }

    # ─────────────────────────────────────────────────────────────────
    # Issue #77 Phase 2 — last-auto-restore banner state
    # ─────────────────────────────────────────────────────────────────
    @app.get("/api/system/last-restore")
    async def last_restore():
        """Returns the most recent auto-restore event for the current
        process, or ``{"restored": false}`` if no restore happened at
        startup. Frontend uses this to decide whether to draw the
        yellow forensic-preserved banner.
        """
        info = getattr(app.state, "last_auto_restore", None)
        if not info or not getattr(info, "restored", False):
            return {"restored": False}
        return {
            "restored": True,
            "snapshot_id": info.snapshot_id,
            "broken_path": info.broken_path,
            "ts": info.ts,
            "audit_event_id": info.audit_event_id,
        }

    # ─────────────────────────────────────────────────────────────────
    # System backups (#77) — read + manual snapshot + restore
    # ─────────────────────────────────────────────────────────────────

    def _get_backup_manager():
        existing = getattr(app.state, "backup_manager", None)
        if existing is not None:
            return existing
        from src.storage.backup_manager import BackupManager
        db_path = ((config or {}).get("database") or {}).get("path") \
            or "data/file_activity.db"
        mgr = BackupManager(db_path, config or {})
        app.state.backup_manager = mgr
        return mgr

    @app.get("/api/system/backups")
    async def list_backups():
        """List snapshot metadata rows from the manifest.

        Always returns a 200 with ``rows: []`` even when the backup feature
        is disabled — the frontend draws a feature-flag banner from
        ``enabled``/``configured`` so we don't want to 4xx here.
        """
        try:
            mgr = _get_backup_manager()
        except Exception as e:  # pragma: no cover - defensive
            return {"enabled": False, "configured": False,
                    "reason": f"manager_init_failed: {e}", "rows": []}
        rows = [m.to_dict() for m in mgr.list_snapshots()]
        return {
            "enabled": bool(mgr.enabled),
            "configured": True,
            "backup_dir": mgr.backup_dir,
            "keep_last_n": mgr.keep_last_n,
            "keep_weekly": mgr.keep_weekly,
            "rows": rows,
        }

    @app.post("/api/system/backups/snapshot")
    async def create_snapshot(body: dict):
        """Take a manual snapshot. Body: ``{"reason": "manual", "confirm": true}``."""
        body = body or {}
        if not bool(body.get("confirm", False)):
            raise HTTPException(400, "confirm: true required")
        reason = (body.get("reason") or "manual").strip() or "manual"
        try:
            mgr = _get_backup_manager()
        except Exception as e:
            raise HTTPException(500, f"manager_init_failed: {e}")
        if not mgr.enabled:
            raise HTTPException(400, "backup feature disabled in config.yaml")
        with _track_op(
            "snapshot",
            f"DB anlik goruntu: {reason}",
            metadata={"reason": reason},
        ):
            try:
                meta = mgr.snapshot(reason=reason)
            except Exception as e:
                logger.error("snapshot failed: %s", e)
                raise HTTPException(500, f"snapshot_failed: {e}")
        return {"ok": True, **meta.to_dict()}

    def _do_snapshot_restore(payload: dict) -> dict:
        """Execute a snapshot restore from a payload dict.

        Shared by the direct endpoint (when approvals are disabled or
        the op isn't gated) and ``ApprovalRegistry.execute`` so the same
        code path runs whether or not the two-person rule is in effect.
        Raises ``HTTPException`` so both callers can surface the same
        error semantics.
        """
        snapshot_id = (payload or {}).get("snapshot_id")
        if not snapshot_id:
            raise HTTPException(400, "snapshot_id required in payload")
        try:
            mgr = _get_backup_manager()
        except Exception as e:
            raise HTTPException(500, f"manager_init_failed: {e}")
        if not mgr.enabled:
            raise HTTPException(400, "backup feature disabled in config.yaml")
        try:
            mgr.restore(snapshot_id)
        except KeyError:
            raise HTTPException(404, f"unknown snapshot id: {snapshot_id}")
        except FileNotFoundError as e:
            raise HTTPException(404, str(e))
        except RuntimeError as e:
            raise HTTPException(409, str(e))
        except Exception as e:
            logger.error("restore failed: %s", e)
            raise HTTPException(500, f"restore_failed: {e}")
        return {"ok": True, "restored": snapshot_id}

    # Map of operation_type -> executor callable. Used by
    # ``POST /api/approvals/{id}/execute`` to dispatch the approved
    # payload back through the original code path. Phase 1 only wires
    # snapshot_restore — archive_bulk, purge_bulk, retention_apply
    # follow in subsequent PRs.
    app.state.approval_executors = {
        "snapshot_restore": _do_snapshot_restore,
    }

    @app.post("/api/system/backups/restore/{snapshot_id}")
    async def restore_snapshot(snapshot_id: str, body: dict, request: Request):
        """Restore the live DB from ``snapshot_id``. Refuses if a live
        connection holds the DB lock — the caller must stop the dashboard
        first. Body must include ``confirm: true`` and
        ``safety_token: "RESTORE"`` (audit M-3, mirrors the PURGE /
        QUARANTINE triple-gate from PR #109/#110).

        When ``approvals.enabled=true`` AND ``'snapshot_restore'`` is in
        ``approvals.require_for``, this endpoint queues a pending
        approval row instead of executing immediately. The caller must
        then have a *different* operator approve via
        ``POST /api/approvals/{id}/approve`` and finally trigger
        execution via ``POST /api/approvals/{id}/execute``.
        """
        body = body or {}
        if not bool(body.get("confirm", False)):
            raise HTTPException(400, "confirm: true required")
        # Defence-in-depth: even with confirm=true, an attacker who can
        # forge a request still has to know the literal "RESTORE" token.
        # Mirrors PURGE_SAFETY_TOKEN_VALUE in DuplicateCleaner.
        if body.get("safety_token", "") != "RESTORE":
            raise HTTPException(
                400, "safety_token must equal 'RESTORE'"
            )

        # Route through approvals when gated. The framework is opt-in;
        # default config keeps this branch dormant.
        registry = getattr(app.state, "approval_registry", None)
        if registry is not None and registry.is_required("snapshot_restore"):
            from src.security.identity import resolve_user
            requested_by = resolve_user(request, config, body)
            req = registry.request(
                "snapshot_restore",
                {"snapshot_id": snapshot_id},
                requested_by,
            )
            return {
                "pending_approval_id": req.id,
                "status": "pending",
                "operation_type": req.operation_type,
                "expires_at": req.expires_at,
                "requested_by": req.requested_by,
                "message": "Awaiting second-person approval",
            }

        return _do_snapshot_restore({"snapshot_id": snapshot_id})

    @app.get("/api/system/backups/export")
    async def export_backups_xlsx():
        """Export the snapshot manifest as XLSX (uses openpyxl if
        available, else CSV fallback). Mirrors the pattern other dashboard
        pages use — the frontend just hits this URL and saves the blob."""
        try:
            mgr = _get_backup_manager()
            rows = [m.to_dict() for m in mgr.list_snapshots()]
        except Exception as e:
            raise HTTPException(500, f"manager_init_failed: {e}")
        headers = ["id", "created_at", "reason", "size_bytes",
                   "sha256", "path"]
        try:
            import io
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Backups"
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h, "") for h in headers])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            from fastapi.responses import Response
            filename = f"backups_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return Response(
                content=buf.getvalue(),
                media_type=("application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet"),
                headers={"Content-Disposition":
                         f"attachment; filename={filename}"},
            )
        except ImportError:
            # CSV fallback — no openpyxl available.
            import io
            import csv
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(headers)
            for r in rows:
                w.writerow([r.get(h, "") for h in headers])
            from fastapi.responses import Response
            filename = f"backups_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            return Response(
                content=buf.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition":
                         f"attachment; filename={filename}"},
            )

    # ─────────────────────────────────────────────────────────────────
    # GDPR PII detection + retention engine (#58)
    # ─────────────────────────────────────────────────────────────────

    def _get_pii_engine():
        existing = getattr(app.state, "pii_engine", None)
        if existing is not None:
            return existing
        from src.compliance.pii_engine import PiiEngine
        engine = PiiEngine(db, config)
        app.state.pii_engine = engine
        return engine

    def _get_retention_engine():
        existing = getattr(app.state, "retention_engine", None)
        if existing is not None:
            return existing
        from src.compliance.retention import RetentionEngine
        # Lazy archive engine — only constructed if a policy actually
        # needs it (action='archive' + non-dry-run apply).
        archive_engine = None
        try:
            from src.archiver.archive_engine import ArchiveEngine
            archive_engine = ArchiveEngine(db, config)
        except Exception as e:  # pragma: no cover - defensive
            logger.debug("ArchiveEngine init skipped for retention: %s", e)
        engine = RetentionEngine(db, config, archive_engine=archive_engine)
        app.state.retention_engine = engine
        return engine

    @app.post("/api/compliance/pii/scan/{source_id}")
    async def pii_scan(source_id: int,
                       max_files: Optional[int] = None,
                       overwrite_existing: bool = False):
        """Run PiiEngine.scan_source against the latest scan of source_id."""
        engine = _get_pii_engine()
        src = _get_source(db, source_id)
        import asyncio

        def _run():
            return engine.scan_source(
                src.id,
                max_files=max_files,
                overwrite_existing=overwrite_existing,
            )

        result = await asyncio.get_event_loop().run_in_executor(None, _run)
        result["source_id"] = src.id
        return result

    @app.get("/api/compliance/pii/findings")
    async def pii_findings(pattern: Optional[str] = None,
                           source_id: Optional[int] = None,
                           page: int = Query(1, ge=1),
                           page_size: int = Query(50, ge=1, le=1000)):
        """Browse persisted pii_findings rows.

        Optional filters:
        * ``pattern``    — exact ``pattern_name`` match (e.g. ``email``)
        * ``source_id``  — restrict to scans belonging to a single source

        Issue #81: ``source_id`` filter joins through ``scan_runs`` so
        the dashboard's "Compliance > PII Findings" page can scope rows
        to the source the operator picked in the source selector.
        """
        _get_pii_engine()  # ensure engine constructable / config sane
        offset = (page - 1) * page_size
        params: list = []
        clauses: list = []
        if pattern:
            clauses.append("p.pattern_name = ?")
            params.append(pattern)
        if source_id is not None:
            # ``pii_findings.scan_id`` -> ``scan_runs.source_id``.
            clauses.append(
                "p.scan_id IN (SELECT id FROM scan_runs WHERE source_id = ?)"
            )
            params.append(int(source_id))
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        with db.get_cursor() as cur:
            cur.execute(
                f"SELECT COUNT(*) AS cnt FROM pii_findings p {where}",
                params,
            )
            total = cur.fetchone()["cnt"]
            cur.execute(
                f"""SELECT p.id, p.scan_id, p.file_path, p.pattern_name,
                           p.hit_count, p.sample_snippet, p.detected_at
                    FROM pii_findings p {where}
                    ORDER BY p.detected_at DESC, p.id DESC
                    LIMIT ? OFFSET ?""",
                params + [page_size, offset],
            )
            rows = [dict(r) for r in cur.fetchall()]
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "findings": rows,
        }

    @app.get("/api/compliance/pii/patterns")
    async def pii_patterns():
        """Return the active pattern dictionary (built-ins + operator
        overrides), so the dashboard can populate a filter dropdown
        without hardcoding pattern names. Issue #81.
        """
        engine = _get_pii_engine()
        # Engine.patterns may be empty when the Hyperscan backend is
        # in use; merge from DEFAULT_PATTERNS + user overrides instead.
        from src.compliance.pii_engine import PiiEngine as _PE
        merged = dict(_PE.DEFAULT_PATTERNS)
        cfg = ((config or {}).get("compliance", {}) or {}).get("pii", {}) or {}
        user_patterns = cfg.get("patterns") or {}
        if isinstance(user_patterns, dict):
            for name in user_patterns:
                if name:
                    merged[str(name)] = "<custom>"
        return {
            "patterns": sorted(merged.keys()),
            "backend": engine.engine_name,
        }

    @app.get("/api/compliance/pii/findings/export.xlsx")
    async def pii_findings_export_xlsx(
        pattern: Optional[str] = None,
        source_id: Optional[int] = None,
        ids: Optional[str] = Query(
            None,
            description="Comma-separated pii_findings.id values; empty = all matching rows",
        ),
        format: Optional[str] = Query(
            None,
            description="Set to 'csv' to bypass Excel's 1,048,576-row limit (issue #122)",
        ),
    ):
        """Export persisted PII findings as XLSX (issue #81; #122 streaming).

        Mirrors the mit-naming export pattern: optional ``ids`` filter
        scopes the workbook to a subset selected in the dashboard's
        entity-list. ``pattern`` and ``source_id`` honour the same
        filters as the JSON listing endpoint. Pass ``?format=csv`` for an
        unbounded streaming CSV (issue #122 fallback).
        """
        from fastapi.responses import StreamingResponse
        import io
        from datetime import datetime

        from src.utils.xlsx_writer import write_large_workbook, stream_csv

        _get_pii_engine()

        params: list = []
        clauses: list = []
        if pattern:
            clauses.append("p.pattern_name = ?")
            params.append(pattern)
        if source_id is not None:
            clauses.append(
                "p.scan_id IN (SELECT id FROM scan_runs WHERE source_id = ?)"
            )
            params.append(int(source_id))

        id_filter: Optional[set] = None
        if ids:
            id_filter = set()
            for tok in ids.split(","):
                tok = tok.strip()
                if not tok:
                    continue
                try:
                    id_filter.add(int(tok))
                except ValueError:
                    continue

        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

        columns = [
            {"key": "id", "header": "id", "width": 10},
            {"key": "scan_id", "header": "scan_id", "width": 10},
            {"key": "file_path", "header": "file_path", "width": 60},
            {"key": "pattern_name", "header": "pattern_name", "width": 18},
            {"key": "hit_count", "header": "hit_count", "width": 10},
            {"key": "sample_snippet", "header": "sample_snippet", "width": 40},
            {"key": "detected_at", "header": "detected_at", "width": 22},
        ]

        # Stream rows directly out of the cursor — keeps memory flat even
        # when the findings table grows past 1M rows on big PII scans.
        def _finding_rows():
            with db.get_cursor() as cur:
                cur.execute(
                    f"""SELECT p.id, p.scan_id, p.file_path, p.pattern_name,
                               p.hit_count, p.sample_snippet, p.detected_at
                        FROM pii_findings p {where}
                        ORDER BY p.detected_at DESC, p.id DESC""",
                    params,
                )
                for r in cur:
                    if id_filter is not None and r["id"] not in id_filter:
                        continue
                    yield {
                        "id": r["id"],
                        "scan_id": r["scan_id"],
                        "file_path": r["file_path"] or "",
                        "pattern_name": r["pattern_name"] or "",
                        "hit_count": r["hit_count"] or 0,
                        "sample_snippet": r["sample_snippet"] or "",
                        "detected_at": r["detected_at"] or "",
                    }

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        # ---- CSV fallback (issue #122) ------------------------------------
        if (format or "").lower() == "csv":
            return StreamingResponse(
                stream_csv(_finding_rows(), columns),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition":
                        f"attachment; filename=pii_findings_{ts}.csv",
                    "X-Format-Fallback": "csv",
                },
            )

        # ---- XLSX (default) ----------------------------------------------
        buf = io.BytesIO()
        meta = write_large_workbook(_finding_rows(), columns, buf)
        buf.seek(0)
        filename = f"pii_findings_{ts}.xlsx"
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Total-Rows": str(meta["total_rows"]),
                "X-Sheet-Count": str(meta["sheet_count"]),
            },
        )

    @app.get("/api/compliance/pii/subject")
    async def pii_subject(term: str = Query(..., min_length=1),
                          format: str = Query("json", pattern="^(json|csv)$")):
        """Article 17/30 export — every file mentioning ``term``."""
        engine = _get_pii_engine()
        if format == "csv":
            from fastapi.responses import StreamingResponse
            import io
            import csv as _csv
            rows = engine.find_for_subject(term, limit=100_000)
            buf = io.StringIO()
            writer = _csv.writer(buf)
            writer.writerow([
                "file_path", "match_count", "last_modify_time",
                "owner", "patterns", "sample_snippets",
            ])
            for r in rows:
                patterns = ";".join(h["pattern_name"] for h in r["hits"])
                snippets = ";".join(
                    h.get("sample_snippet") or "" for h in r["hits"]
                )
                writer.writerow([
                    r["file_path"], r["match_count"],
                    r["last_modify_time"] or "",
                    r["owner"] or "",
                    patterns, snippets,
                ])
            buf.seek(0)
            safe_term = "".join(c if c.isalnum() else "_" for c in term)[:40]
            return StreamingResponse(
                iter([buf.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition":
                        f"attachment; filename=pii_subject_{safe_term}.csv"
                },
            )
        results = engine.find_for_subject(term)
        return {"term": term, "matches": len(results), "files": results}

    @app.get("/api/compliance/pii/backend")
    async def pii_backend_status():
        """Capability probe (issue #64) — which regex backend the
        PII engine is actually using and the optional package version.

        Returns ``{"backend": "hyperscan"|"re",
                   "version": <pkg version or null>,
                   "hyperscan_available": bool,
                   "configured": <raw engine config value>}``.
        Mirrors the ``OrphanSidAnalyzer.is_supported`` pattern of
        surfacing optional-acceleration availability over the API.
        """
        from src.compliance._pii_backends import (
            hyperscan_available, hyperscan_version,
        )
        engine = _get_pii_engine()
        configured = (
            ((config or {}).get("compliance", {}) or {}).get("pii", {}) or {}
        ).get("engine", "auto")
        return {
            "backend": engine.engine_name,
            "version": hyperscan_version() if engine.engine_name == "hyperscan" else None,
            "hyperscan_available": hyperscan_available(),
            "configured": configured,
        }

    @app.get("/api/compliance/retention/policies")
    async def retention_policies_list():
        engine = _get_retention_engine()
        return {"policies": engine.list_policies()}

    class _RetentionPolicyCreate(BaseModel):
        name: str
        pattern_match: Optional[str] = ""
        retain_days: int
        action: str

        @field_validator("action")
        @classmethod
        def _validate_action(cls, v: str) -> str:
            if v not in ("archive", "delete"):
                raise ValueError("action must be 'archive' or 'delete'")
            return v

    @app.post("/api/compliance/retention/policies")
    async def retention_policy_create(data: _RetentionPolicyCreate):
        engine = _get_retention_engine()
        try:
            pid = engine.add_policy(
                data.name,
                data.pattern_match or "",
                data.retain_days,
                data.action,
            )
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:
            # Most likely a UNIQUE-name collision.
            raise HTTPException(400, str(e))
        return {"id": pid, "name": data.name}

    @app.delete("/api/compliance/retention/policies/{name}")
    async def retention_policy_remove(name: str):
        engine = _get_retention_engine()
        if not engine.remove_policy(name):
            raise HTTPException(404, f"Policy not found: {name}")
        return {"removed": True, "name": name}

    @app.post("/api/compliance/retention/apply/{policy_name}")
    async def retention_policy_apply(policy_name: str,
                                     dry_run: bool = True):
        engine = _get_retention_engine()
        import asyncio

        def _run():
            return engine.apply(policy_name, dry_run=dry_run)

        try:
            result = await asyncio.get_event_loop().run_in_executor(None, _run)
        except ValueError as e:
            raise HTTPException(404, str(e))
        except RuntimeError as e:
            raise HTTPException(503, str(e))
        return result

    @app.get("/api/compliance/retention/attestation")
    async def retention_attestation(since_days: int = Query(30, ge=1, le=3650)):
        engine = _get_retention_engine()
        return engine.attestation_report(since_days=since_days)

    @app.get("/api/compliance/retention/attestation/export.xlsx")
    async def retention_attestation_xlsx(
        since_days: int = Query(30, ge=1, le=3650),
    ):
        """Attestation report as XLSX (issue #81).

        Two sheets: ``By Policy`` (one row per policy/action group) and
        ``Events`` (one row per audit event in the window). Mirrors the
        mit-naming export pattern.
        """
        from fastapi.responses import StreamingResponse
        import io
        from datetime import datetime

        try:
            from openpyxl import Workbook
        except ImportError as e:  # pragma: no cover - in requirements.txt
            raise HTTPException(
                500,
                "openpyxl is not installed; add openpyxl>=3.1.0 to requirements.txt",
            ) from e

        engine = _get_retention_engine()
        report = engine.attestation_report(since_days=since_days)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "By Policy"
        by_policy_headers = [
            "policy", "action", "count", "first_event", "last_event",
        ]
        ws1.append(by_policy_headers)
        for r in (report.get("by_policy") or []):
            ws1.append([r.get(h, "") for h in by_policy_headers])

        ws2 = wb.create_sheet(title="Events")
        event_headers = ["id", "event_time", "event_type", "file_path", "details"]
        ws2.append(event_headers)
        for ev in (report.get("events") or []):
            ws2.append([ev.get(h, "") for h in event_headers])

        # Summary cell on a separate sheet for auditors.
        ws3 = wb.create_sheet(title="Summary")
        ws3.append(["since_days", report.get("since_days")])
        ws3.append(["generated_at", report.get("generated_at")])
        totals = report.get("totals") or {}
        ws3.append(["total_archive", totals.get("archive", 0)])
        ws3.append(["total_delete", totals.get("delete", 0)])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = (
            f"retention_attestation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
            },
        )

    @app.get("/api/compliance/config")
    async def compliance_config():
        """Expose the three compliance feature flags so the dashboard
        can render a "feature disabled" banner without hardcoding the
        config path. Issue #81.
        """
        cfg = (config or {}).get("compliance") or {}
        pii_cfg = cfg.get("pii") or {}
        retention_cfg = cfg.get("retention") or {}
        legal_hold_cfg = cfg.get("legal_hold") or {}
        return {
            "pii": {"enabled": bool(pii_cfg.get("enabled", False))},
            "retention": {"enabled": bool(retention_cfg.get("enabled", False))},
            "legal_hold": {"enabled": bool(legal_hold_cfg.get("enabled", True))},
        }

    # ──────────────────────────────────────────────
    # Compliance — Legal Hold Registry (issue #59)
    # ──────────────────────────────────────────────
    # Constructed once per app; the registry itself is stateless beyond
    # the DB handle, so a single shared instance is safe across requests.
    from src.compliance.legal_hold import LegalHoldRegistry
    app.state.legal_hold = LegalHoldRegistry(db, config)

    @app.get("/api/compliance/legal-holds/active")
    async def legal_holds_active():
        return {"holds": app.state.legal_hold.list_active()}

    @app.get("/api/compliance/legal-holds/history")
    async def legal_holds_history(page: int = Query(1, ge=1),
                                  page_size: int = Query(50, ge=1, le=500)):
        return app.state.legal_hold.list_history(page=page, page_size=page_size)

    @app.post("/api/compliance/legal-holds")
    async def legal_holds_add(body: dict):
        pattern = (body or {}).get("pattern")
        reason = (body or {}).get("reason")
        case_ref = (body or {}).get("case_ref")
        created_by = (body or {}).get("created_by") or "dashboard"
        try:
            hold_id = app.state.legal_hold.add_hold(
                pattern=pattern,
                reason=reason,
                case_ref=case_ref,
                created_by=created_by,
            )
        except ValueError as e:
            raise HTTPException(400, str(e))
        return {"id": hold_id, "ok": True}

    @app.post("/api/compliance/legal-holds/{hold_id}/release")
    async def legal_holds_release(hold_id: int, body: dict):
        released_by = (body or {}).get("released_by") or "dashboard"
        try:
            ok = app.state.legal_hold.release_hold(hold_id, released_by)
        except ValueError as e:
            raise HTTPException(400, str(e))
        if not ok:
            # Either unknown id or already released — same response so
            # the dashboard can treat it idempotently.
            return {"ok": False, "reason": "not_found_or_already_released"}
        return {"ok": True}

    @app.get("/api/compliance/legal-holds/check")
    async def legal_holds_check(path: str = Query(..., min_length=1)):
        held = app.state.legal_hold.is_held(path)
        return {"path": path, "is_held": held is not None, "hold": held}

    @app.get("/api/compliance/legal-holds/badge")
    async def legal_holds_badge():
        """Sidebar badge data — cheap polling endpoint."""
        registry = app.state.legal_hold
        actives = registry.list_active()
        try:
            held_paths = registry.count_held_paths()
        except Exception as e:  # pragma: no cover - defensive only
            logger.warning("legal_hold count_held_paths failed: %s", e)
            held_paths = 0
        return {
            "active_count": len(actives),
            "held_paths_count": held_paths,
        }

    # ──────────────────────────────────────────────
    # Two-person approval framework (issue #112)
    # ──────────────────────────────────────────────
    # Registry lives on app.state so the snapshot-restore endpoint
    # (above) and these endpoints share a single instance. Constructed
    # unconditionally — when ``approvals.enabled=false`` every
    # ``is_required`` call short-circuits to False and existing ops run
    # straight through, so this is zero-cost backwards compat.
    from src.security.approvals import (
        ApprovalRegistry, ApprovalNotFound, SelfApprovalError,
        InvalidStateError, ApprovalExpiredError, ApprovalError,
    )
    from src.security.identity import resolve_user, warn_if_unsafe
    # Issue #158 (H-2) — refuse the unsafe combo at boot.
    #
    # ApprovalRegistry.__init__ raises RuntimeError when
    # ``approvals.enabled=true`` AND ``identity_source='client_supplied'``
    # because the second-person rule degenerates to "any caller can claim
    # the requester's name" in that mode. We catch the error here, log
    # CRITICAL, and leave ``app.state.approval_registry = None``. The
    # approval endpoints below check for None and return HTTP 503 so
    # callers get a clear "approvals disabled until config fixed" signal
    # rather than a cryptic 500.
    try:
        app.state.approval_registry = ApprovalRegistry(db, config)
    except RuntimeError as e:
        logger.critical(
            "Approval framework disabled: %s. Edit config.yaml to set "
            "approvals.identity_source to 'windows' or 'header', or "
            "set approvals.enabled=false, then restart.",
            e,
        )
        app.state.approval_registry = None
    warn_if_unsafe(config)

    def _approval_to_json(req) -> dict:
        return req.to_dict()

    @app.get("/api/approvals/config")
    async def approvals_config():
        """Expose the runtime config of the approval framework so the
        frontend can render the right banners + disable/enable buttons.
        Safe to read while disabled — returns ``enabled=false``."""
        cfg = (config or {}).get("approvals") or {}
        return {
            "enabled": bool(cfg.get("enabled", False)),
            "require_for": list(cfg.get("require_for") or []),
            "expiry_hours": int(cfg.get("expiry_hours", 24) or 24),
            "identity_source": (cfg.get("identity_source")
                                or "client_supplied"),
        }

    # Issue #158 (H-2) helper — every endpoint below dereferences the
    # registry; this guards against the "unsafe combo refused at boot"
    # path where the registry is None.
    _APPROVALS_DISABLED_MSG = (
        "Approvals framework disabled - check server log for the "
        "config error (issue #158 H-2). Likely cause: "
        "approvals.enabled=true with identity_source='client_supplied'."
    )

    def _require_approval_registry():
        registry = getattr(app.state, "approval_registry", None)
        if registry is None:
            raise HTTPException(503, _APPROVALS_DISABLED_MSG)
        return registry

    @app.get("/api/approvals/pending")
    async def approvals_list_pending():
        registry = _require_approval_registry()
        rows = registry.list_pending()
        return {"rows": [_approval_to_json(r) for r in rows]}

    @app.get("/api/approvals/history")
    async def approvals_history(limit: int = Query(50, ge=1, le=1000)):
        registry = _require_approval_registry()
        rows = registry.list_history(limit=limit)
        return {"rows": [_approval_to_json(r) for r in rows]}

    @app.post("/api/approvals/{approval_id}/approve")
    async def approvals_approve(approval_id: int, body: dict, request: Request):
        body = body or {}
        registry = _require_approval_registry()
        approved_by = (body.get("approved_by")
                       or resolve_user(request, config, body))
        try:
            req = registry.approve(approval_id, approved_by)
        except ApprovalNotFound:
            raise HTTPException(404, f"approval {approval_id} not found")
        except SelfApprovalError as e:
            raise HTTPException(403, str(e))
        except ApprovalExpiredError as e:
            raise HTTPException(409, str(e))
        except InvalidStateError as e:
            raise HTTPException(409, str(e))
        except ApprovalError as e:
            raise HTTPException(400, str(e))
        return {"ok": True, "approval": _approval_to_json(req)}

    @app.post("/api/approvals/{approval_id}/reject")
    async def approvals_reject(approval_id: int, body: dict, request: Request):
        body = body or {}
        registry = _require_approval_registry()
        rejected_by = (body.get("rejected_by")
                       or resolve_user(request, config, body))
        reason = (body.get("reason") or "").strip()
        try:
            req = registry.reject(approval_id, rejected_by, reason)
        except ApprovalNotFound:
            raise HTTPException(404, f"approval {approval_id} not found")
        except InvalidStateError as e:
            raise HTTPException(409, str(e))
        except ApprovalError as e:
            raise HTTPException(400, str(e))
        return {"ok": True, "approval": _approval_to_json(req)}

    @app.post("/api/approvals/{approval_id}/execute")
    async def approvals_execute(approval_id: int, body: dict):
        """Run the executor mapped to the approval's operation_type.

        Body may carry an ``executor_token`` reserved for future
        signing (issue #112 follow-up); Phase 1 ignores it but the
        field is documented in the API surface so client code can
        pre-stage support.
        """
        body = body or {}
        registry = _require_approval_registry()
        try:
            req = registry.get(approval_id)
        except ApprovalNotFound:
            raise HTTPException(404, f"approval {approval_id} not found")

        executor_map = getattr(app.state, "approval_executors", {}) or {}
        executor = executor_map.get(req.operation_type)
        if executor is None:
            raise HTTPException(
                400,
                f"no executor registered for operation_type "
                f"{req.operation_type!r} (Phase 1 wires snapshot_restore only)",
            )
        try:
            result = registry.execute(approval_id, executor)
        except InvalidStateError as e:
            raise HTTPException(409, str(e))
        except HTTPException:
            # Executor surfaced its own HTTP error — propagate.
            raise
        except ApprovalError as e:
            raise HTTPException(400, str(e))
        return {"ok": True, "result": result}

    # ─────────────────────────────────────────────────────────────────────
    # Chargeback / cost-center reports (issue #111).
    #
    # CRUD: cost centers + owner patterns (manual mapping; AD auto-discovery
    # deferred per #111 Phase 1). Compute: per-scan aggregation. Export: XLSX
    # workbook with FORMULAS so auditors can edit the rate cell on the
    # Settings sheet and have totals recompute automatically.
    # ─────────────────────────────────────────────────────────────────────

    def _chargeback_report():
        from src.reports.chargeback import ChargebackReport
        return ChargebackReport(db, config)

    @app.get("/api/chargeback/centers")
    async def chargeback_list_centers():
        return {"centers": _chargeback_report().list_centers()}

    @app.post("/api/chargeback/centers")
    async def chargeback_add_center(body: dict):
        if not isinstance(body, dict):
            raise HTTPException(400, "JSON body bekleniyor")
        name = (body.get("name") or "").strip()
        if not name:
            raise HTTPException(400, "name gerekli")
        try:
            cid = _chargeback_report().add_center(
                name=name,
                description=body.get("description") or "",
                cost_per_gb_month=body.get("cost_per_gb_month") or 0,
            )
        except ValueError as e:
            raise HTTPException(400, str(e))
        except Exception as e:
            # UNIQUE name violation surfaces as IntegrityError on SQLite.
            raise HTTPException(409, f"Eklenemedi: {e}")
        return {"id": cid, "ok": True}

    @app.put("/api/chargeback/centers/{center_id}")
    async def chargeback_update_center(center_id: int, body: dict):
        if not isinstance(body, dict):
            raise HTTPException(400, "JSON body bekleniyor")
        # Strip unknown keys to make the endpoint idempotent and safe.
        allowed = {"name", "description", "cost_per_gb_month"}
        fields = {k: v for k, v in body.items() if k in allowed}
        try:
            ok = _chargeback_report().update_center(center_id, **fields)
        except ValueError as e:
            raise HTTPException(400, str(e))
        if not ok:
            # No matching row OR no fields to update: treat as 404 only when
            # the center genuinely does not exist (idempotent vs no-op).
            existing = _chargeback_report().get_center(center_id)
            if not existing:
                raise HTTPException(404, "cost_center bulunamadi")
        return {"ok": True}

    @app.delete("/api/chargeback/centers/{center_id}")
    async def chargeback_remove_center(center_id: int):
        # Idempotent: deleting an already-deleted center returns ok=True
        # with deleted=False so callers can call this on stale UI state.
        deleted = _chargeback_report().remove_center(center_id)
        return {"ok": True, "deleted": deleted}

    @app.post("/api/chargeback/centers/{center_id}/owners")
    async def chargeback_add_owner(center_id: int, body: dict):
        if not isinstance(body, dict):
            raise HTTPException(400, "JSON body bekleniyor")
        pat = (body.get("owner_pattern") or "").strip()
        if not pat:
            raise HTTPException(400, "owner_pattern gerekli")
        try:
            added = _chargeback_report().add_owner(center_id, pat)
        except ValueError as e:
            raise HTTPException(404 if "bulunamadi" in str(e) else 400, str(e))
        return {"ok": True, "added": added, "owner_pattern": pat}

    @app.delete("/api/chargeback/centers/{center_id}/owners/{owner_pattern:path}")
    async def chargeback_remove_owner(center_id: int, owner_pattern: str):
        # ``:path`` lets the pattern contain backslashes / slashes from the
        # owner field (e.g. ``CONTOSO\jdoe``) without double-encoding.
        deleted = _chargeback_report().remove_owner(center_id, owner_pattern)
        return {"ok": True, "deleted": deleted}

    @app.get("/api/chargeback/{source_id}")
    async def chargeback_compute(source_id: int):
        """Compute the chargeback report for the latest scan of a source."""
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, "Tamamlanmis scan yok")
        result = _chargeback_report().compute(scan_id)
        return result.to_dict()

    @app.get("/api/chargeback/{source_id}/export.xlsx")
    async def chargeback_export_xlsx(
        source_id: int,
        format: Optional[str] = Query(
            None,
            description="Set to 'csv' for a flat CSV (issue #122 fallback)",
        ),
    ):
        """Chargeback workbook (issue #111).

        XLSX preserves the auditor-editable Settings!$B$2 rate cell + the
        per-row =Settings!$B$2*C{r} formulas byte-identically (chargeback
        rows are top-N owners + unmapped, never anywhere near Excel's
        1M-row cap, so #122's split logic does not apply here). For
        completeness with the rest of the export surface we still expose
        ``?format=csv``, which flattens Detail to a single CSV — the rate
        is materialised per-row instead of formula-driven.
        """
        from fastapi.responses import StreamingResponse
        import io as _io
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, "Tamamlanmis scan yok")

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        if (format or "").lower() == "csv":
            from src.utils.xlsx_writer import stream_csv
            cb = _chargeback_report()
            result = cb.compute(scan_id)
            # Flatten owners + unmapped buckets into a single row stream.
            # The CSV materialises monthly_cost rather than carrying the
            # =Settings!$B$2 formula since CSV has no formula support.
            _BYTES_PER_GB = 1024 ** 3
            # Pick the first non-zero rate as the global default — same
            # heuristic the XLSX export uses for Settings!$B$2.
            default_rate = 0.0
            non_zero = [c.cost_per_gb_month for c in result.centers
                        if c.cost_per_gb_month]
            if non_zero:
                default_rate = float(non_zero[0])

            def _detail_rows():
                for ct in result.centers:
                    for o in ct.top_owners:
                        gb = float(o.get("total_bytes") or 0) / _BYTES_PER_GB
                        rate = (ct.cost_per_gb_month
                                if ct.cost_per_gb_month else default_rate)
                        yield {
                            "cost_center": ct.name,
                            "owner": o.get("owner") or "",
                            "total_gb": round(gb, 6),
                            "file_count": int(o.get("file_count") or 0),
                            "monthly_cost": round(gb * rate, 4),
                        }
                for u in result.unmapped_owners or []:
                    gb = float(u.get("total_bytes") or 0) / _BYTES_PER_GB
                    yield {
                        "cost_center": "__unmapped__",
                        "owner": u.get("owner") or "",
                        "total_gb": round(gb, 6),
                        "file_count": int(u.get("file_count") or 0),
                        "monthly_cost": round(gb * default_rate, 4),
                    }

            csv_columns = [
                {"key": "cost_center", "header": "cost_center"},
                {"key": "owner", "header": "owner"},
                {"key": "total_gb", "header": "total_gb"},
                {"key": "file_count", "header": "file_count"},
                {"key": "monthly_cost", "header": "monthly_cost"},
            ]
            return StreamingResponse(
                stream_csv(_detail_rows(), csv_columns),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition":
                        f"attachment; filename=Chargeback_source{source_id}"
                        f"_scan{scan_id}_{ts}.csv",
                    "X-Format-Fallback": "csv",
                },
            )

        try:
            blob = _chargeback_report().export_xlsx(scan_id)
        except RuntimeError as e:
            raise HTTPException(500, str(e))
        filename = f"Chargeback_source{source_id}_scan{scan_id}_{ts}.xlsx"
        return StreamingResponse(
            _io.BytesIO(blob),
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
            },
        )

    # ─────────────────────────────────────────────────────────────────────
    # Forecast / capacity planning (issue #113).
    #
    # GET /api/forecast/{source_id}?horizon_days=180&model=linear
    #   Linear-regression projection over scan_runs history. Returns the
    #   ForecastResult dataclass shape.
    #
    # GET /api/forecast/{source_id}/export.xlsx
    #   Workbook with Summary / History / Settings sheets. Settings.B2 holds
    #   the editable threshold (% of disk).
    # ─────────────────────────────────────────────────────────────────────

    def _scan_history_for(source_id: int) -> list:
        """Pull (started_at, total_size) pairs from scan_runs for one source.

        Excludes still-running scans because their total_size is interim.
        Sorted ascending so forecast_growth() can use the first row as t0.
        """
        with db.get_cursor() as cur:
            cur.execute(
                """
                SELECT started_at, total_size, total_files
                FROM scan_runs
                WHERE source_id = ?
                  AND status = 'completed'
                  AND total_size IS NOT NULL
                  AND total_size > 0
                ORDER BY started_at ASC
                """,
                (int(source_id),),
            )
            return [dict(r) for r in cur.fetchall()]

    def _forecast_config() -> dict:
        cfg = (config or {}).get("forecast") or {}
        return {
            "enabled": bool(cfg.get("enabled", True)),
            "default_horizon_days": int(cfg.get("default_horizon_days", 180)),
            "capacity_threshold_pct": float(cfg.get("capacity_threshold_pct", 85)),
            "alarm_email": (cfg.get("alarm_email") or "").strip(),
            "alarm_lead_days": int(cfg.get("alarm_lead_days", 30)),
        }

    def _disk_total_bytes() -> int:
        """Best-effort: total bytes on the volume holding the SQLite DB.

        Same logic as /api/system/health — kept inline so the forecast
        endpoints don't have to round-trip through the health probe.
        """
        try:
            import shutil as _shutil
            target = os.path.dirname(os.path.abspath(db.db_path)) or "."
            return int(_shutil.disk_usage(target).total)
        except Exception:
            return 0

    @app.get("/api/forecast/{source_id}")
    async def forecast_endpoint(
        source_id: int,
        horizon_days: int = Query(180, ge=1, le=3650),
        model: str = Query("linear"),
        threshold_bytes: Optional[int] = Query(None, ge=0),
    ):
        """Capacity forecast for one source (issue #113).

        Currently only ``model=linear`` is supported; future iterations may
        add seasonal / ARIMA — keeping the query parameter so the URL doesn't
        break when that arrives.
        """
        from src.reports.forecast import forecast_growth

        fc_cfg = _forecast_config()
        if not fc_cfg["enabled"]:
            raise HTTPException(404, "forecast.enabled=false")

        if model != "linear":
            raise HTTPException(
                400, f"unsupported model {model!r} (only 'linear' for now)"
            )

        # Verify the source exists so callers get a real 404 (not an empty
        # forecast with samples_used=0).
        with db.get_cursor() as cur:
            cur.execute("SELECT 1 FROM sources WHERE id = ?", (source_id,))
            if not cur.fetchone():
                raise HTTPException(404, "source bulunamadi")

        rows = _scan_history_for(source_id)

        # Threshold resolution: explicit query param wins; otherwise fall
        # back to (capacity_threshold_pct% × disk_total_bytes).
        thr = threshold_bytes
        disk_total = _disk_total_bytes()
        if thr is None and disk_total > 0:
            thr = int(disk_total * fc_cfg["capacity_threshold_pct"] / 100.0)

        result = forecast_growth(
            rows,
            horizon_days=horizon_days,
            capacity_threshold_bytes=thr,
            source_id=source_id,
        )
        body = result.to_dict()
        # Surface the threshold + disk total so the frontend can recompute
        # alarm dates client-side without a second call.
        body["threshold_bytes"] = int(thr) if thr is not None else None
        body["disk_total_bytes"] = int(disk_total) if disk_total else None
        body["capacity_threshold_pct"] = fc_cfg["capacity_threshold_pct"]
        body["model"] = "linear"
        return body

    @app.get("/api/forecast/{source_id}/export.xlsx")
    async def forecast_export_xlsx(
        source_id: int,
        horizon_days: int = Query(180, ge=1, le=3650),
        threshold_bytes: Optional[int] = Query(None, ge=0),
        format: Optional[str] = Query(
            None,
            description="Set to 'csv' for a flat CSV of the History sheet (issue #122)",
        ),
    ):
        """XLSX workbook: Summary + History + Settings (editable threshold).

        Issue #122: ``?format=csv`` flattens the History sheet to a single
        streaming CSV with no row cap (Summary/Settings are tiny; CSV
        callers don't need them).
        """
        from fastapi.responses import StreamingResponse
        import io as _io
        from src.reports.forecast import forecast_growth
        from src.utils.xlsx_writer import stream_csv

        fc_cfg = _forecast_config()
        if not fc_cfg["enabled"]:
            raise HTTPException(404, "forecast.enabled=false")

        with db.get_cursor() as cur:
            cur.execute(
                "SELECT id, name FROM sources WHERE id = ?", (source_id,)
            )
            src_row = cur.fetchone()
            if not src_row:
                raise HTTPException(404, "source bulunamadi")
        source_name = src_row["name"]

        rows = _scan_history_for(source_id)
        disk_total = _disk_total_bytes()
        thr = threshold_bytes
        if thr is None and disk_total > 0:
            thr = int(disk_total * fc_cfg["capacity_threshold_pct"] / 100.0)

        result = forecast_growth(
            rows,
            horizon_days=horizon_days,
            capacity_threshold_bytes=thr,
            source_id=source_id,
        )

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        # ---- CSV fallback (issue #122) ------------------------------------
        if (format or "").lower() == "csv":
            csv_columns = [
                {"key": "started_at", "header": "started_at"},
                {"key": "total_size_bytes", "header": "total_size_bytes"},
                {"key": "total_files", "header": "total_files"},
            ]

            def _hist_rows():
                for r in rows:
                    yield {
                        "started_at": r.get("started_at") or "",
                        "total_size_bytes": int(r.get("total_size") or 0),
                        "total_files": int(r.get("total_files") or 0),
                    }
            return StreamingResponse(
                stream_csv(_hist_rows(), csv_columns),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition":
                        f"attachment; filename=Forecast_source{source_id}"
                        f"_h{horizon_days}d_{ts}.csv",
                    "X-Format-Fallback": "csv",
                },
            )

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError as e:
            raise HTTPException(
                500, f"openpyxl not available: {e}"
            )

        wb = Workbook()

        # ---- Summary sheet (active by default) ------------------------
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum["A1"] = "Field"
        ws_sum["B1"] = "Value"
        for c in (ws_sum["A1"], ws_sum["B1"]):
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
        rows_sum = [
            ("Source ID", source_id),
            ("Source name", source_name),
            ("Horizon (days)", result.horizon_days),
            ("Samples used", result.samples_used),
            ("R²", round(result.r_squared, 6)),
            ("Slope (bytes/day)", round(result.slope_bytes_per_day, 2)),
            ("Predicted bytes", int(result.predicted_bytes)),
            ("Predicted GiB",
             round(result.predicted_bytes / (1024 ** 3), 4)),
            ("CI low (bytes)", int(result.ci_low_bytes)),
            ("CI high (bytes)", int(result.ci_high_bytes)),
            ("Threshold (bytes)", int(thr) if thr is not None else ""),
            ("Disk total (bytes)", int(disk_total) if disk_total else ""),
            ("Capacity alarm at", result.capacity_alarm_at or ""),
            ("Generated at (UTC)", datetime.utcnow().isoformat() + "Z"),
        ]
        for i, (k, v) in enumerate(rows_sum, start=2):
            ws_sum.cell(row=i, column=1, value=k)
            ws_sum.cell(row=i, column=2, value=v)
        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 32

        # ---- History sheet (raw scan_runs data) -----------------------
        ws_hist = wb.create_sheet("History")
        ws_hist.append(["started_at", "total_size_bytes", "total_files"])
        for cell in ws_hist[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        for r in rows:
            ws_hist.append([
                r.get("started_at") or "",
                int(r.get("total_size") or 0),
                int(r.get("total_files") or 0),
            ])
        ws_hist.column_dimensions["A"].width = 24
        ws_hist.column_dimensions["B"].width = 22
        ws_hist.column_dimensions["C"].width = 16

        # ---- Settings sheet (editable threshold cell) -----------------
        ws_set = wb.create_sheet("Settings")
        ws_set["A1"] = "Setting"
        ws_set["B1"] = "Value"
        for c in (ws_set["A1"], ws_set["B1"]):
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
        # B2 is the editable threshold percentage. B3 holds the disk total
        # so auditors can compute (B2/100)*B3 manually if they need to.
        ws_set["A2"] = "Threshold (% of disk)"
        ws_set["B2"] = fc_cfg["capacity_threshold_pct"]
        ws_set["A3"] = "Disk total bytes"
        ws_set["B3"] = int(disk_total) if disk_total else ""
        ws_set["A4"] = "Threshold bytes (=B2/100*B3)"
        ws_set["B4"] = "=B2/100*B3" if disk_total else int(thr or 0)
        ws_set["A5"] = "Note"
        ws_set["B5"] = (
            "Edit B2 to change the threshold; B4 recomputes automatically. "
            "Re-run the forecast endpoint to refresh the alarm date."
        )
        ws_set.column_dimensions["A"].width = 32
        ws_set.column_dimensions["B"].width = 32

        buf = _io.BytesIO()
        wb.save(buf)
        blob = buf.getvalue()

        filename = (
            f"Forecast_source{source_id}_h{horizon_days}d_{ts}.xlsx"
        )
        return StreamingResponse(
            _io.BytesIO(blob),
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
            },
        )

    return app
