"""XLS ve PDF rapor export modulu.

openpyxl ile Excel, reportlab ile PDF rapor olusturur.
"""

import logging
from datetime import datetime
from io import BytesIO

logger = logging.getLogger("file_activity.report_exporter_v2")


class XLSExporter:
    """Excel (XLSX) rapor olusturucu."""

    def __init__(self, db, config):
        self.db = db
        self.config = config

    def export_full_report(self, source_id: int) -> bytes:
        """Tam analiz raporu olustur (tum sheet'ler)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from src.utils.size_formatter import format_size

        wb = Workbook()
        scan_id = self.db.get_latest_scan_id(source_id)

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        def style_header(ws, cols):
            for i, col_name in enumerate(cols, 1):
                cell = ws.cell(row=1, column=i, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

        def auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value else ""
                        max_len = max(max_len, len(val))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        # ---- Sheet 1: Ozet ----
        ws = wb.active
        ws.title = "Ozet"
        source = self.db.get_source_by_id(source_id)
        summary = self.db.get_status_summary(source_id, scan_id) if scan_id else {}

        ws["A1"] = "FILE ACTIVITY - Analiz Raporu"
        ws["A1"].font = Font(bold=True, size=16, color="1E40AF")
        ws["A3"] = "Kaynak:"
        ws["B3"] = source.name if source else str(source_id)
        ws["A4"] = "Yol:"
        ws["B4"] = source.unc_path if source else ""
        ws["A5"] = "Rapor Tarihi:"
        ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A7"] = "Toplam Dosya:"
        ws["B7"] = summary.get("total_files", 0)
        ws["A8"] = "Toplam Boyut:"
        ws["B8"] = format_size(summary.get("total_size", 0))
        ws["A9"] = "Dosya Turu Sayisi:"
        ws["B9"] = summary.get("type_count", 0)
        auto_width(ws)

        if not scan_id:
            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        # ---- Sheet 2: Erisim Sikligi ----
        ws2 = wb.create_sheet("Erisim Sikligi")
        freq_data = self.db.get_frequency_analysis(
            source_id, scan_id,
            self.config.get("analysis", {}).get("frequency_buckets", [30, 90, 180, 365, 730])
        )
        style_header(ws2, ["Kriter", "Dosya Sayisi", "Toplam Boyut", "Boyut (Formatli)"])
        for i, f in enumerate(freq_data, 2):
            ws2.cell(row=i, column=1, value=f["label"])
            ws2.cell(row=i, column=2, value=f["file_count"])
            ws2.cell(row=i, column=3, value=f["total_size"])
            ws2.cell(row=i, column=4, value=format_size(f["total_size"]))
        auto_width(ws2)

        # ---- Sheet 3: Dosya Turleri ----
        ws3 = wb.create_sheet("Dosya Turleri")
        type_data = self.db.get_type_analysis(source_id, scan_id)
        style_header(ws3, ["Uzanti", "Dosya Sayisi", "Toplam Boyut", "Boyut (Formatli)", "Ortalama", "Min", "Max"])
        for i, t in enumerate(type_data, 2):
            ws3.cell(row=i, column=1, value=t["extension"])
            ws3.cell(row=i, column=2, value=t["file_count"])
            ws3.cell(row=i, column=3, value=t["total_size"])
            ws3.cell(row=i, column=4, value=format_size(t["total_size"] or 0))
            ws3.cell(row=i, column=5, value=format_size(t.get("avg_size", 0) or 0))
            ws3.cell(row=i, column=6, value=format_size(t.get("min_size", 0) or 0))
            ws3.cell(row=i, column=7, value=format_size(t.get("max_size", 0) or 0))
        auto_width(ws3)

        # ---- Sheet 4: Boyut Dagilimi ----
        ws4 = wb.create_sheet("Boyut Dagilimi")
        size_buckets = self.config.get("analysis", {}).get("size_buckets", {
            "tiny": 102400, "small": 1048576, "medium": 10485760,
            "large": 104857600, "xlarge": 1073741824
        })
        size_data = self.db.get_size_analysis(source_id, scan_id, size_buckets)
        style_header(ws4, ["Kategori", "Min (bytes)", "Max (bytes)", "Dosya Sayisi", "Toplam Boyut", "Boyut (Formatli)"])
        for i, s in enumerate(size_data, 2):
            ws4.cell(row=i, column=1, value=s["label"])
            ws4.cell(row=i, column=2, value=s["min_bytes"])
            ws4.cell(row=i, column=3, value=s.get("max_bytes") or "")
            ws4.cell(row=i, column=4, value=s["file_count"])
            ws4.cell(row=i, column=5, value=s["total_size"])
            ws4.cell(row=i, column=6, value=format_size(s["total_size"] or 0))
        auto_width(ws4)

        # ---- Sheet 5: Sahiplik ----
        ws5 = wb.create_sheet("Sahiplik")
        owner_data = self.db.get_file_owners_stats(source_id, scan_id)
        style_header(ws5, ["Sahip", "Dosya Sayisi", "Toplam Boyut", "Boyut (Formatli)"])
        for i, o in enumerate(owner_data, 2):
            ws5.cell(row=i, column=1, value=o["owner"])
            ws5.cell(row=i, column=2, value=o["file_count"])
            ws5.cell(row=i, column=3, value=o["total_size"])
            ws5.cell(row=i, column=4, value=format_size(o["total_size"] or 0))
        auto_width(ws5)

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_drilldown(self, files: list, title: str) -> bytes:
        """Drill-down dosya listesini XLS olarak export et."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from src.utils.size_formatter import format_size

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name max 31 chars

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        cols = ["Dosya Adi", "Yol", "Uzanti", "Boyut", "Boyut (Formatli)",
                "Olusturma", "Son Erisim", "Son Degisiklik", "Sahip"]
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill

        for i, f in enumerate(files, 2):
            ws.cell(row=i, column=1, value=f.get("file_name", ""))
            ws.cell(row=i, column=2, value=f.get("file_path", ""))
            ws.cell(row=i, column=3, value=f.get("extension", ""))
            ws.cell(row=i, column=4, value=f.get("file_size", 0))
            ws.cell(row=i, column=5, value=format_size(f.get("file_size", 0)))
            ws.cell(row=i, column=6, value=f.get("creation_time", ""))
            ws.cell(row=i, column=7, value=f.get("last_access_time", ""))
            ws.cell(row=i, column=8, value=f.get("last_modify_time", ""))
            ws.cell(row=i, column=9, value=f.get("owner", ""))

        output = BytesIO()
        wb.save(output)
        return output.getvalue()


class PDFExporter:
    """PDF rapor olusturucu (reportlab)."""

    def __init__(self, db, config):
        self.db = db
        self.config = config

    def export_full_report(self, source_id: int) -> bytes:
        """Tam PDF rapor olustur."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from src.utils.size_formatter import format_size

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4,
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                      fontSize=20, spaceAfter=20,
                                      textColor=colors.HexColor("#1E40AF"))
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'],
                                        fontSize=14, spaceAfter=10, spaceBefore=20,
                                        textColor=colors.HexColor("#1E3A5F"))
        normal_style = styles['Normal']

        scan_id = self.db.get_latest_scan_id(source_id)
        source = self.db.get_source_by_id(source_id)
        summary = self.db.get_status_summary(source_id, scan_id) if scan_id else {}

        # Title
        elements.append(Paragraph("FILE ACTIVITY - Analiz Raporu", title_style))
        elements.append(Spacer(1, 10))

        # Source info
        info_data = [
            ["Kaynak:", source.name if source else str(source_id)],
            ["Yol:", source.unc_path if source else ""],
            ["Rapor Tarihi:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Toplam Dosya:", str(summary.get("total_files", 0))],
            ["Toplam Boyut:", format_size(summary.get("total_size", 0))],
            ["Dosya Turu:", str(summary.get("type_count", 0))],
        ]
        info_table = Table(info_data, colWidths=[120, 350])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor("#374151")),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))

        if not scan_id:
            elements.append(Paragraph("Tarama verisi bulunamadi.", normal_style))
            doc.build(elements)
            return output.getvalue()

        # --- Table style ---
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ])

        # --- Frequency ---
        elements.append(Paragraph("Erisim Sikligi", heading_style))
        freq_data = self.db.get_frequency_analysis(
            source_id, scan_id,
            self.config.get("analysis", {}).get("frequency_buckets", [30, 90, 180, 365, 730])
        )
        if freq_data:
            t_data = [["Kriter", "Dosya Sayisi", "Toplam Boyut"]]
            for f in freq_data:
                t_data.append([f["label"], str(f["file_count"]), format_size(f["total_size"])])
            t = Table(t_data, colWidths=[200, 100, 120])
            t.setStyle(table_style)
            elements.append(t)
        elements.append(Spacer(1, 15))

        # --- Types ---
        elements.append(Paragraph("Dosya Turleri", heading_style))
        type_data = self.db.get_type_analysis(source_id, scan_id)
        if type_data:
            t_data = [["Uzanti", "Sayi", "Toplam Boyut", "Ortalama"]]
            for tp in type_data[:30]:
                t_data.append([
                    "." + tp["extension"],
                    str(tp["file_count"]),
                    format_size(tp["total_size"] or 0),
                    format_size(tp.get("avg_size", 0) or 0),
                ])
            t = Table(t_data, colWidths=[100, 80, 120, 120])
            t.setStyle(table_style)
            elements.append(t)
        elements.append(Spacer(1, 15))

        # --- Size Distribution ---
        elements.append(Paragraph("Boyut Dagilimi", heading_style))
        size_buckets = self.config.get("analysis", {}).get("size_buckets", {
            "tiny": 102400, "small": 1048576, "medium": 10485760,
            "large": 104857600, "xlarge": 1073741824
        })
        size_data = self.db.get_size_analysis(source_id, scan_id, size_buckets)
        if size_data:
            t_data = [["Kategori", "Dosya Sayisi", "Toplam Boyut"]]
            for s in size_data:
                t_data.append([s["label"], str(s["file_count"]), format_size(s["total_size"] or 0)])
            t = Table(t_data, colWidths=[150, 100, 120])
            t.setStyle(table_style)
            elements.append(t)
        elements.append(Spacer(1, 15))

        # --- Owners ---
        elements.append(Paragraph("Dosya Sahipligi", heading_style))
        owner_data = self.db.get_file_owners_stats(source_id, scan_id)
        if owner_data:
            t_data = [["Sahip", "Dosya Sayisi", "Toplam Boyut"]]
            for o in owner_data[:30]:
                t_data.append([o["owner"], str(o["file_count"]), format_size(o["total_size"] or 0)])
            t = Table(t_data, colWidths=[200, 100, 120])
            t.setStyle(table_style)
            elements.append(t)

        doc.build(elements)
        return output.getvalue()
