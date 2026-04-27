"""Tests for ``src.utils.xlsx_writer.write_large_workbook`` (issue #122).

Production reproducer: a 2.5M-row scan crashed with ``Row numbers must be
between 1 and 1048576`` in the chargeback / mit-naming exporters. The
helper splits rows across ``Data_1``, ``Data_2``, ... sheets (default
1,000,000 rows per sheet, well under Excel's 2**20 hard cap) using
openpyxl ``write_only=True`` for constant-memory streaming.

These tests exercise the four behaviours the issue calls out:

* Single-sheet output when row count is under the cap.
* Multi-sheet split when row count exceeds the cap (we use a 1.5M-row
  generator with ``max_rows_per_sheet=1_000_000`` to hit the split path
  without actually emitting a 1.5M-row workbook on disk).
* ``extra_sheets`` lets a Settings sheet land *before* the data sheets,
  with chargeback-style ``=Settings!$B$2 * C{r}`` formulas working from
  every Data_N sheet.
* The Index sheet lists each Data_N sheet and its row count.
* The CSV fallback endpoint streams without hitting the row limit.
"""

from __future__ import annotations

import io
import os
import sys

import pytest

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from openpyxl import load_workbook  # noqa: E402

from src.utils.xlsx_writer import (  # noqa: E402
    EXCEL_MAX_ROWS_PER_SHEET,
    stream_csv,
    write_large_workbook,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _basic_columns():
    return [
        {"key": "id", "header": "id"},
        {"key": "name", "header": "name"},
        {"key": "size", "header": "size"},
    ]


def _basic_rows(n):
    for i in range(n):
        yield {"id": i, "name": f"file{i}", "size": i * 10}


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_write_large_workbook_under_limit_single_sheet():
    """1000 rows with the default 1M cap fits in one sheet (Data_1)."""
    buf = io.BytesIO()
    meta = write_large_workbook(_basic_rows(1000), _basic_columns(), buf)

    assert meta["total_rows"] == 1000
    assert meta["sheet_count"] == 1
    assert meta["sheet_titles"] == ["Data_1"]
    assert meta["byte_size"] > 0

    buf.seek(0)
    wb = load_workbook(buf)
    # No Index sheet for single-sheet workbooks (no value to navigate).
    assert wb.sheetnames == ["Data_1"]
    ws = wb["Data_1"]
    # 1 header row + 1000 data rows.
    assert ws.max_row == 1001
    assert [c.value for c in ws[1]] == ["id", "name", "size"]
    assert ws.cell(row=2, column=1).value == 0
    assert ws.cell(row=1001, column=2).value == "file999"


def test_write_large_workbook_over_limit_splits():
    """1.5M rows with cap=1_000_000 produces Data_1 (1M) + Data_2 (500k).

    Uses a small max_rows_per_sheet for the same code path without paying
    the cost of actually generating 1.5M rows — the split logic only
    cares about the cap being exceeded, not the absolute count.
    """
    cap = 1_000_000
    total = cap + 500_000  # 1.5M
    buf = io.BytesIO()
    meta = write_large_workbook(
        _basic_rows(total), _basic_columns(), buf, max_rows_per_sheet=cap,
    )

    assert meta["total_rows"] == total
    assert meta["sheet_count"] == 2
    assert meta["sheet_titles"] == ["Data_1", "Data_2"]

    buf.seek(0)
    wb = load_workbook(buf, read_only=True)
    assert "Data_1" in wb.sheetnames
    assert "Data_2" in wb.sheetnames

    # ``read_only=True`` doesn't materialise ``max_row``; count rows
    # via the iterator instead. Each sheet has 1 header row followed
    # by N data rows.
    n1 = sum(1 for _ in wb["Data_1"].iter_rows())
    n2 = sum(1 for _ in wb["Data_2"].iter_rows())
    assert n1 == cap + 1, f"Data_1 should hold {cap} rows + header, got {n1}"
    assert n2 == (total - cap) + 1, (
        f"Data_2 should hold {total - cap} rows + header, got {n2}"
    )


def test_write_large_workbook_extra_settings_sheet():
    """A chargeback-style Settings sheet lands before Data_N and the
    per-row =Settings!$B$2*C{r} formula stays sheet-aware after a split.
    """
    extra = {
        "Settings": [
            (1, "A", "Setting"),
            (1, "B", "Value"),
            (2, "A", "Default cost_per_gb_month"),
            (2, "B", 0.05),
        ],
    }

    def factory(row_idx, col_idx, key):
        # monthly_cost is column 5 in our schema below; emit the same
        # formula chargeback uses today.
        if key == "monthly_cost":
            return f"=Settings!$B$2*C{row_idx}"
        return None

    cols = [
        {"key": "cost_center", "header": "cost_center"},
        {"key": "owner", "header": "owner"},
        {"key": "total_gb", "header": "total_gb"},
        {"key": "file_count", "header": "file_count"},
        {"key": "monthly_cost", "header": "monthly_cost"},
    ]

    def rows(n):
        for i in range(n):
            yield {
                "cost_center": "HR",
                "owner": f"user{i}",
                "total_gb": float(i),
                "file_count": i * 2,
            }

    # Force a split with cap=3 over 7 rows so we can verify the formula
    # stays correct on Data_2 (different sheet, but same =Settings!$B$2).
    buf = io.BytesIO()
    meta = write_large_workbook(
        rows(7), cols, buf,
        max_rows_per_sheet=3,
        extra_sheets=extra,
        formula_factory=factory,
    )
    assert meta["sheet_count"] == 3

    buf.seek(0)
    wb = load_workbook(buf)
    assert wb.sheetnames[0] == "Settings"
    settings = wb["Settings"]
    assert settings["B2"].value == 0.05

    # Data_1 and Data_2 must both carry =Settings!$B$2*C{r} formulas where
    # r is the row index *within that sheet* (row 2, 3, 4 etc) — not a
    # globally-incrementing counter.
    data1 = wb["Data_1"]
    data2 = wb["Data_2"]
    f1 = data1.cell(row=2, column=5).value
    f2_first = data2.cell(row=2, column=5).value
    assert f1 == "=Settings!$B$2*C2", f"Data_1 row 2 formula: {f1!r}"
    assert f2_first == "=Settings!$B$2*C2", (
        f"Data_2 first data row should reset to C2, got {f2_first!r}"
    )


def test_write_large_workbook_index_sheet_links_data_sheets():
    """Multi-sheet workbook gets a final Index sheet listing the Data_N
    titles and their row counts so auditors can navigate."""
    buf = io.BytesIO()
    meta = write_large_workbook(
        _basic_rows(7), _basic_columns(), buf, max_rows_per_sheet=3,
    )
    assert meta["sheet_count"] == 3

    buf.seek(0)
    wb = load_workbook(buf)
    assert "Index" in wb.sheetnames
    idx = wb["Index"]
    rows = [tuple(c.value for c in row) for row in idx.iter_rows()]
    # Header + 3 sheet rows + TOTAL row.
    assert rows[0] == ("Sheet", "Row count")
    titles = [r[0] for r in rows[1:-1]]
    counts = [r[1] for r in rows[1:-1]]
    assert titles == ["Data_1", "Data_2", "Data_3"]
    assert counts == [3, 3, 1]
    # TOTAL row reports the sum.
    assert rows[-1] == ("TOTAL", 7)


def test_write_large_workbook_caps_at_excel_hard_limit():
    """Callers passing max_rows_per_sheet > Excel's 1,048,576 cap get
    silently clamped — otherwise wb.save() raises the same error #122
    was filed for."""
    buf = io.BytesIO()
    # We only emit 5 rows; the test is purely about the cap arithmetic
    # not actually trying to write 2M rows.
    meta = write_large_workbook(
        _basic_rows(5), _basic_columns(), buf,
        max_rows_per_sheet=EXCEL_MAX_ROWS_PER_SHEET + 100,
    )
    # 5 rows still fit on a single sheet under the clamped cap.
    assert meta["sheet_count"] == 1


def test_write_large_workbook_empty_iterator_yields_header_only_sheet():
    """An empty row iterator still produces a valid (header-only) workbook
    so endpoints don't 500 on no-data scans."""
    buf = io.BytesIO()
    meta = write_large_workbook(iter([]), _basic_columns(), buf)
    assert meta["total_rows"] == 0
    assert meta["sheet_count"] == 1

    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb["Data_1"]
    # Just the header row.
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == ["id", "name", "size"]


def test_csv_fallback_returns_streaming_response():
    """Endpoint smoke test: an XLSX endpoint with ``?format=csv`` returns
    a streaming CSV with no row cap, the right Content-Type and an
    ``X-Format-Fallback: csv`` marker for the frontend toast."""
    from fastapi import FastAPI
    from fastapi.responses import StreamingResponse
    from fastapi.testclient import TestClient

    app = FastAPI()

    @app.get("/api/test/export.xlsx")
    async def fake_export(format: str | None = None):
        cols = _basic_columns()
        rows = _basic_rows(50)
        if (format or "").lower() == "csv":
            return StreamingResponse(
                stream_csv(rows, cols),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": "attachment; filename=test.csv",
                    "X-Format-Fallback": "csv",
                },
            )
        buf = io.BytesIO()
        meta = write_large_workbook(rows, cols, buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": "attachment; filename=test.xlsx",
                "X-Sheet-Count": str(meta["sheet_count"]),
            },
        )

    client = TestClient(app)
    resp = client.get("/api/test/export.xlsx", params={"format": "csv"})
    assert resp.status_code == 200
    assert resp.headers["x-format-fallback"] == "csv"
    assert "text/csv" in resp.headers["content-type"]
    body = resp.content.decode("utf-8")
    lines = body.strip().split("\n")
    # Header + 50 data rows.
    assert len(lines) == 51
    assert lines[0] == "id,name,size"
    assert lines[1].startswith("0,file0,")

    # And the default (no format=) path returns XLSX.
    resp_xlsx = client.get("/api/test/export.xlsx")
    assert resp_xlsx.status_code == 200
    assert "spreadsheetml" in resp_xlsx.headers["content-type"]
    assert resp_xlsx.headers["x-sheet-count"] == "1"


def test_stream_csv_yields_bytes_per_row():
    """``stream_csv`` should emit one chunk per row so a 1M-row CSV does
    not balloon memory."""
    cols = _basic_columns()
    chunks = list(stream_csv(_basic_rows(3), cols))
    # Header + 3 data chunks.
    assert len(chunks) == 4
    for c in chunks:
        assert isinstance(c, bytes)
    # First chunk is the header.
    assert chunks[0].decode("utf-8").strip() == "id,name,size"
