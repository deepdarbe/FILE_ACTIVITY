"""Tests for the chargeback / cost-center report (issue #111).

Covers:

* compute() with no centers -> all owners go to ``unmapped_owners``
* compute() with an exact owner match
* compute() with an fnmatch glob (``CONTOSO\\hr_*``) matching multiple
* export_xlsx() emits FORMULAS (key cells start with ``=``)
* export_xlsx() rate cell on the Settings sheet is referenced by Detail
  formulas (so editing the rate updates totals)
* API smoke tests for the CRUD endpoints
"""

from __future__ import annotations

import io
import os
import sys

import pytest
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import StreamingResponse
from fastapi.testclient import TestClient

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from openpyxl import load_workbook  # noqa: E402

from src.reports.chargeback import (  # noqa: E402
    UNMAPPED_BUCKET,
    ChargebackReport,
)
from src.storage.database import Database  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _seed_db(tmp_path):
    """Create a DB with a source, a completed scan, and 5 files spanning
    a few owner shapes::

        CONTOSO\\jdoe         (HR director — exact-match candidate)
        CONTOSO\\hr_alice     (glob CONTOSO\\hr_* candidate)
        CONTOSO\\hr_bob       (glob CONTOSO\\hr_* candidate)
        FABRIKAM\\bcoder      (engineering)
        (empty owner)         (orphan / unattributed)

    Returns ``(db, source_id, scan_id)``.
    """
    db = Database({"path": str(tmp_path / "cb.db")})
    db.connect()
    with db.get_cursor() as cur:
        cur.execute(
            "INSERT INTO sources (name, unc_path) VALUES ('s1', '/share')"
        )
        source_id = cur.lastrowid
        cur.execute(
            "INSERT INTO scan_runs (source_id, status, completed_at) "
            "VALUES (?, 'completed', datetime('now'))",
            (source_id,),
        )
        scan_id = cur.lastrowid

        rows = [
            # 2 GiB exactly — easy to assert on
            (source_id, scan_id, "/share/hr/handbook.pdf", "hr/handbook.pdf",
             "handbook.pdf", "pdf", 2 * (1024 ** 3), "CONTOSO\\jdoe"),
            (source_id, scan_id, "/share/hr/payroll.xlsx", "hr/payroll.xlsx",
             "payroll.xlsx", "xlsx", 1 * (1024 ** 3), "CONTOSO\\hr_alice"),
            (source_id, scan_id, "/share/hr/staff.csv", "hr/staff.csv",
             "staff.csv", "csv", 512 * (1024 ** 2), "CONTOSO\\hr_bob"),
            (source_id, scan_id, "/share/eng/build.zip", "eng/build.zip",
             "build.zip", "zip", 4 * (1024 ** 3), "FABRIKAM\\bcoder"),
            (source_id, scan_id, "/share/misc/orphan.bin", "misc/orphan.bin",
             "orphan.bin", "bin", 100 * (1024 ** 2), ""),
        ]
        cur.executemany(
            """INSERT INTO scanned_files
               (source_id, scan_id, file_path, relative_path, file_name,
                extension, file_size, owner)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            rows,
        )
    return db, source_id, scan_id


# ---------------------------------------------------------------------------
# compute() tests
# ---------------------------------------------------------------------------


def test_compute_with_no_centers_unmapped(tmp_path):
    db, _, scan_id = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})
    result = cb.compute(scan_id)

    # No centers -> empty list
    assert result.centers == []
    # Every owner that appeared in a file is in the unmapped bucket. Empty
    # owners collapse into the synthetic "(no owner)" key.
    unmapped = {u["owner"] for u in result.unmapped_owners}
    assert "CONTOSO\\jdoe" in unmapped
    assert "CONTOSO\\hr_alice" in unmapped
    assert "CONTOSO\\hr_bob" in unmapped
    assert "FABRIKAM\\bcoder" in unmapped
    assert "(no owner)" in unmapped

    assert result.total_file_count == 5
    assert result.total_bytes > 0
    # Totals on the result match the dataclass property
    assert result.total_monthly_cost == 0.0


def test_compute_with_exact_owner_match(tmp_path):
    db, _, scan_id = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})

    cid = cb.add_center("HR Director", "Top of HR org", 0.10)
    cb.add_owner(cid, "CONTOSO\\jdoe")

    result = cb.compute(scan_id)
    assert len(result.centers) == 1
    hr = result.centers[0]
    # 2 GiB, 1 file
    assert hr.file_count == 1
    assert hr.total_bytes == 2 * (1024 ** 3)
    assert hr.total_gb == pytest.approx(2.0, rel=1e-9)
    # rate × gb -> 0.10 × 2 GiB = 0.20
    assert hr.monthly_cost == pytest.approx(0.20, rel=1e-9)
    # The other 4 files (4 owners) end up unmapped.
    unmapped_owners = {u["owner"] for u in result.unmapped_owners}
    assert "CONTOSO\\jdoe" not in unmapped_owners
    assert "CONTOSO\\hr_alice" in unmapped_owners
    assert "FABRIKAM\\bcoder" in unmapped_owners


def test_compute_with_glob_owner_match(tmp_path):
    db, _, scan_id = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})

    cid = cb.add_center("HR", "Human Resources", 0.05)
    cb.add_owner(cid, "CONTOSO\\hr_*")

    result = cb.compute(scan_id)
    assert len(result.centers) == 1
    hr = result.centers[0]
    # hr_alice (1 GiB) + hr_bob (0.5 GiB) = 1.5 GiB, 2 files
    assert hr.file_count == 2
    assert hr.total_bytes == (1 * (1024 ** 3)) + (512 * (1024 ** 2))
    assert hr.total_gb == pytest.approx(1.5, rel=1e-9)
    # Both hr_* owners appear in top_owners (capped at 10 — we have 2)
    owner_names = {o["owner"] for o in hr.top_owners}
    assert owner_names == {"CONTOSO\\hr_alice", "CONTOSO\\hr_bob"}


def test_compute_first_pattern_wins_on_overlap(tmp_path):
    """When two centers both could match, the deterministic ordering (by
    center name asc) ensures a stable first-pattern-wins outcome."""
    db, _, scan_id = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})

    a = cb.add_center("A_first", "alpha", 0.01)
    b = cb.add_center("Z_last", "omega", 0.99)
    cb.add_owner(a, "CONTOSO\\hr_*")
    cb.add_owner(b, "CONTOSO\\hr_alice")  # would also match alice

    result = cb.compute(scan_id)
    centers_by_name = {c.name: c for c in result.centers}
    # A_first (sorted before Z_last) absorbs both hr files; Z_last gets nothing.
    assert centers_by_name["A_first"].file_count == 2
    assert centers_by_name["Z_last"].file_count == 0


# ---------------------------------------------------------------------------
# CRUD invariants
# ---------------------------------------------------------------------------


def test_add_remove_center_idempotent(tmp_path):
    db, _, _ = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})
    cid = cb.add_center("X", "", 1.5)
    assert cb.remove_center(cid) is True
    # Second remove is a no-op (False), not an exception
    assert cb.remove_center(cid) is False


def test_add_owner_validates_center_exists(tmp_path):
    db, _, _ = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})
    with pytest.raises(ValueError):
        cb.add_owner(9999, "CONTOSO\\nope")


def test_add_owner_idempotent_on_duplicate(tmp_path):
    db, _, _ = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})
    cid = cb.add_center("HR", "", 0.1)
    assert cb.add_owner(cid, "CONTOSO\\jdoe") is True
    # Re-add same pattern -> False (no-op), no exception
    assert cb.add_owner(cid, "CONTOSO\\jdoe") is False


# ---------------------------------------------------------------------------
# XLSX export tests
# ---------------------------------------------------------------------------


def _build_export_with_one_center(tmp_path):
    db, _, scan_id = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})
    cid = cb.add_center("HR", "Human Resources", 0.05)
    cb.add_owner(cid, "CONTOSO\\hr_*")
    blob = cb.export_xlsx(scan_id)
    return blob, cb, scan_id


def test_xlsx_uses_formulas(tmp_path):
    blob, _, _ = _build_export_with_one_center(tmp_path)
    wb = load_workbook(io.BytesIO(blob))

    # Required sheets
    assert "Summary" in wb.sheetnames
    assert "Detail" in wb.sheetnames
    assert "Settings" in wb.sheetnames

    summary = wb["Summary"]
    # Summary row 2 column F (monthly_cost) is a formula referencing C2*D2
    f_cell = summary.cell(row=2, column=6).value
    assert isinstance(f_cell, str) and f_cell.startswith("="), (
        f"monthly_cost should be a formula, got {f_cell!r}"
    )
    assert "C2" in f_cell and "D2" in f_cell

    # Total row's monthly_cost is a SUM formula
    last_row = summary.max_row
    sum_cell = summary.cell(row=last_row, column=6).value
    assert isinstance(sum_cell, str) and sum_cell.startswith("=SUM("), (
        f"total monthly_cost should be a SUM formula, got {sum_cell!r}"
    )


def test_xlsx_rate_cell_referenced(tmp_path):
    """Settings!B2 holds the rate, and Detail-sheet formulas reference it
    so editing B2 updates the rendered totals."""
    blob, _, _ = _build_export_with_one_center(tmp_path)
    wb = load_workbook(io.BytesIO(blob))

    settings = wb["Settings"]
    # B2 must be a numeric value (the editable rate cell)
    assert isinstance(settings["B2"].value, (int, float)), (
        f"Settings!B2 should hold a numeric rate, got {settings['B2'].value!r}"
    )

    # At least one Detail formula must reference Settings!$B$2
    detail = wb["Detail"]
    found = False
    for row in detail.iter_rows(min_row=2):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "Settings!$B$2" in v and v.startswith("="):
                found = True
                break
        if found:
            break
    assert found, "No Detail-sheet formula references Settings!$B$2"


def test_xlsx_summary_unmapped_excluded_from_centers(tmp_path):
    """Unmapped owners must not pollute the Summary sheet's center rows.
    They live on the Detail sheet under the (unmapped) bucket only."""
    blob, _, _ = _build_export_with_one_center(tmp_path)
    wb = load_workbook(io.BytesIO(blob))

    summary = wb["Summary"]
    # Skip header row + 1 center row + 1 total row -> max_row 3
    names = [summary.cell(row=r, column=1).value for r in range(2, summary.max_row + 1)]
    assert "HR" in names
    assert UNMAPPED_BUCKET not in names

    detail = wb["Detail"]
    detail_first_col = [
        detail.cell(row=r, column=1).value
        for r in range(2, detail.max_row + 1)
    ]
    assert UNMAPPED_BUCKET in detail_first_col


def test_xlsx_with_zero_rate_falls_back_to_settings(tmp_path):
    """A center with rate=0 should reference Settings!$B$2 instead of
    hardcoding 0, so auditors can flip on a global rate."""
    db, _, scan_id = _seed_db(tmp_path)
    cb = ChargebackReport(db, {})
    cid = cb.add_center("Free", "", 0)
    cb.add_owner(cid, "CONTOSO\\hr_*")
    blob = cb.export_xlsx(scan_id)

    wb = load_workbook(io.BytesIO(blob))
    summary = wb["Summary"]
    rate_cell = summary.cell(row=2, column=3).value
    assert isinstance(rate_cell, str) and rate_cell == "=Settings!$B$2", (
        f"zero-rate center should reference Settings!$B$2, got {rate_cell!r}"
    )


# ---------------------------------------------------------------------------
# API smoke tests
# ---------------------------------------------------------------------------


def _build_api_app(db):
    """Mount the chargeback CRUD + compute endpoints onto a minimal app.
    Mirrors the production handlers verbatim — the production handlers
    are exercised via the ``create_app`` factory in dashboard tests; this
    smoke harness avoids the heavy fixture setup."""
    from typing import Optional as _Opt
    from src.reports.chargeback import ChargebackReport as _CB

    app = FastAPI()

    def _cb():
        return _CB(db, {})

    @app.get("/api/chargeback/centers")
    async def list_centers():
        return {"centers": _cb().list_centers()}

    @app.post("/api/chargeback/centers")
    async def add_center(body: dict):
        try:
            cid = _cb().add_center(
                name=(body.get("name") or "").strip(),
                description=body.get("description") or "",
                cost_per_gb_month=body.get("cost_per_gb_month") or 0,
            )
        except ValueError as e:
            raise HTTPException(400, str(e))
        return {"id": cid, "ok": True}

    @app.put("/api/chargeback/centers/{cid}")
    async def update_center(cid: int, body: dict):
        allowed = {"name", "description", "cost_per_gb_month"}
        fields = {k: v for k, v in (body or {}).items() if k in allowed}
        try:
            ok = _cb().update_center(cid, **fields)
        except ValueError as e:
            raise HTTPException(400, str(e))
        return {"ok": True, "updated": ok}

    @app.delete("/api/chargeback/centers/{cid}")
    async def del_center(cid: int):
        return {"ok": True, "deleted": _cb().remove_center(cid)}

    @app.post("/api/chargeback/centers/{cid}/owners")
    async def add_owner(cid: int, body: dict):
        try:
            added = _cb().add_owner(cid, (body.get("owner_pattern") or "").strip())
        except ValueError as e:
            raise HTTPException(400, str(e))
        return {"ok": True, "added": added}

    @app.delete("/api/chargeback/centers/{cid}/owners/{owner_pattern:path}")
    async def del_owner(cid: int, owner_pattern: str):
        return {"ok": True, "deleted": _cb().remove_owner(cid, owner_pattern)}

    @app.get("/api/chargeback/{source_id}")
    async def compute(source_id: int):
        scan_id = db.get_latest_scan_id(source_id, include_running=False)
        if not scan_id:
            raise HTTPException(404, "Tamamlanmis scan yok")
        return _cb().compute(scan_id).to_dict()

    return app


def test_api_crud_roundtrip(tmp_path):
    db, source_id, _ = _seed_db(tmp_path)
    client = TestClient(_build_api_app(db))

    # Initially empty
    r = client.get("/api/chargeback/centers")
    assert r.status_code == 200
    assert r.json()["centers"] == []

    # Add center
    r = client.post(
        "/api/chargeback/centers",
        json={"name": "HR", "description": "Human Resources", "cost_per_gb_month": 0.07},
    )
    assert r.status_code == 200
    cid = r.json()["id"]
    assert isinstance(cid, int) and cid > 0

    # Update rate
    r = client.put(f"/api/chargeback/centers/{cid}", json={"cost_per_gb_month": 0.09})
    assert r.status_code == 200

    # Add owner pattern (glob)
    r = client.post(
        f"/api/chargeback/centers/{cid}/owners",
        json={"owner_pattern": "CONTOSO\\hr_*"},
    )
    assert r.status_code == 200
    assert r.json()["added"] is True

    # Idempotent re-add
    r = client.post(
        f"/api/chargeback/centers/{cid}/owners",
        json={"owner_pattern": "CONTOSO\\hr_*"},
    )
    assert r.status_code == 200
    assert r.json()["added"] is False

    # Compute returns the HR center
    r = client.get(f"/api/chargeback/{source_id}")
    assert r.status_code == 200
    body = r.json()
    assert len(body["centers"]) == 1
    assert body["centers"][0]["name"] == "HR"
    assert body["centers"][0]["file_count"] == 2

    # Delete owner pattern (idempotent)
    r = client.delete(
        f"/api/chargeback/centers/{cid}/owners/{'CONTOSO%5Chr_%2A'}"
    )
    # path-param decode handles the URL-encoded pattern
    assert r.status_code == 200

    # Delete center (idempotent)
    r = client.delete(f"/api/chargeback/centers/{cid}")
    assert r.status_code == 200
    assert r.json()["deleted"] is True

    r = client.delete(f"/api/chargeback/centers/{cid}")
    assert r.status_code == 200
    assert r.json()["deleted"] is False  # already gone


def test_api_compute_404_when_no_completed_scan(tmp_path):
    """Mirror the dashboard contract: a source with only a running scan
    (no completed_at) should NOT yield a chargeback report."""
    db = Database({"path": str(tmp_path / "empty.db")})
    db.connect()
    with db.get_cursor() as cur:
        cur.execute(
            "INSERT INTO sources (name, unc_path) VALUES ('s1', '/share')"
        )
        source_id = cur.lastrowid
        cur.execute(
            "INSERT INTO scan_runs (source_id, status) VALUES (?, 'running')",
            (source_id,),
        )

    client = TestClient(_build_api_app(db))
    r = client.get(f"/api/chargeback/{source_id}")
    assert r.status_code == 404
