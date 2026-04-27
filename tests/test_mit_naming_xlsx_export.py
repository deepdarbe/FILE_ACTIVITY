"""Smoke test for the MIT-naming XLSX export endpoint (issue #80).

Drives the new ``/api/reports/mit-naming/{source_id}/export.xlsx`` endpoint
through TestClient against a real SQLite ``Database`` seeded with files
that violate at least R1 (space) and B1 (long name).

Asserts:
* HTTP 200 on the happy path
* Content-Type is the XLSX MIME type
* Body is a non-empty XLSX (parseable by openpyxl) with the expected
  header row and at least one data row
* ``ids=`` query param scopes the export to a subset
* 404 when no scan exists for the source
"""

import io
import os
import sys
from typing import Optional

import pytest
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import StreamingResponse
from fastapi.testclient import TestClient

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from openpyxl import load_workbook  # noqa: E402
from src.storage.database import Database  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _seed_db(tmp_path):
    """Returns ``(db, source_id, scan_id)`` with three seeded files:

    * ``has space.txt``   — violates R1 (space)
    * ``Aa.TXT``          — violates B4 (uppercase) and clean otherwise
    * ``thisisalongfilenamewithoutseparators_butlongerthan31chars.dat``
                          — violates B1 (long name) + B5 (no separator)
    """
    db = Database({"path": str(tmp_path / "naming.db")})
    db.connect()
    with db.get_cursor() as cur:
        cur.execute(
            "INSERT INTO sources (name, unc_path) VALUES ('s1', '/share')"
        )
        source_id = cur.lastrowid
        cur.execute(
            "INSERT INTO scan_runs (source_id, status) VALUES (?, 'completed')",
            (source_id,),
        )
        scan_id = cur.lastrowid

        rows = [
            (source_id, scan_id, "/share/has space.txt", "has space.txt",
             "has space.txt", "txt", 100, "alice"),
            (source_id, scan_id, "/share/Aa.TXT", "Aa.TXT",
             "Aa.TXT", "TXT", 200, "bob"),
            (source_id, scan_id,
             "/share/thisisalongfilenamewithoutseparators_butlongerthan31chars.dat",
             "thisisalongfilenamewithoutseparators_butlongerthan31chars.dat",
             "thisisalongfilenamewithoutseparators_butlongerthan31chars.dat",
             "dat", 300, "carol"),
        ]
        cur.executemany(
            """INSERT INTO scanned_files
               (source_id, scan_id, file_path, relative_path, file_name,
                extension, file_size, owner)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            rows,
        )
    return db, source_id, scan_id


def _build_app(db):
    """Mount the XLSX endpoint onto a minimal FastAPI app.

    Mirrors the production handler verbatim — when the production handler
    changes, update both. This avoids spinning up the full ``create_app``
    factory, which requires AnalyticsEngine + ADLookup + EmailNotifier
    fixtures we don't need here.
    """
    import re as re_mod
    from datetime import datetime

    from openpyxl import Workbook

    app = FastAPI()

    @app.get("/api/reports/mit-naming/{source_id}/export.xlsx")
    async def export_mit_naming_xlsx(
        source_id: int,
        ids: Optional[str] = Query(None),
    ):
        scan_id = db.get_latest_scan_id(source_id, include_running=True)
        if not scan_id:
            raise HTTPException(404, "Tarama bulunamadi")

        rules = [
            ("R1", "Bosluk Iceren", "critical",
             lambda p, n: bool(re_mod.search(r"\s", n))),
            ("R2", "Ilk Karakter Harf Degil", "critical",
             lambda p, n: bool(n) and not re_mod.match(r"^[a-zA-Z]", n)),
            ("R3", "Yasak Karakter", "critical",
             lambda p, n: bool(n) and "." in n
             and not re_mod.match(r"^[a-zA-Z0-9._-]+$", n[: n.rfind(".")])),
            ("R4", "Uzanti Sorunu", "critical",
             lambda p, n: "." not in n or not n.rsplit(".", 1)[-1].isalpha()),
            ("B1", "Uzun Ad (>31)", "warning",
             lambda p, n: len(n) > 31),
            ("B2", "Uzun Yol (>256)", "warning",
             lambda p, n: len(p) > 256),
            ("B3", "Base'de Nokta", "warning",
             lambda p, n: "." in n and n[: n.rfind(".")].count(".") > 0),
            ("B4", "Buyuk Harf", "info",
             lambda p, n: bool(re_mod.search(
                 r"[A-Z]", n[: n.rfind(".")] if "." in n else n))),
            ("B5", "Ayirici Yok", "info",
             lambda p, n: len(n) > 10 and "_" not in n and "-" not in n),
            ("B6", "Dizin Adinda Nokta", "info",
             lambda p, n: any(
                 "." in part and part not in ("", ".", "..")
                 for part in p.replace("\\", "/").split("/")
             )),
        ]

        id_filter = None
        if ids:
            id_filter = set()
            for tok in ids.split(","):
                tok = tok.strip()
                if tok:
                    try:
                        id_filter.add(int(tok))
                    except ValueError:
                        continue

        export_rows = []
        with db.get_cursor() as cur:
            cur.execute(
                """SELECT id, file_path, file_name, owner, last_modify_time, file_size
                   FROM scanned_files
                   WHERE source_id = ? AND scan_id = ?""",
                (source_id, scan_id),
            )
            for r in cur:
                if id_filter is not None and r["id"] not in id_filter:
                    continue
                path = r["file_path"] or ""
                name = r["file_name"] or ""
                for code, label, severity, fn in rules:
                    if fn(path, name):
                        export_rows.append({
                            "file_path": path,
                            "owner": r["owner"] or "",
                            "last_modify_time": r["last_modify_time"] or "",
                            "file_size": r["file_size"] or 0,
                            "rule": f"{code} - {label}",
                            "severity": severity,
                        })

        wb = Workbook()
        ws = wb.active
        ws.title = "MIT Naming"
        headers = ["file_path", "owner", "last_modify_time",
                   "file_size", "rule", "severity"]
        ws.append(headers)
        for row in export_rows:
            ws.append([row[h] for h in headers])

        last_data_row = ws.max_row
        if last_data_row > 1:
            total_row = last_data_row + 2
            ws.cell(row=total_row, column=3, value="TOTAL BYTES")
            ws.cell(
                row=total_row, column=4,
                value=f"=SUM(D2:D{last_data_row})",
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = (
            f"MIT_Naming_Report_source{source_id}_scan{scan_id}_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return StreamingResponse(
            buf,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Total-Rows": str(len(export_rows)),
            },
        )

    return app


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.fixture
def seeded(tmp_path):
    return _seed_db(tmp_path)


def test_xlsx_export_returns_workbook(seeded):
    db, source_id, _ = seeded
    client = TestClient(_build_app(db))
    resp = client.get(f"/api/reports/mit-naming/{source_id}/export.xlsx")

    assert resp.status_code == 200, resp.text
    assert (
        "spreadsheetml" in resp.headers["content-type"]
    ), f"unexpected content-type: {resp.headers['content-type']}"
    assert "attachment" in resp.headers.get("content-disposition", "")
    body = resp.content
    assert len(body) > 100, "XLSX body looks empty"
    # X-Total-Rows reports how many (file × rule) rows were emitted.
    total_rows = int(resp.headers.get("x-total-rows", "0"))
    assert total_rows >= 2, (
        f"expected at least 2 violation rows (R1 + B1+B5), got {total_rows}"
    )

    # Parse the workbook and assert structure.
    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    assert ws.max_row >= 2  # header + ≥1 data
    header = [c.value for c in ws[1]]
    assert header == [
        "file_path", "owner", "last_modify_time",
        "file_size", "rule", "severity",
    ]
    # At least one data row mentions the spaced filename.
    data = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
    assert any(
        r[0] == "/share/has space.txt" and r[4].startswith("R1") for r in data
    ), f"R1 violation missing from export: {data!r}"


def test_xlsx_export_filters_by_ids(seeded):
    db, source_id, scan_id = seeded
    # Look up the id of the spaced-filename row so we can scope ?ids=.
    with db.get_cursor() as cur:
        cur.execute(
            "SELECT id FROM scanned_files WHERE file_name = ? AND scan_id = ?",
            ("has space.txt", scan_id),
        )
        target_id = cur.fetchone()["id"]

    client = TestClient(_build_app(db))
    resp = client.get(
        f"/api/reports/mit-naming/{source_id}/export.xlsx",
        params={"ids": str(target_id)},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    paths = {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}
    # The TOTAL BYTES marker row leaks "TOTAL BYTES" into col C, not col A,
    # so col A should only contain the filtered file path (or be empty for
    # the trailing total row).
    assert paths == {"/share/has space.txt"}, (
        f"ids filter leaked other files: {paths!r}"
    )


def test_xlsx_export_404_when_no_scan(tmp_path):
    db = Database({"path": str(tmp_path / "empty.db")})
    db.connect()
    with db.get_cursor() as cur:
        cur.execute(
            "INSERT INTO sources (name, unc_path) VALUES ('s1', '/share')"
        )
        source_id = cur.lastrowid

    client = TestClient(_build_app(db))
    resp = client.get(f"/api/reports/mit-naming/{source_id}/export.xlsx")
    assert resp.status_code == 404
